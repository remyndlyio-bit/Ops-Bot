"""
Run the full Intent_Test_Matrix (148 rows) against the live bot and write a
results workbook.

    python -m tests.e2e.run_matrix                 # everything
    python -m tests.e2e.run_matrix --category "Invoice Generation"
    python -m tests.e2e.run_matrix --language EN   # the 134 English rows
    python -m tests.e2e.run_matrix --limit 5       # smoke test the harness

Needs a real AI_KEY and SUPABASE_DB_URL. It makes real (paid) AI calls — two
per row, one for the bot and one for the judge — and writes to a real
database, always under a synthetic `e2etest:` user id that is torn down at
the end.

Why a script and not a pytest module
------------------------------------
pytest wants independent test functions; these rows are 13 ordered
conversations. Forcing them into parametrised tests would either lose the
ordering guarantee or smuggle shared mutable state through fixtures, and a
run of this shape is a REPORT (a graded spreadsheet a human reads) rather
than a pass/fail gate. The 29-row regression suite in test_scenarios.py is
the gate; this is the survey. They deliberately serve different purposes.

State discipline: memory and rows are reset at every category boundary, so a
mid-sheet failure can't cascade into the next category. Within a category
state is deliberately preserved — that's what makes multi-turn rows real.
"""

from __future__ import annotations

import argparse
import os
import sys
import traceback
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from tests.e2e import grade as grade_mod
from tests.e2e import matrix as matrix_mod
from tests.e2e import seed as seed_mod


def _run_config_fingerprint() -> list:
    """(label, value) pairs identifying which dispatch config this run
    actually exercised. P2-4 (PLAN_OF_ACTION.md): a prior run's 62.5%
    baseline was graded with FLOW_MACHINE_V2 unset (legacy path) while
    production's default was assumed to be v2-on — nothing in the workbook
    itself recorded which one actually ran, so the number was silently
    ambiguous. Every run's workbook now carries this fingerprint."""
    v2_raw = (os.getenv("FLOW_MACHINE_V2", "") or "").strip()
    v2_display = v2_raw if v2_raw else "(unset — legacy dispatch for every user)"
    try:
        import subprocess
        commit = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True, text=True, timeout=5,
            cwd=os.path.dirname(os.path.abspath(__file__)),
        ).stdout.strip() or "(unknown)"
    except Exception:
        commit = "(unknown)"
    return [
        ("Run timestamp", datetime.now().isoformat(timespec="seconds")),
        ("FLOW_MACHINE_V2", v2_display),
        ("STRICT_PLAN_VALIDATION", os.getenv("STRICT_PLAN_VALIDATION", "1 (default)")),
        ("KNOWLEDGE_BOOK", os.getenv("KNOWLEDGE_BOOK", "0 (default, off)")),
        ("Git commit", commit),
    ]


def _reset_memory(svc, uid: str) -> None:
    try:
        from utils.memory_service import _DEFAULT
        svc.memory._write_raw(uid, dict(_DEFAULT))
    except Exception:
        svc.memory.update_user_memory(uid, {
            "last_intent": None, "uscf_context": {}, "answer_ledger": [],
            "flow_v2": None, "last_generated_invoice": None,
            "pending_disambiguation": None, "suggested_next_action": None,
        })


def _fresh_unonboarded_uid() -> str:
    """A synthetic id with NO profile row, so the Onboarding rows exercise
    the real first-run path. Deliberately not seeded — seeding sets
    onboarded_at, which is precisely what those 8 rows are testing."""
    return seed_mod.new_user_id()


def run(rows, service, gemini, sink=None):
    """Execute every row and grade it.

    Isolation model, driven by the sheet's own annotations (see matrix.py):

    * `(after …)` rows CONTINUE from the previous row — no reset, because
      the previous turn's state is the whole point of the test.
    * every other row starts from a clean account: rows re-seeded, memory
      wiped. A row must never be able to observe what an earlier row did.
    * Onboarding rows get a fresh, UN-seeded id (no profile) plus an
      explicit setup path that drives the conversation to the step under
      test.
    * `(check PDF)` rows are recorded MANUAL without being sent — they ask
      a human to open an artifact.
    """
    results = sink if sink is not None else []
    current_category = None
    uid = None
    seeded_uid = None
    # P2-4 (PLAN_OF_ACTION.md): precondition rows ("no bank saved", "no
    # billing info", ...) get their OWN dedicated fixture user instead of
    # the shared one — the shared fixture always has bank details and full
    # Nike billing/POC data by design, so those preconditions could never
    # actually hold against it. Tracked separately from seeded_uid so a
    # variant seed never clobbers the shared fixture other rows depend on.
    variant_uids: list = []

    for row in rows:
        if row.category != current_category:
            current_category = row.category
            print(f"\n─── {current_category} ───", flush=True)

        if row.is_manual:
            results.append({
                "num": row.num, "category": row.category, "scenario": row.scenario,
                "message": row.message, "expected": row.expected,
                "language": row.language,
                "actual": "", "operation": "", "verdict": "MANUAL",
                "reason": (
                    "requires inspecting a generated PDF/email — not gradeable "
                    "from the chat reply"
                    if row.send_text.strip() else
                    f"not a chat message ({row.annotation or 'no message'}) — "
                    "a background job, button tap, or platform delivery detail; "
                    "unreachable through process_request"
                ),
            })
            print(f"  · {row.num:>3} {row.scenario[:44]:<44} MANUAL", flush=True)
            continue

        setup_trace = []
        if row.continues_previous and uid is not None:
            pass                       # keep the previous row's state
        elif row.needs_fresh_account:
            uid = _fresh_unonboarded_uid()
            for msg in row.setup_messages:
                service.process_request(uid, msg)
                setup_trace.append(msg)
        elif row.precondition_variant:
            # A fresh variant user EVERY time (not reused across rows) — a
            # row asserting "no bank saved" must never accidentally run
            # against another precondition row's already-mutated fixture.
            variant_uid = seed_mod.seed_variant(row.precondition_variant)
            variant_uids.append(variant_uid)
            uid = variant_uid
            _reset_memory(service, uid)
        else:
            if seeded_uid is None:
                seeded_uid = seed_mod.seed()
            else:
                seed_mod.seed(seeded_uid)
            uid = seeded_uid
            _reset_memory(service, uid)

        try:
            result = service.process_request(uid, row.send_text) or {}
            actual = (result.get("response") or "").strip()
            operation = result.get("operation") or ""
            error = ""
        except Exception as e:
            actual, operation = "", ""
            error = f"{type(e).__name__}: {e}"
            traceback.print_exc()

        if error:
            verdict, reason = "FAIL", f"bot raised {error}"
        else:
            g = grade_mod.grade(gemini, row.scenario, row.send_text,
                                row.expected, actual)
            verdict, reason = g["verdict"], g["reason"]

        if row.has_data_precondition and not row.precondition_variant:
            # P2-4 (PLAN_OF_ACTION.md): every _PRECONDITION_VARIANTS hint
            # now gets a dedicated fixture (see seed_mod.seed_variant) where
            # the precondition genuinely holds — this hedge only fires for
            # a hint that somehow matched has_data_precondition without a
            # mapped variant, which shouldn't happen but must not silently
            # mis-blame the bot if it ever does.
            reason = (reason + "  [row asserts a data precondition "
                      f"({row.annotation}) with no seed variant mapped for "
                      "it — the fixture may not meet it]").strip()

        results.append({
            "num": row.num, "category": row.category, "scenario": row.scenario,
            "message": row.message, "expected": row.expected,
            "language": row.language, "actual": actual, "operation": operation,
            "verdict": verdict, "reason": reason,
            "setup": " → ".join(setup_trace),
        })

        mark = {"PASS": "✓", "FAIL": "✗"}.get(verdict, "?")
        print(f"  {mark} {row.num:>3} {row.scenario[:44]:<44} {verdict}", flush=True)

    if seeded_uid:
        seed_mod.teardown(seeded_uid)
    for variant_uid in variant_uids:
        try:
            seed_mod.teardown(variant_uid)
        except Exception as e:
            # A failed teardown on one variant user must not lose the
            # results of an otherwise-complete run.
            print(f"  ! teardown failed for {variant_uid}: {e}", flush=True)
    return results


def write_workbook(results, path: str) -> None:
    """Same column shape as the previous live run's workbook, plus the
    judge's reason — a bare FAIL without a reason is not actionable."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["#", "Category", "Test Scenario", "User Message",
               "Expected Bot Behavior", "Language",
               f"Actual Bot Response (Live Run, {datetime.now():%Y-%m-%d})",
               "Operation", "Result", "Judge Reason", "Setup Replayed"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    fills = {
        "PASS":    PatternFill("solid", fgColor="C6EFCE"),
        "FAIL":    PatternFill("solid", fgColor="FFC7CE"),
        "UNCLEAR": PatternFill("solid", fgColor="FFEB9C"),
        "MANUAL":  PatternFill("solid", fgColor="D9D9D9"),
    }

    for r in results:
        ws.append([r["num"], r["category"], r["scenario"], r["message"],
                   r["expected"], r["language"], r["actual"][:2000],
                   r["operation"], r["verdict"], r["reason"], r.get("setup", "")])
        cell = ws.cell(row=ws.max_row, column=9)
        if r["verdict"] in fills:
            cell.fill = fills[r["verdict"]]

    for col, width in zip("ABCDEFGHIJK", (5, 20, 34, 34, 40, 10, 60, 20, 10, 44, 24)):
        ws.column_dimensions[col].width = width

    # Per-category summary — mirrors the Summary sheet in the source workbook.
    ws2 = wb.create_sheet("Summary")

    # Run-config fingerprint (P2-4, PLAN_OF_ACTION.md): every prior run's
    # score was silently ambiguous about which dispatch brain it measured —
    # the 62.5% baseline turned out to have been graded with FLOW_MACHINE_V2
    # OFF (legacy path), not the production-default config, and nothing in
    # the workbook itself would have told you that. Recorded here so a
    # score can never again be read without knowing what it actually ran
    # against.
    _fingerprint = _run_config_fingerprint()
    for label, value in _fingerprint:
        ws2.append([label, value])
    for row in ws2.iter_rows(min_row=1, max_row=len(_fingerprint)):
        row[0].font = Font(bold=True)
    ws2.append([])

    ws2.append(["Category", "Total", "Pass", "Fail", "Unclear", "Manual",
                "Pass % (of graded)"])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True)

    def _counts(rs):
        p = sum(1 for r in rs if r["verdict"] == "PASS")
        f = sum(1 for r in rs if r["verdict"] == "FAIL")
        u = sum(1 for r in rs if r["verdict"] == "UNCLEAR")
        m = sum(1 for r in rs if r["verdict"] == "MANUAL")
        # MANUAL rows are excluded from the denominator: they were never
        # graded, so counting them would understate quality by fiat.
        graded = p + f + u
        return p, f, u, m, (round(100 * p / graded, 1) if graded else 0.0)

    cats = []
    for r in results:
        if r["category"] not in cats:
            cats.append(r["category"])
    for cat in cats:
        rs = [r for r in results if r["category"] == cat]
        p, f, u, m, pct = _counts(rs)
        ws2.append([cat, len(rs), p, f, u, m, pct])

    p, f, u, m, pct = _counts(results)
    ws2.append(["TOTAL", len(results), p, f, u, m, pct])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True)

    wb.save(path)


def regrade(path: str, gemini):
    """Re-judge a completed run from its workbook, without touching the bot.

    The bot's replies are already recorded, so a grader fix does not need a
    second live run: re-judging costs one call per row instead of two plus
    all the DB work, and it isolates the change — any difference is the
    grader's, not the bot behaving differently on a re-ask.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb["Results"]

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        (num, category, scenario, message, expected, language,
         actual, operation, verdict, reason) = row[:10]
        setup = row[10] if len(row) > 10 else ""

        if verdict != "MANUAL":
            g = grade_mod.grade(gemini, scenario or "", message or "",
                                expected or "", actual or "")
            if g["verdict"] != verdict:
                print(f"  {num:>3} {str(scenario)[:40]:<40} {verdict} → {g['verdict']}",
                      flush=True)
            verdict, reason = g["verdict"], g["reason"]

        results.append({
            "num": num, "category": category, "scenario": scenario,
            "message": message, "expected": expected, "language": language,
            "actual": actual or "", "operation": operation or "",
            "verdict": verdict, "reason": reason, "setup": setup or "",
        })
    return results


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--regrade", metavar="XLSX",
                    help="re-judge a completed run's stored replies "
                         "(no bot calls, no DB writes)")
    ap.add_argument("--category", action="append",
                    help="only these categories (repeatable)")
    ap.add_argument("--language", help="only this language (EN / Hindi / Hinglish)")
    ap.add_argument("--limit", type=int, help="first N rows — harness smoke test")
    ap.add_argument("--out", default=None, help="output .xlsx path")
    args = ap.parse_args()

    # P2-4 (PLAN_OF_ACTION.md): announce the config every run measures, up
    # front — not just in the workbook after the fact. A run against the
    # wrong FLOW_MACHINE_V2 state is easy to kick off by accident (it's an
    # ambient env var, not a flag this script takes) and expensive to
    # discover after 148 rows of paid AI calls.
    print("Run config:")
    for label, value in _run_config_fingerprint():
        print(f"  {label}: {value}")
    print()

    from tests.conftest import has_real_ai_key, has_real_db_url
    if not has_real_ai_key():
        print("AI_KEY is missing or a placeholder — this run needs a real key.")
        return 2

    if args.regrade:
        from services.gemini_service import GeminiService
        print(f"Re-grading {args.regrade} (judge only, no bot calls)")
        results = regrade(args.regrade, GeminiService())
        out = args.out or args.regrade
        write_workbook(results, out)
        p = sum(1 for r in results if r["verdict"] == "PASS")
        f = sum(1 for r in results if r["verdict"] == "FAIL")
        u = sum(1 for r in results if r["verdict"] == "UNCLEAR")
        m = sum(1 for r in results if r["verdict"] == "MANUAL")
        graded = p + f + u
        print(f"\n{p}/{graded} PASS   {f} FAIL   {u} UNCLEAR   {m} MANUAL")
        print(f"wrote {out}")
        return 0

    if not has_real_db_url():
        print("SUPABASE_DB_URL is missing — this run needs a real database.")
        return 2

    rows = matrix_mod.load()
    if args.category:
        wanted = {c.lower() for c in args.category}
        rows = [r for r in rows if r.category.lower() in wanted]
    if args.language:
        rows = [r for r in rows if r.language.lower() == args.language.lower()]
    if args.limit:
        rows = rows[:args.limit]

    if not rows:
        print("no rows selected")
        return 2

    print(f"Running {len(rows)} matrix rows "
          f"({len({r.category for r in rows})} categories)")

    from services.intent_service import IntentService
    from services.gemini_service import GeminiService
    service = IntentService()
    gemini = GeminiService()

    out = args.out or os.path.join(
        os.path.dirname(__file__),
        f"matrix_results_{datetime.now():%Y%m%d_%H%M}.xlsx")

    # Collect into a list the runner appends to, and write it out even if the
    # run dies partway. A dropped database connection at row 122 of 148
    # destroyed a complete run's worth of paid AI calls because the workbook
    # was only written at the end — partial results are worth far more than
    # nothing.
    results: list = []
    try:
        results = run(rows, service, gemini, sink=results)
    except BaseException as e:
        print(f"\nRUN ABORTED after {len(results)} rows: {type(e).__name__}: {e}")
        if results:
            write_workbook(results, out)
            print(f"wrote partial results to {out}")
        raise

    write_workbook(results, out)

    p = sum(1 for r in results if r["verdict"] == "PASS")
    f = sum(1 for r in results if r["verdict"] == "FAIL")
    u = sum(1 for r in results if r["verdict"] == "UNCLEAR")
    m = sum(1 for r in results if r["verdict"] == "MANUAL")
    graded = p + f + u
    print(f"\n{p}/{graded} PASS   {f} FAIL   {u} UNCLEAR   {m} MANUAL "
          f"({100 * p / graded:.1f}% of graded)" if graded else "\nnothing graded")
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
