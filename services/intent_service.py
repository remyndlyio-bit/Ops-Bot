from services.gemini_service import GeminiService
from services.resend_email_service import ResendEmailService
from services.supabase_service import SupabaseService, JOB_ENTRIES_COLUMNS, _COLUMN_SCHEMA_FROM_ENV
from utils.date_utils import month_name_to_number, number_to_month_name
from services.sql_generator import generate_sql
from services.sql_validator import validate_sql
from services.query_planner import execute_query_plan
from services.query_router import (
    route_common_query,
    format_client_list,
    format_payment_status,
    ROWS as _RENDER_ROWS,
    AGGREGATE as _RENDER_AGGREGATE,
    CLIENT_LIST as _RENDER_CLIENT_LIST,
    PAYMENT_STATUS as _RENDER_PAYMENT_STATUS,
)
from services.response_formatter import (
    format_response,
    ASSISTANT_MODE,
    REMINDER_MODE,
    ERROR_MODE,
    clarify_phrase,
    error_calm_phrase,
    query_invalid_phrase,
    unsupported_feature_phrase,
)
from services.response_synthesis import build_clean_payload, build_field_answer_payload
from utils.memory_service import MemoryService
from utils.pending_reminders import get_pending, clear_pending, remove_single
from utils.logger import logger
from typing import Dict, List, Optional
import json
import os
import re
import time


def _is_full_job_row(row: dict) -> bool:
    """True when the row is a SELECT * from job_entries (not an aggregate)."""
    return "bill_no" in row or "job_date" in row


def _is_aggregate_sql(sql: str) -> bool:
    """True when the SQL is an aggregate / GROUP BY query (has GROUP BY, an
    aggregate function, or an `AS result` alias). Such queries must NOT be
    rewritten to `SELECT *` for history questions — doing so drops the aliased
    aggregate that ORDER BY / HAVING still reference, causing Postgres
    'column "result" does not exist'."""
    if not sql:
        return False
    return bool(
        re.search(r"\bGROUP\s+BY\b", sql, re.IGNORECASE)
        or re.search(r"\b(SUM|AVG|COUNT|MIN|MAX)\s*\(", sql, re.IGNORECASE)
        or re.search(r"\bAS\s+result\b", sql, re.IGNORECASE)
    )


def _format_job_card(row: dict) -> str:
    client = (row.get("client_name") or row.get("brand_name") or
              row.get("production_house") or "—").strip()
    brand = (row.get("brand_name") or "").strip()
    poc_name = (row.get("poc_name") or "").strip()
    poc_email = (row.get("poc_email") or "").strip()
    if poc_name and poc_email:
        poc = f"{poc_name} ({poc_email})"
    elif poc_name:
        poc = poc_name
    elif poc_email:
        poc = poc_email
    else:
        poc = "—"

    fees = row.get("fees")
    try:
        amount = f"₹{int(float(fees)):,}" if fees is not None else "—"
    except (ValueError, TypeError):
        amount = str(fees) if fees else "—"

    inv_date_raw = row.get("invoice_date")
    if inv_date_raw:
        try:
            from datetime import datetime as _dt
            inv_date_str = _dt.strptime(str(inv_date_raw)[:10], "%Y-%m-%d").strftime("%-d %b %Y")
        except Exception:
            inv_date_str = str(inv_date_raw)[:10]
    else:
        inv_date_str = "Not sent"

    bill_no = (row.get("bill_no") or "—").strip()

    lines = [f"Client: {client}"]
    if brand and brand.lower() != client.lower():
        lines.append(f"Brand: {brand}")
    lines.append(f"POC: {poc}")
    lines.append(f"Amount: {amount}")
    lines.append(f"Invoice Date: {inv_date_str}")
    lines.append(f"Invoice No: {bill_no}")
    return "\n".join(lines)


def _format_job_cards(rows: list) -> str:
    cards = []
    for i, row in enumerate(rows, 1):
        prefix = f"Job {i}\n" if len(rows) > 1 else ""
        cards.append(prefix + _format_job_card(row))
    return "\n\n".join(cards)


def _format_aggregate_fallback(payload: dict, user_message: str) -> str:
    """
    Deterministic formatter for simple aggregate payloads when AI synthesis returns empty.
    Handles type='aggregate' (SUM/COUNT/AVG result) and type='multi_record' GROUP BY results.
    Never returns the generic 'couldn't format' error string.
    """
    msg = user_message.strip().lower()
    p_type = payload.get("type", "")

    if p_type == "aggregate":
        val = (payload.get("data") or {}).get("result", 0)
        is_zero = val is None or val == 0
        # Decide if it's a money answer or a count answer based on message keywords
        is_count = any(k in msg for k in ("how many", "count", "kitne", "number of"))
        if is_count:
            n = int(val or 0)
            return f"{'No' if n == 0 else n} {'jobs' if 'job' in msg else 'records'} found."
        else:
            amount = int(val or 0)
            if is_zero:
                return "₹0 for that period — no matching records."
            return f"₹{amount:,}"

    if p_type in ("multi_record", "job_summary", "job_list"):
        data = payload.get("data") or []
        if not data:
            return "No matching records found."
        if isinstance(data, dict):
            data = [data]
        # GROUP BY result: each row has client_name + result
        lines = []
        for row in data[:10]:
            client = row.get("client_name") or row.get("brand_name") or "Unknown"
            amount = row.get("result") or row.get("fee") or row.get("fees") or 0
            try:
                lines.append(f"• {client}: ₹{int(float(amount)):,}")
            except (ValueError, TypeError):
                lines.append(f"• {client}: {amount}")
        return "\n".join(lines) if lines else "No matching records found."

    return "No matching records found."


def _generate_jobs_excel(rows: list, user_id: str) -> str:
    """Generate a branded Remyndly xlsx. Returns the file path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Fixed columns — exactly the discussed job format, no extras
    COLS = [
        ("client_name",  "Client Name"),
        ("brand_name",   "Brand Name"),
        ("_poc",         "POC"),           # computed: poc_name (poc_email)
        ("fees",         "Amount"),
        ("invoice_date", "Invoice Date"),
        ("bill_no",      "Invoice No"),
    ]
    NUM_COLS = len(COLS)

    BRAND_COLOR  = "1A1A2E"   # dark navy — Remyndly brand
    HEADER_COLOR = "2D6BE4"   # blue column headers
    ALT_COLOR    = "EEF3FC"   # alternating row tint
    WHITE        = "FFFFFF"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # ── Row 1: Remyndly branding banner ─────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    brand_cell = ws.cell(row=1, column=1, value="Remyndly")
    brand_cell.font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
    brand_cell.fill      = PatternFill("solid", fgColor=BRAND_COLOR)
    brand_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: column headers ────────────────────────────────────────
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)
    for ci, (_, header) in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.font      = Font(name="Calibri", bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[2].height = 20

    # ── Rows 3+: data ────────────────────────────────────────────────
    for ri, row in enumerate(rows, 3):
        fill = PatternFill("solid", fgColor=ALT_COLOR) if ri % 2 == 1 else None
        for ci, (field, _) in enumerate(COLS, 1):
            if field == "_poc":
                poc_name  = (row.get("poc_name")  or "").strip()
                poc_email = (row.get("poc_email") or "").strip()
                if poc_name and poc_email:
                    val = f"{poc_name} ({poc_email})"
                else:
                    val = poc_name or poc_email or ""
            else:
                raw = row.get(field)
                if field == "fees" and raw is not None:
                    try:
                        val = int(float(raw))
                    except (ValueError, TypeError):
                        val = raw
                elif field == "invoice_date" and raw:
                    try:
                        from datetime import datetime as _dt
                        val = _dt.strptime(str(raw)[:10], "%Y-%m-%d").strftime("%-d %b %Y")
                    except Exception:
                        val = str(raw)[:10]
                else:
                    val = str(raw).strip() if raw is not None else ""

            cell = ws.cell(row=ri, column=ci, value=val if val != "" else None)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

        ws.row_dimensions[ri].height = 16

    # ── Column widths ────────────────────────────────────────────────
    # Row 1 is the merged Remyndly banner — its non-anchor cells are MergedCell
    # instances that don't expose .column_letter. Use get_column_letter(idx) instead.
    from openpyxl.utils import get_column_letter
    WIDTHS = [22, 22, 32, 14, 16, 18]
    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    safe_uid = re.sub(r'[^a-zA-Z0-9]', '_', str(user_id))[-20:]
    filename = f"jobs_{safe_uid}_{int(time.time())}.xlsx"
    # Ensure the output directory exists. It is gitignored, so a fresh worktree
    # or deploy won't have it — without this, every >4-row query crashes with
    # FileNotFoundError when saving the spreadsheet.
    os.makedirs("output", exist_ok=True)
    path = os.path.join("output", filename)
    wb.save(path)

    # Sister CSV file at the same path with .csv extension. WhatsApp's
    # WhatsApp Business API rejects xlsx deliveries somewhat unpredictably
    # (Twilio 63019 = Meta-side internal failure on xlsx) — CSV is plain
    # text, universally accepted, and opens in Excel / Sheets / Numbers.
    # main.py picks per-platform: WhatsApp → .csv, Telegram → .xlsx.
    try:
        import csv as _csv
        csv_path = path[:-5] + ".csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            w.writerow(["Remyndly — Jobs Export"])
            w.writerow([h for _, h in COLS])
            for row in rows:
                line = []
                for field, _ in COLS:
                    if field == "_poc":
                        poc_name  = (row.get("poc_name")  or "").strip()
                        poc_email = (row.get("poc_email") or "").strip()
                        if poc_name and poc_email:
                            line.append(f"{poc_name} ({poc_email})")
                        else:
                            line.append(poc_name or poc_email or "")
                    elif field == "fees":
                        raw = row.get("fees")
                        try:
                            line.append(int(float(raw)) if raw is not None else "")
                        except (ValueError, TypeError):
                            line.append(raw or "")
                    elif field == "invoice_date":
                        raw = row.get("invoice_date")
                        if raw:
                            try:
                                from datetime import datetime as _dt
                                line.append(_dt.strptime(str(raw)[:10], "%Y-%m-%d").strftime("%-d %b %Y"))
                            except Exception:
                                line.append(str(raw)[:10])
                        else:
                            line.append("")
                    else:
                        v = row.get(field)
                        line.append(str(v).strip() if v is not None else "")
                w.writerow(line)
        logger.info(f"[EXCEL] Wrote sister CSV: {csv_path}")
    except Exception as _csv_err:
        logger.warning(f"[EXCEL] CSV sister write failed (xlsx still usable): {_csv_err}")

    # Sister PDF — Twilio's WhatsApp channel only reliably accepts PDF for
    # outbound documents. xlsx → 63019, csv → 63005 (channel doesn't support).
    # main.py picks .pdf for WhatsApp.
    try:
        from fpdf import FPDF
        pdf_path = path[:-5] + ".pdf"
        pdf = FPDF(orientation="L", unit="mm", format="A4")  # landscape for wider table
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.add_page()
        # Brand banner
        pdf.set_fill_color(26, 26, 46)  # BRAND_COLOR
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Remyndly - Jobs Export", ln=1, align="C", fill=True)
        pdf.ln(2)
        # Column widths in mm — must sum ≤ 277 (A4 landscape printable width)
        col_widths = [45, 40, 60, 30, 35, 35]
        headers = [h for _, h in COLS]
        # Header row
        pdf.set_fill_color(45, 107, 228)  # HEADER_COLOR
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 10)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 8, h, border=1, align="C", fill=True)
        pdf.ln(8)
        # Data rows
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(20, 20, 20)
        for ri, row in enumerate(rows):
            # Alternating row colour
            if ri % 2 == 0:
                pdf.set_fill_color(238, 243, 252)  # ALT_COLOR
                use_fill = True
            else:
                use_fill = False
            cells = []
            for field, _ in COLS:
                if field == "_poc":
                    poc_name  = (row.get("poc_name")  or "").strip()
                    poc_email = (row.get("poc_email") or "").strip()
                    if poc_name and poc_email:
                        v = f"{poc_name} ({poc_email})"
                    else:
                        v = poc_name or poc_email or ""
                elif field == "fees":
                    raw = row.get("fees")
                    try:
                        v = f"Rs {int(float(raw)):,}" if raw is not None else ""
                    except (ValueError, TypeError):
                        v = str(raw) if raw else ""
                elif field == "invoice_date":
                    raw = row.get("invoice_date")
                    if raw:
                        try:
                            from datetime import datetime as _dt
                            v = _dt.strptime(str(raw)[:10], "%Y-%m-%d").strftime("%-d %b %Y")
                        except Exception:
                            v = str(raw)[:10]
                    else:
                        v = ""
                else:
                    raw = row.get(field)
                    v = str(raw).strip() if raw is not None else ""
                # fpdf2 with the built-in Helvetica font only supports
                # Latin-1. Any non-Latin glyph raises — so we normalise all
                # the common offenders to ASCII equivalents.
                v = (
                    v.replace("₹", "Rs ")
                     .replace("—", "-").replace("–", "-")  # em + en dash
                     .replace("'", "'").replace("'", "'")  # smart single quotes
                     .replace(""", '"').replace(""", '"')  # smart double quotes
                     .replace("…", "...")                  # horizontal ellipsis (THIS bit us)
                )
                cells.append(v)
            for i, v in enumerate(cells):
                # Truncate to fit; the table is for at-a-glance, not exhaustive.
                # MUST use ASCII '...' not '…' — Helvetica rejects U+2026
                # and the whole PDF generation fails (we already lost a
                # production day to that one).
                max_chars = max(8, int(col_widths[i] / 1.8))
                if len(v) > max_chars:
                    v = v[:max_chars - 3] + "..."
                pdf.cell(col_widths[i], 6, v, border=1, align="L", fill=use_fill)
            pdf.ln(6)
        pdf.output(pdf_path)
        logger.info(f"[EXCEL] Wrote sister PDF: {pdf_path}")
    except Exception as _pdf_err:
        logger.warning(f"[EXCEL] PDF sister write failed (xlsx/csv still usable): {_pdf_err}")
    return path


class IntentService:
    # Cache AI-generated schema by column names so we don't call the AI on every message
    _schema_cache: Dict[tuple, str] = {}

    # Required fields for smart capture job creation
    _SMART_CAPTURE_REQUIRED = ["brand_name", "job_date", "job_description_details"]

    # Trigger phrases for bank detail commands
    _UPDATE_BANK_TRIGGERS = [
        "update bank details", "update bank detail", "change bank details",
        "set bank details", "edit bank details", "add bank details",
        "update my bank", "change my bank", "set my bank",
        "save bank details", "new bank details",
    ]
    _VIEW_BANK_TRIGGERS = [
        "my bank details", "show bank details", "view bank details",
        "what are my bank details", "bank details", "show my bank",
        "get bank details", "see bank details", "check bank details",
    ]

    # Small-talk trigger words / phrases (case-insensitive, matched as whole tokens)
    _SMALL_TALK_TRIGGERS = {
        "hi", "hey", "hello", "hiya", "howdy", "yo", "sup", "heya",
        "how are you", "how r u", "how are u", "how are you doing",
        "how\'s it going", "hows it going", "how do you do",
        "what\'s up", "whats up", "wassup",
        "thanks", "thank you", "thx", "ty", "cheers",
        "bye", "goodbye", "good bye", "see you", "see ya", "cya", "ttyl",
        "ok", "okay", "cool", "got it", "great", "nice", "awesome",
        "good morning", "good afternoon", "good evening", "good night",
        "morning", "afternoon", "evening",
        # dismissal / "nothing needed" replies
        "nothing", "nothing thanks", "nothing thank you", "nothing, thanks",
        "no thanks", "no thank you", "nope thanks", "nah thanks",
        "all good", "all good thanks", "i'm good", "im good", "i'm fine", "im fine",
        "that's all", "thats all", "that's it", "thats it",
        "no need", "not needed", "never mind", "nevermind", "nvm",
        "i'm ok", "im ok", "i'm okay", "im okay",
        # Hindi / Hinglish greetings & small talk
        "namaste", "namaskar", "jai hind",
        "kya haal hai", "kya hal hai", "kya haal h", "kya hal h",
        "kaise ho", "kaisa chal raha hai", "kya chal raha hai", "kya scene hai",
        "sab theek", "sab thik", "theek hoon", "thik hoon", "mast hoon",
        "shukriya", "dhanyavaad", "shukran",
        "haan", "haan ji", "nahi", "nahi ji",
        "alvida", "bye bye", "phir milenge",
        "kal milte hain", "baad mein baat karte hain",
        "bas karo", "rehne do", "chhoddo",
    }

    _SMALL_TALK_RESPONSES = {
        "greeting": [
            "Hey! What can I help you with today?",
            "Hi there! Need an invoice, a query, or something else?",
            "Hello! Ready when you are — just tell me what you need.",
        ],
        "how_are_you": [
            "Doing great, thanks for asking! What can I pull up for you?",
            "All good on my end! What do you need today?",
            "Running smoothly! What can I help with?",
        ],
        "thanks": [
            "Happy to help! Anything else?",
            "Anytime! Let me know if you need more.",
            "Of course! Just ask if there\'s anything else.",
        ],
        "bye": [
            "Take care! Come back anytime.",
            "Goodbye! Have a great day.",
            "See you! I\'ll be here whenever you need me.",
        ],
        "affirmation": [
            "Got it! Let me know if there\'s anything else.",
            "Sure thing! Anything else I can help with?",
        ],
        "time_of_day": [
            "Good to hear from you! What do you need?",
            "Hope your day\'s going well! What can I help with?",
        ],
        "good_morning": [
            "Good morning! What can I help you with today?",
            "Morning! Ready when you are — what do you need?",
        ],
        "good_afternoon": [
            "Good afternoon! What can I help you with?",
            "Afternoon! What do you need?",
        ],
        "good_evening": [
            "Good evening! What can I help you with?",
            "Evening! What do you need?",
        ],
        "good_night": [
            "Good night! Take care.",
            "Good night! Rest well.",
        ],
    }

    @staticmethod
    def _is_valid_email(email: str) -> bool:
        """Strict email format check (single token, local@domain.tld, no spaces)."""
        if not email:
            return False
        return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', str(email).strip()))

    def __init__(self):
        self.gemini = GeminiService()
        self.email = ResendEmailService()
        self.supabase = SupabaseService()
        self.memory = MemoryService()
        # Column schema (if provided via COLUMN_SCHEMA env) for AI validation.
        self.column_schema = _COLUMN_SCHEMA_FROM_ENV or {}
        # FlowMachine v2 (session 2+). Owned alongside the legacy awaiting_*
        # flag bag during migration; either path can reset state if v2 is
        # flag-flipped on/off mid-flow.
        from services.flow_machine import FlowMachine as _FlowMachine
        self.flow_machine = _FlowMachine(self.memory)

    def _reconcile_legacy_to_flow_machine(self, user_id: str, user_mem: Dict) -> None:
        """Once per message: if FlowMachine is IDLE but a legacy awaiting_* flag
        (or a form_state) was armed on the previous turn, sync FlowMachine to
        match. This lets dispatch_in_flow take over without modifying every
        legacy arm site. No-op when FlowMachine is already tracking a flow."""
        from services.flow_machine import (
            FLOW_IDLE,
            FLOW_INVOICE_AWAIT_SEND_CONFIRM, FLOW_INVOICE_NEED_BILLING,
            FLOW_INVOICE_NEED_POC_NAME, FLOW_INVOICE_NEED_POC_EMAIL,
            FLOW_SMART_CAPTURE_NEED_DESCRIPTION, FLOW_SMART_CAPTURE_CONFIRM_PENDING,
        )
        # Already tracking — nothing to reconcile.
        if self.flow_machine.current_flow(user_id) != FLOW_IDLE:
            return

        # Order matters: form_state (smart-capture confirm) is the "deepest"
        # state, so it wins if multiple flags happen to be set.
        if self.memory.get_form_state(user_id):
            self.flow_machine.set_state(
                user_id, FLOW_SMART_CAPTURE_CONFIRM_PENDING,
                {"source": "reconcile_form_state"},
            )
            return

        if user_mem.get("awaiting_send_confirmation"):
            pend = user_mem.get("pending_send_invoice") or {}
            self.flow_machine.set_state(
                user_id, FLOW_INVOICE_AWAIT_SEND_CONFIRM,
                {
                    "client_name": pend.get("client_name"),
                    "month":       pend.get("month"),
                    "year":        pend.get("year"),
                    "poc_email":   pend.get("poc_email"),
                },
            )
            return

        if user_mem.get("awaiting_client_billing"):
            self.flow_machine.set_state(
                user_id, FLOW_INVOICE_NEED_BILLING,
                {"client_name": user_mem.get("pending_billing_client")},
            )
            return

        if user_mem.get("awaiting_poc_name"):
            self.flow_machine.set_state(
                user_id, FLOW_INVOICE_NEED_POC_NAME,
                {"client_name": user_mem.get("pending_poc_client")},
            )
            return

        if user_mem.get("awaiting_poc_email"):
            pend = user_mem.get("pending_send_invoice") or {}
            self.flow_machine.set_state(
                user_id, FLOW_INVOICE_NEED_POC_EMAIL,
                {
                    "client_name": (user_mem.get("poc_email_client")
                                    or pend.get("client_name")),
                },
            )
            return

        if user_mem.get("awaiting_job_input"):
            self.flow_machine.set_state(
                user_id, FLOW_SMART_CAPTURE_NEED_DESCRIPTION, {},
            )
            return

    def _store_conversation(self, user_id: str, user_message: str, bot_response: str):
        """Store user message and bot response in conversation history."""
        self.memory.add_message(user_id, "user", user_message)
        self.memory.add_message(user_id, "assistant", bot_response)

    # ── Structured intent tracking for context reconstruction ──

    def _save_last_intent(self, user_id: str, *, operation: str = None, client_name: str = None,
                          month: str = None, year=None, entity: str = None,
                          pending_clarification: str = None, extra: dict = None):
        """Persist the most recent structured intent so follow-ups can inherit it."""
        intent = {k: v for k, v in {
            "operation": operation,
            "client_name": client_name,
            "month": month,
            "year": year,
            "entity": entity,
            "pending_clarification": pending_clarification,
        }.items() if v is not None}
        if extra:
            intent.update(extra)
        self.memory.update_user_memory(user_id, {"last_intent": intent})
        logger.info(f"[CONTEXT] Saved last_intent for {user_id}: {intent}")

    def _reconstruct_message(self, user_id: str, message: str, conversation_history: List[Dict]) -> str:
        """
        Context reconstruction: if the message is short/ambiguous, merge it
        with stored last_intent and recent conversation to produce a fully
        self-contained query.  Returns the original message unchanged when no
        reconstruction is needed.
        """
        msg_lower = message.strip().lower()
        word_count = len(message.strip().split())

        # Skip reconstruction for messages that are already self-contained
        # (long enough AND contain an action verb + entity)
        _ACTION_VERBS = {"generate", "create", "send", "show", "get", "list", "give",
                         "update", "delete", "remove", "add", "fetch", "download",
                         "make", "prepare", "invoice", "query", "find", "search"}
        has_action = any(v in msg_lower for v in _ACTION_VERBS)
        if word_count >= 4 and has_action:
            return message  # Already self-contained

        # Short or ambiguous message — try to reconstruct from context
        user_mem = self.memory.get_user_memory(user_id)
        last_intent = user_mem.get("last_intent", {})
        if not last_intent:
            return message  # No prior context to merge

        pending = last_intent.get("pending_clarification", "")
        operation = last_intent.get("operation", "")
        client_name = last_intent.get("client_name", "")
        month_val = last_intent.get("month", "")
        entity = last_intent.get("entity", "")

        # Get last assistant message to understand what was asked
        last_assistant_msg = ""
        if conversation_history:
            assistant_msgs = [m for m in conversation_history if m.get("role") == "assistant"]
            if assistant_msgs:
                last_assistant_msg = assistant_msgs[-1].get("content", "").lower()

        reconstructed = None

        # Case 0: Bot suggested an alternate month and user confirms ("okay", "yes",
        # "okay generate", "sure", "go ahead", "proceed"). Persisted last_intent has
        # pending_clarification == 'confirm_alt_month' with client + month + year.
        _AFFIRM_TOKENS = {"okay", "ok", "yes", "yep", "yeah", "sure", "go", "ahead",
                          "proceed", "fine", "alright", "yup", "haan", "thik", "theek",
                          "generate", "create", "send"}
        _msg_words = set(re.findall(r"[a-z]+", msg_lower))
        if pending == "confirm_alt_month" and client_name and month_val and (_msg_words & _AFFIRM_TOKENS) and word_count <= 4:
            year_part = f" {last_intent.get('year')}" if last_intent.get("year") else ""
            verb = "Send" if "send" in operation.lower() else "Generate"
            reconstructed = f"{verb} invoice for {client_name} for {month_val}{year_part}"

        # Case 1: Bot asked "Which month?" and user replied with a month name
        _MONTH_NAMES = {"january", "february", "march", "april", "may", "june",
                        "july", "august", "september", "october", "november", "december",
                        "jan", "feb", "mar", "apr", "jun", "jul", "aug", "sep", "oct", "nov", "dec"}
        first_word = msg_lower.split()[0] if msg_lower.split() else ""
        is_month_reply = first_word in _MONTH_NAMES or any(m in msg_lower for m in _MONTH_NAMES)
        if is_month_reply and pending == "month" and client_name:
            verb = "Send" if "send" in operation.lower() else "Generate"
            reconstructed = f"{verb} invoice for {client_name} for {message.strip()}"

        # Case 2: Bot asked for a client name and user replied with one.
        # Guard: skip reconstruction if the reply looks like a new query rather than a name.
        elif pending == "client_name" and word_count <= 4:
            _QUERY_SIGNALS = {
                "total", "earning", "earnings", "income", "revenue", "fees",
                "last month", "this month", "last year", "this year", "last week",
                "all", "how many", "how much", "show", "list", "what", "when",
                "unpaid", "paid", "pending", "outstanding",
            }
            _is_new_query = any(sig in msg_lower for sig in _QUERY_SIGNALS)
            if not _is_new_query:
                if "invoice" in entity or "invoice" in operation.lower():
                    month_part = f" for {month_val}" if month_val else ""
                    reconstructed = f"Generate invoice for {message.strip()}{month_part}"
                else:
                    reconstructed = f"Show jobs for {message.strip()}"

        # Case 3: User says something like "for March" or "for Garnier"
        elif msg_lower.startswith("for ") and word_count <= 4:
            rest = message.strip()[4:]  # strip "for "
            rest_lower = rest.lower().strip()
            if any(m in rest_lower for m in _MONTH_NAMES) and client_name:
                verb = "Send" if "send" in operation.lower() else "Generate"
                reconstructed = f"{verb} invoice for {client_name} for {rest.strip()}"
            elif operation and client_name:
                reconstructed = f"{operation} for {client_name} for {rest.strip()}"

        # Case 4: Very short replies (1-2 words) with a pending clarification
        elif word_count <= 2 and pending and client_name:
            if pending == "month":
                verb = "Send" if "send" in operation.lower() else "Generate"
                reconstructed = f"{verb} invoice for {client_name} for {message.strip()}"
            elif pending == "confirm":
                pass  # Let awaiting_* handlers deal with yes/no

        # Case 5: "this month", "last month", "this year" relative time with prior client
        elif client_name and any(t in msg_lower for t in ["this month", "last month", "this year", "last year"]):
            # If the last assistant turn was a query result (showing jobs/details),
            # the user is refining that query — NOT requesting a new invoice.
            # last_intent.entity may still say "invoice" from a much earlier flow,
            # so we trust the immediate conversation over stale stored intent.
            _query_result_signals = (
                "here are", "here's", "details for", "your jobs", "i found",
                "showing", "your last job", "most recent", "•",
            )
            just_answered_query = any(s in last_assistant_msg for s in _query_result_signals)
            invoice_context = ("invoice" in entity or "invoice" in operation.lower())
            if invoice_context and not just_answered_query:
                reconstructed = f"Generate invoice for {client_name} for {message.strip()}"
            else:
                reconstructed = f"Show jobs for {client_name} for {message.strip()}"

        if reconstructed and reconstructed.strip().lower() != message.strip().lower():
            logger.info(f"[CONTEXT] Reconstructed message: '{message}' → '{reconstructed}' "
                        f"(last_intent={last_intent})")
            return reconstructed

        return message

    def _get_schema_and_columns(self, records: List[Dict]) -> tuple:
        """Return (schema_description, allowed_columns, date_column). Prefer AI-generated schema; fallback to rule-based."""
        from services.business_logic_service import BusinessLogicService
        logic = BusinessLogicService()
        cols = list(records[0].keys()) if records else []
        if not cols:
            column_map = logic._get_column_names(None)
            cols = list({c for v in column_map.values() for c in v})
        # Prefer AI-generated schema; cache by column names so we don't call the AI every message
        cache_key = tuple(sorted(cols))
        schema_description = IntentService._schema_cache.get(cache_key)
        if not schema_description:
            schema_description = self.gemini.generate_schema_from_columns(
                cols,
                sample_row=records[0] if records else None,
            )
            if schema_description:
                IntentService._schema_cache[cache_key] = schema_description
        if not schema_description:
            schema_description = logic.get_schema_for_intent(cols)
        column_map = logic._get_column_names(cols)
        date_cols = column_map.get("invoice_date", []) or ([c for c in cols if "date" in c.lower()] if cols else [])
        date_column = date_cols[0] if date_cols else (cols[0] if cols else "Date")
        return schema_description, cols, date_column

    def _resolve_response_mode(self, result: Dict, cmd: Dict) -> str:
        """
        Determine ResponseMode based on priority:
        1. SINGLE_FIELD: return_fields has 1 field OR metric=value with column
        2. RECORD: multiple return_fields
        3. COUNT: metric=count with no specific field requested
        4. AGGREGATION: sum/avg/min/max
        5. GROUPED: group_by present
        6. CLARIFY: otherwise
        """
        metric = result.get("metric") or cmd.get("metric", "count")
        return_fields = result.get("return_fields") or cmd.get("return_fields") or []
        column = result.get("column") or cmd.get("column")
        group_by = cmd.get("group_by")

        # Priority 1: Single field requested
        if len(return_fields) == 1:
            return "SINGLE_FIELD"
        if metric == "value" and column:
            return "SINGLE_FIELD"

        # Priority 2: Multiple return fields
        if len(return_fields) > 1:
            return "RECORD"

        # Priority 5: Grouped results
        if group_by or ("labels" in result and result["labels"]):
            return "GROUPED"

        # Priority 3/4: Aggregation metrics
        if metric in ("sum", "avg", "min", "max"):
            return "AGGREGATION"

        if metric == "count":
            return "COUNT"

        return "CLARIFY"

    def _build_filter_context(self, filters: Dict) -> str:
        """Build human-readable context from filters (e.g., 'the Apple job')."""
        if not filters:
            return ""
        parts = []
        for k, v in filters.items():
            if v and not str(k).startswith("_"):
                k_lower = str(k).lower()
                if "client" in k_lower or "name" in k_lower:
                    parts.append(str(v))
                elif "date" in k_lower:
                    parts.append(f"on {v}")
        return " ".join(parts) if parts else ""

    def _format_uscf_result(self, result: Dict, cmd: Dict) -> str:
        """Format USCF executor result as factual output for response maker."""
        if not result.get("ok"):
            return result.get("message", "I don't see this information in my records.")

        operation = result.get("operation") or cmd.get("operation")

        # CREATE result
        if operation == "create":
            msg = result.get("message", "Record created.")
            bill_no = result.get("bill_number")
            if bill_no:
                msg += f" Invoice/Bill #: {bill_no}."
            return msg

        # UPDATE result
        if operation == "update":
            return result.get("message", f"Updated {result.get('count', 0)} record(s).")

        # DELETE result
        if operation == "delete":
            return result.get("message", f"Deleted {result.get('count', 0)} record(s).")

        # QUERY result - use ResponseMode resolver
        mode = self._resolve_response_mode(result, cmd)
        filters = result.get("filters") or cmd.get("filters") or {}
        context = self._build_filter_context(filters)
        column = result.get("column") or cmd.get("column", "")
        count = result.get("count", 0)
        metric = result.get("metric") or cmd.get("metric", "count")

        logger.info(f"[RESPONSE] mode={mode}, column={column}, metric={metric}, filters={filters}")

        # Date result (special case for max on date column)
        if result.get("value_type") == "date":
            val = result.get("value")
            if val is None:
                return result.get("message", "No date found.")
            try:
                from datetime import datetime
                dt = datetime.strptime(str(val)[:10], "%Y-%m-%d")
                return f"date: {dt.strftime('%d %b %Y')}"
            except ValueError:
                return f"date: {val}"

        # SINGLE_FIELD mode: return specific field value
        if mode == "SINGLE_FIELD":
            val = result.get("value")
            rows = result.get("rows", [])
            return_fields = result.get("return_fields") or [column]
            target_field = return_fields[0] if return_fields else column

            # Try to get value from result or first row
            if val is None and rows:
                val = rows[0].get(target_field)

            if val is None or (isinstance(val, str) and not val.strip()):
                return f"No {target_field} found{' for ' + context if context else ''}."

            # Format based on value type
            if isinstance(val, (int, float)):
                return f"{target_field}{' for ' + context if context else ''}: ₹{val:,.2f}"
            else:
                return f"{target_field}{' for ' + context if context else ''}: {val}"

        # RECORD mode: return multiple fields
        if mode == "RECORD":
            rows = result.get("rows", [])
            return_fields = result.get("return_fields", [])
            if not rows:
                return "No matching records found."
            lines = []
            for row in rows[:10]:
                parts = [f"{f}: {row.get(f, 'N/A')}" for f in return_fields]
                lines.append("• " + ", ".join(parts))
            return "\n".join(lines)

        # GROUPED mode: labels + values
        if mode == "GROUPED":
            labels = result.get("labels", [])
            values = result.get("values", [])
            if not labels:
                return "No grouped results."
            lines = []
            for idx, label in enumerate(labels[:30]):
                line = f"• {label}"
                if idx < len(values):
                    v = values[idx]
                    if isinstance(v, (int, float)):
                        if metric == "count":
                            line += f": {int(v)}"
                        else:
                            line += f": ₹{v:,.2f}"
                    else:
                        line += f": {v}"
                lines.append(line)
            if len(labels) > 30:
                lines.append(f"... and {len(labels) - 30} more.")
            return "\n".join(lines)

        # AGGREGATION mode: sum/avg/min/max
        if mode == "AGGREGATION":
            value = result.get("value", 0)
            if not isinstance(value, (int, float)):
                return str(value) if value else "No result."
            prefix = context + " " if context else ""
            if metric == "sum":
                return f"total {column}{' for ' + context if context else ''}: ₹{value:,.2f}"
            elif metric == "avg":
                return f"average {column}{' for ' + context if context else ''}: ₹{value:,.2f} (across {count} records)"
            elif metric == "min":
                return f"minimum {column}{' for ' + context if context else ''}: ₹{value:,.2f}"
            elif metric == "max":
                return f"maximum {column}{' for ' + context if context else ''}: ₹{value:,.2f}"

        # COUNT mode (only when no specific field requested)
        if mode == "COUNT":
            value = result.get("value", result.get("count", 0))
            return f"count{' for ' + context if context else ''}: {int(value)}"

        # CLARIFY fallback
        return "Could you clarify what specific information you're looking for?"

    def _format_sql_result(self, rows: List[Dict]) -> str:
        """Format SQL result rows into a short factual reply. Never returns empty when rows exist."""
        if not rows:
            return "No matching records found."
        def _fmt_val(v):
            return v if v is not None else "N/A"
        if len(rows) == 1 and len(rows[0]) <= 3:
            parts = [f"{k}: {_fmt_val(v)}" for k, v in rows[0].items()]
            out = ", ".join(parts)
            return out if out.strip() else "1 row (no values)"
        if len(rows) == 1:
            lines = [f"• {k}: {_fmt_val(v)}" for k, v in list(rows[0].items())[:15]]
            out = "\n".join(lines)
            return out if out.strip() else "1 row (no values)"
        keys = list(rows[0].keys())[:6]
        lines = []
        for r in rows[:20]:
            parts = [f"{k}: {_fmt_val(r.get(k))}" for k in keys]
            lines.append("• " + ", ".join(parts))
        if len(rows) > 20:
            lines.append(f"... and {len(rows) - 20} more.")
        return "\n".join(lines)

    # Whitelist of columns the modify flow is allowed to touch.
    _MODIFY_ALLOWED_FIELDS = {
        "fees", "paid", "client_name", "brand_name",
        "job_description_details", "invoice_date", "poc_email",
        "poc_name", "deadline_date", "job_date", "production_house",
    }
    _MODIFY_FIELD_ALIASES = {
        "fee": "fees", "amount": "fees", "price": "fees", "cost": "fees",
        "payment_status": "paid", "payment": "paid", "status": "paid",
        "client": "client_name", "brand": "brand_name",
        "description": "job_description_details", "details": "job_description_details",
        "job_description": "job_description_details",
        "billing_date": "invoice_date", "bill_date": "invoice_date",
        "email": "poc_email", "contact_email": "poc_email",
        "poc": "poc_name", "contact": "poc_name", "contact_person": "poc_name",
        "deadline": "deadline_date", "due_date": "deadline_date",
        "date": "job_date",
    }

    def _normalize_modify_field(self, raw: str) -> Optional[str]:
        if not raw:
            return None
        k = str(raw).strip().lower().replace(" ", "_")
        k = self._MODIFY_FIELD_ALIASES.get(k, k)
        return k if k in self._MODIFY_ALLOWED_FIELDS else None

    def _handle_modify_intent(self, user_id: str, message: str, user_mem: Dict) -> Optional[Dict]:
        """
        B → A modify flow. Extract field+value via Gemini; if either is missing,
        ask a clarifying question. Apply via update_job_entry_field with whitelist.
        Returns response dict on success, or None to fall through.
        """
        import re as _re
        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
        last_row = ctx.get("last_row_data") or {}

        awaiting = user_mem.get("awaiting_modify_field")
        # Carry forward a pinned row_id if we already asked which row to change.
        pinned_row_id = user_mem.get("modify_row_id") or last_row.get("id")

        parsed = self.gemini.extract_modify_intent(message, last_row if last_row else None)
        if not parsed:
            parsed = {}

        field = self._normalize_modify_field(parsed.get("field"))
        value = parsed.get("value")
        client_filter = parsed.get("client_filter") or ""
        bill_filter = parsed.get("bill_no_filter") or ""

        # No field/value parsed — ask
        if not field or value is None or str(value).strip() == "":
            # If we have a row in context, ask scoped to it.
            if pinned_row_id:
                _client = last_row.get("client_name") or last_row.get("brand_name") or "this job"
                resp = (
                    f"What would you like to change about {_client}? "
                    "Reply like: `fee: 25000`, `paid: yes`, `contact email: x@y.com`."
                )
                self.memory.update_user_memory(user_id, {
                    "awaiting_modify_field": True,
                    "modify_row_id": pinned_row_id,
                })
                self._store_conversation(user_id, message, resp)
                return {"operation": "modify_prompt", "response": resp, "trigger_invoice": False, "invoice_data": {}}
            # No context — let the normal pipeline have a go
            return None

        # Resolve target row id
        target_id = None
        if not client_filter and not bill_filter and pinned_row_id:
            target_id = pinned_row_id
        else:
            # Look up by filter
            where = []
            _profile = self.supabase.get_user_profile(user_id) or {}
            data_user_id = self._resolve_data_user_id(user_id, _profile.get("data", {}))
            params_user = str(data_user_id).replace("'", "''")
            if bill_filter:
                where.append(f"bill_no ILIKE '{str(bill_filter).replace(chr(39), chr(39)*2)}'")
            if client_filter:
                cf = str(client_filter).replace("'", "''")
                where.append(f"(client_name ILIKE '%{cf}%' OR brand_name ILIKE '%{cf}%' OR production_house ILIKE '%{cf}%')")
            where_clause = " AND ".join(where) if where else "TRUE"
            sql = (
                f"SELECT id, client_name, brand_name, bill_no, fees FROM public.job_entries "
                f"WHERE user_id = '{params_user}' AND ({where_clause}) "
                f"ORDER BY created_at DESC LIMIT 5"
            )
            res = self.supabase.execute_sql(sql)
            rows = res.get("rows", []) if res.get("ok") else []
            if len(rows) == 0:
                resp = f"I couldn't find a job matching that. Tell me the client name or bill number."
                self._store_conversation(user_id, message, resp)
                return {"operation": "modify_no_match", "response": resp, "trigger_invoice": False, "invoice_data": {}}
            if len(rows) > 1:
                lines = ["Multiple matches — which one?"]
                for i, r in enumerate(rows, 1):
                    lines.append(f"{i}. {r.get('client_name') or r.get('brand_name') or '—'} · bill {r.get('bill_no') or '—'} · ₹{r.get('fees') or '—'}")
                resp = "\n".join(lines)
                # Store pending disambiguation so user can reply with a number
                self.memory.update_user_memory(user_id, {
                    "pending_disambiguation": {
                        "type": "modify",
                        "rows": rows,
                        "field": field,
                        "value": value,
                    },
                    "awaiting_modify_field": False,
                })
                self._store_conversation(user_id, message, resp)
                return {"operation": "modify_disambiguate", "response": resp, "trigger_invoice": False, "invoke_data": {}}
            target_id = rows[0]["id"]

        # Normalize value
        if field == "paid":
            sv = str(value).strip().lower()
            value = "Yes" if sv in ("yes", "true", "1", "paid", "y") else "No"
        elif field == "fees":
            try:
                value = int(float(str(value).replace(",", "").replace("₹", "").strip()))
            except (ValueError, TypeError):
                resp = f"I couldn't parse '{value}' as a fee amount. Try a number like 25000."
                self._store_conversation(user_id, message, resp)
                return {"operation": "modify_bad_value", "response": resp, "trigger_invoice": False, "invoice_data": {}}

        # Fetch old value + existing notes before overwriting
        _safe_id = str(target_id).replace("'", "''")
        _safe_field = field.replace('"', '')
        pre = self.supabase.execute_sql(
            f'SELECT "{_safe_field}", notes FROM public.job_entries WHERE id = \'{_safe_id}\''
        )
        old_value = None
        existing_notes = ""
        if pre.get("ok") and pre.get("rows"):
            old_value = pre["rows"][0].get(field)
            existing_notes = pre["rows"][0].get("notes") or ""

        # Build change-history entry to append to notes
        from datetime import date as _date
        _label = self._MODIFY_FIELD_ALIASES.get(field, field).replace("_", " ")
        _old_disp = (f"₹{int(float(old_value)):,}" if field == "fees" and old_value is not None
                     else str(old_value) if old_value is not None else "—")
        _new_disp = f"₹{value:,}" if field == "fees" else str(value)
        history_entry = f"[{_date.today().strftime('%d %b %Y')}] {_label}: {_old_disp} → {_new_disp}"
        new_notes = (existing_notes + "\n" + history_entry).strip() if existing_notes else history_entry

        # Apply field update + notes append + RETURNING * in one shot
        logger.info(f"[MODIFY] Writing history: {history_entry}")
        _val_param = value if not isinstance(value, str) else value.replace("'", "''")
        _notes_param = new_notes.replace("'", "''")
        update_sql = (
            f'UPDATE public.job_entries SET "{field}" = \'{_val_param}\', '
            f'notes = \'{_notes_param}\' '
            f'WHERE id = \'{_safe_id}\' RETURNING *'
        )
        result = self.supabase.execute_sql(update_sql)

        # Clear awaiting state regardless
        self.memory.update_user_memory(user_id, {
            "awaiting_modify_field": False,
            "modify_row_id": None,
        })
        if not result.get("ok"):
            resp = f"Update failed: {result.get('error', 'unknown error')}"
            self._store_conversation(user_id, message, resp)
            return {"operation": "modify_failed", "response": resp, "trigger_invoice": False, "invoice_data": {}}

        # Refresh context with full updated row (includes new notes with history)
        updated_rows = result.get("rows") or []
        if updated_rows:
            self._update_sql_context(user_id, updated_rows)

        resp = f"✅ Updated {_label} to {_new_disp}."
        self._store_conversation(user_id, message, resp)
        return {"operation": "modify_success", "response": resp, "trigger_invoice": False, "invoice_data": {}}

    def _update_sql_context(self, user_id: str, rows: List[Dict]):
        """Store first result row for follow-up questions (same shape as USCF context)."""
        if not rows:
            return
        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
        ctx["last_row_data"] = dict(rows[0])
        ctx["last_operation"] = "query"
        self.memory.update_user_memory(user_id, {"uscf_context": ctx})

    def _build_uscf_context(self, user_id: str, conversation_history: List[Dict]) -> Optional[Dict]:
        """Build context for USCF parser (helps resolve 'it', 'that', 'update it')."""
        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
        # Extract info from recent assistant messages (dates, clients mentioned)
        if conversation_history:
            for msg in reversed(conversation_history[-4:]):
                if msg.get("role") == "assistant":
                    content = msg.get("content", "")
                    # Look for dates like "04 Apr 2025" or "2025-04-04"
                    import re
                    date_match = re.search(r"(\d{1,2}\s+\w+\s+\d{4}|\d{4}-\d{2}-\d{2})", content)
                    if date_match and not ctx.get("last_result_date"):
                        ctx["last_result_date"] = date_match.group(1)
                    break
        return ctx if ctx else None

    def _is_followup_field_request(self, message: str, columns: List[str]) -> Optional[str]:
        """
        Check if message is a follow-up request for a specific field from last row.
        Returns the requested field keyword if detected, None otherwise.
        """
        msg_lower = message.lower().strip()

        # Mutation phrases are never field-read follow-ups — they must reach the UPDATE pipeline
        _mutation_verbs = (
            "mark ", "set ", "update ", "change ", "edit ", "modify ",
            "mark as", "set as", "change to", "update to",
        )
        if any(msg_lower.startswith(v) or f" {v}" in f" {msg_lower}" for v in _mutation_verbs):
            return None

        # Explicit list/show queries with a "for <client>" scope are new queries,
        # not follow-up field reads (e.g. "show jobs for nike", "list jobs for X").
        _query_verbs = ("show ", "list ", "get ", "find ", "search ", "fetch ", "give me ", "display ")
        if any(msg_lower.startswith(v) for v in _query_verbs) and " for " in msg_lower:
            return None

        # Common follow-up patterns that indicate user wants info from previous result
        followup_patterns = [
            "and the", "what about", "what's the", "what is the", "how about",
            "the ", "show me the", "tell me the", "give me the", "what was the",
            "what's", "whats",
        ]
        
        # Messages with aggregation, time ranges, or multi-entity scope are standalone queries, not follow-ups
        standalone_indicators = [
            "total earning", "total fee", "total billing", "total income", "total revenue",
            "sum of", "how many",
            "last quarter", "this quarter", "last month", "this month", "last year", "this year",
            "which client", "what client", "which brand", "all client", "all brand",
            "have paid", "haven't paid", "not paid", "unpaid clients", "pending payment",
            "so far", "overall", "across",
        ]
        if any(ind in msg_lower for ind in standalone_indicators):
            return None

        # Month names anywhere in the message indicate a time-scoped query, not a field read
        _MONTH_TOKENS = {
            "january", "february", "march", "april", "may", "june", "july",
            "august", "september", "october", "november", "december",
            "jan", "feb", "mar", "apr", "jun", "jul", "aug", "sep", "sept", "oct", "nov", "dec",
        }
        msg_tokens = set(re.findall(r"[a-z]+", msg_lower))
        if msg_tokens & _MONTH_TOKENS:
            return None

        # Plural "jobs"/"invoices"/"clients" with a scope word means a multi-record query
        if re.search(r"\b(jobs|invoices|clients|brands|payments|records|entries)\b", msg_lower) and \
           re.search(r"\b(in|for|from|of|on|with|by|my|all)\b", msg_lower):
            return None

        # Explicit "all <entity>" / "list/show all" phrasing → multi-record listing
        if re.search(r"\ball\b.*\b(jobs|invoices|clients|brands|payments|records|entries)\b", msg_lower):
            return None

        # Check if message matches follow-up pattern or is a short question
        is_followup = (
            any(msg_lower.startswith(p) for p in followup_patterns) or
            msg_lower.endswith("?") or
            len(msg_lower.split()) <= 4  # Short messages like "language?" or "brand name"
        )

        if not is_followup:
            return None
        
        # Comprehensive field aliases mapping
        field_aliases = {
            "brand": ["brand", "brand_name", "brand name", "brandname"],
            "client": ["client", "client_name", "client name", "clientname", "company"],
            "amount": ["amount", "fees", "fee", "billing", "payment", "cost", "price", "total", "value"],
            "paid": ["paid", "payment_status", "status", "payment status", "ispaid"],
            "date": ["date", "job_date", "job date", "jobdate", "when", "day"],
            "job": ["job", "job_name", "job name", "jobname", "work", "gig", "project", "task"],
            "notes": ["notes", "note", "description", "details", "info", "about"],
            "language": ["language", "lang", "languages"],
            "location": ["location", "place", "city", "venue", "where"],
            "contact": ["contact", "phone", "email", "poc", "person"],
            "production": ["production", "production_house", "production house", "productionhouse", "house"],
            "invoice": ["invoice", "invoice_number", "invoice number", "invoicenumber", "bill"],
            "due": ["due", "due_date", "due date", "duedate", "deadline"],
        }
        
        # First check for exact column match
        for col in columns:
            col_lower = col.lower().replace("_", " ").replace("-", " ")
            col_variants = [col_lower, col_lower.replace(" ", "")]
            for variant in col_variants:
                if variant in msg_lower:
                    return col
        
        # Then check aliases
        for canonical, aliases in field_aliases.items():
            for alias in aliases:
                if alias in msg_lower:
                    return canonical  # Return the canonical name, we'll match it to columns later
        
        return None

    def _try_answer_from_context(self, user_id: str, message: str, columns: List[str]) -> Optional[str]:
        """
        Try to answer follow-up question directly from stored last_row_data.
        Returns factual answer string if possible.
        If we don't find the requested field in context, we now allow a fresh query.
        """
        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
        last_row_data = ctx.get("last_row_data")
        
        if not last_row_data:
            logger.info("[FOLLOWUP] No last_row_data in context - allowing new query")
            return None

        # If the user is asking about a past/historical value, skip the short-circuit
        # so the full-row query path runs and Gemini can read the notes change history.
        if self.gemini.is_history_question(message):
            logger.info("[FOLLOWUP] History question detected by AI — skipping short-circuit, using full row")
            return None

        # Check if this is a follow-up field request
        requested_field = self._is_followup_field_request(message, columns)
        if not requested_field:
            logger.info("[FOLLOWUP] Not a follow-up field request - allowing new query")
            return None
        
        logger.info(f"[FOLLOWUP] Looking for field '{requested_field}' in stored row with keys: {list(last_row_data.keys())}")
        
        # Comprehensive alias mapping for field lookup
        field_aliases = {
            "brand": ["brand", "brand_name", "brandname"],
            "client": ["client", "client_name", "clientname", "company"],
            "amount": ["amount", "fees", "fee", "billing", "total", "cost", "price"],
            "paid": ["paid", "payment_status", "status", "ispaid"],
            "date": ["date", "job_date", "jobdate"],
            "job": ["job", "job_name", "jobname", "work", "project", "task"],
            "notes": ["notes", "note", "description", "details", "about"],
            "language": ["language", "lang", "languages"],
            "location": ["location", "place", "city", "venue"],
            "contact": ["contact", "phone", "email", "poc"],
            "production": ["production", "production_house", "productionhouse", "house"],
            "invoice": ["invoice", "invoice_number", "invoicenumber", "bill"],
            "due": ["due", "due_date", "duedate", "deadline"],
        }
        
        # Get all aliases for the requested field
        search_terms = [requested_field.lower()]
        for canonical, aliases in field_aliases.items():
            if requested_field.lower() == canonical or requested_field.lower() in aliases:
                search_terms = aliases + [canonical]
                break
        
        # Try to find the field value in last_row_data
        value = None
        matched_col = None
        
        for col, val in last_row_data.items():
            col_lower = col.lower().replace("_", "").replace(" ", "")
            col_lower_spaced = col.lower().replace("_", " ")
            
            for term in search_terms:
                term_clean = term.replace("_", "").replace(" ", "")
                if (col_lower == term_clean or 
                    term_clean in col_lower or 
                    col_lower in term_clean or
                    term in col_lower_spaced):
                    value = val
                    matched_col = col
                    break
            if value is not None:
                break
        
        if value is None or (isinstance(value, str) and not value.strip()):
            available_fields = ", ".join(list(last_row_data.keys())[:8])
            logger.info(f"[FOLLOWUP] Field '{requested_field}' not found in stored row. Available: {available_fields}")
            logger.info("[FOLLOWUP] Falling back to a new query for this follow-up.")
            # Allow the main flow to run a new query instead of forcing a 'not found' reply
            return None
        
        logger.info(f"[FOLLOWUP] Serving field from stored row without DB call: {matched_col} = {value}")
        payload = build_field_answer_payload(matched_col, value, last_row_data)
        _hist = self.memory.get_conversation_history(user_id)
        response = self.gemini.synthesize_response(payload, message, conversation_history=_hist)
        if response and response.strip():
            return response
        # Fallback if synthesis fails: minimal natural phrasing (no raw field:value)
        if isinstance(value, (int, float)):
            col_lower = matched_col.lower() if matched_col else ""
            if any(term in col_lower for term in ["amount", "fee", "billing", "cost", "price", "total"]):
                return f"The amount was ₹{value:,.0f}."
            return f"The value is {value}."
        return f"That was {value}."

    def _update_uscf_context(self, user_id: str, cmd: Dict, result: Dict):
        """Update context after command execution for future reference resolution."""
        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
        
        # Only update context if we got successful results with matched rows
        matched_rows = result.get("count", 0)
        rows = result.get("rows", [])
        full_rows = result.get("_full_rows", [])  # Full rows for context (not filtered by return_fields)
        
        if matched_rows == 0 and not rows and not full_rows:
            # Don't store context for empty results
            logger.info("[CONTEXT] No matched rows - not updating context")
            return
        
        filters = cmd.get("filters", {})
        # Store filters for "update it" type references
        if filters:
            ctx["current_filters"] = filters
        
        # Store date from result
        if result.get("value_type") == "date" and result.get("value"):
            ctx["last_result_date"] = result["value"]
        
        # Store operation type
        ctx["last_operation"] = cmd.get("operation")
        
        # Store FULL row data for follow-up questions (prefer _full_rows over rows)
        # This ensures we have ALL columns, not just return_fields
        source_rows = full_rows if full_rows else rows
        if source_rows and len(source_rows) > 0:
            last_row = source_rows[0]
            # Store the ENTIRE row, excluding only internal keys
            ctx["last_row_data"] = {k: v for k, v in last_row.items() if not str(k).startswith("_")}
            ctx["last_row_id"] = last_row.get("_row")
            all_keys = list(ctx["last_row_data"].keys())
            logger.info(f"[CONTEXT] Stored full row with keys: {all_keys}")
            logger.info(f"[CONTEXT] last_row_id={ctx.get('last_row_id')}, total_fields={len(all_keys)}")
        
        self.memory.update_user_memory(user_id, {"uscf_context": ctx})

    def _handle_form_step(self, user_id: str, message: str) -> Dict:
        """Handle smart capture confirmation, missing fields, or edit flow."""
        from datetime import datetime, timedelta
        form = self.memory.get_form_state(user_id)
        if not form:
            return None

        # ── Staleness check: auto-cancel forms older than 30 minutes ──────────
        created_at_str = form.get("created_at")
        if created_at_str:
            try:
                age = datetime.now() - datetime.fromisoformat(created_at_str)
                if age > timedelta(minutes=30):
                    self.memory.cancel_form(user_id)
                    logger.info(f"[FORM] Auto-cancelled stale form for {user_id} (age={int(age.total_seconds()//60)} min)")
                    return None  # Let the message be processed normally
            except (ValueError, TypeError):
                pass  # Malformed timestamp — treat as fresh

        # ── Escape: new job entry (+...) should cancel the old form ───────────
        if message.strip().startswith("+") and len(message.strip()) > 2:
            self.memory.cancel_form(user_id)
            logger.info(f"[FORM] Cancelled stale form — new job entry received for {user_id}")
            return None  # Fall through to smart capture

        # ── Escape: obvious new intent (question words, commands) ─────────────
        _msg_lower = message.strip().lower()
        _new_intent_starts = (
            "show ", "list ", "what ", "how ", "when ", "which ", "who ",
            "mark ", "set ", "update ", "delete ", "send ", "generate ",
            "hi", "hello", "hey", "good morning", "good afternoon", "good evening",
        )
        if any(_msg_lower == w.rstrip() or _msg_lower.startswith(w) for w in _new_intent_starts):
            self.memory.cancel_form(user_id)
            logger.info(f"[FORM] Cancelled stale form — new intent detected for {user_id}: '{message[:50]}'")
            return None

        # ── Cancel if user explicitly says cancel/stop ────────────────────────
        if _msg_lower in ("cancel", "stop", "nevermind", "abort", "exit"):
            self.memory.cancel_form(user_id)
            response = "No problem, cancelled. Let me know if you need anything else."
            self._store_conversation(user_id, message, response)
            return {"operation": "form_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        form_type = form.get("form_type", "smart_capture")

        # --- Smart Capture: awaiting confirmation ---
        if form_type == "smart_capture_confirm":
            return self._handle_smart_capture_confirm(user_id, message, form)

        # --- Smart Capture: awaiting missing fields ---
        if form_type == "smart_capture_missing":
            return self._handle_smart_capture_missing(user_id, message, form)

        # Fallback: cancel unknown form
        self.memory.cancel_form(user_id)
        return None

    def _handle_smart_capture_confirm(self, user_id: str, message: str, form: Dict) -> Dict:
        """Handle Yes/Edit response to smart capture confirmation."""
        msg = message.strip().lower()
        extracted = form.get("values", {})

        # If user reply contains an email or looks like POC info (and POC fields are
        # missing), try to extract those before/instead of treating as Yes/Edit.
        _has_at = "@" in message
        _missing_poc = not extracted.get("poc_name") or not extracted.get("poc_email")
        if _has_at and _missing_poc:
            new_data = self.gemini.extract_job_fields(message) or {}
            # Validate any email Gemini extracted
            if new_data.get("poc_email") and not self._is_valid_email(new_data.get("poc_email")):
                new_data["poc_email"] = None
            updated = False
            for k in ("poc_name", "poc_email"):
                if not extracted.get(k) and new_data.get(k):
                    extracted[k] = new_data[k]
                    updated = True
            # Fallback: regex-extract email if Gemini missed it — but only accept if valid
            if not extracted.get("poc_email"):
                _m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", message)
                if _m and self._is_valid_email(_m.group(0)):
                    extracted["poc_email"] = _m.group(0)
                    updated = True

            # If user clearly tried to send an email but it's malformed, ask again
            if not extracted.get("poc_email"):
                _candidate = re.search(r"\S*@\S*", message)
                _bad = _candidate.group(0) if _candidate else message.strip()
                response = (
                    f"'{_bad}' doesn't look like a valid email. "
                    f"Please send a valid email (e.g. name@company.com) or reply 'Yes' to save without one."
                )
                form["values"] = extracted
                self.memory.start_form(user_id, [], form_override=form)
                self._store_conversation(user_id, message, response)
                return {"operation": "smart_capture_invalid_email", "response": response, "trigger_invoice": False, "invoice_data": {}}

            if updated:
                # Show the updated card and let the user re-confirm
                return self._show_smart_capture_confirmation(user_id, extracted)

        if msg in ("yes", "y", "save", "confirm", "done", "ok", "okay", "sure"):
            # Save to database — POC fields are optional
            return self._save_smart_capture_job(user_id, extracted)

        elif msg in ("no", "nope", "nah", "cancel", "nevermind", "nvm", "abort"):
            # User declines — cancel cleanly without asking them to re-send info
            self.memory.cancel_form(user_id)
            response = "No problem, cancelled. Let me know if you need anything else."
            self._store_conversation(user_id, message, response)
            return {"operation": "smart_capture_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        elif msg in ("edit", "change", "modify", "fix"):
            response = (
                "No problem! Send the corrected job info in one message.\n\n"
                "Example:\n"
                "Brand: Bridgestone\n"
                "Date: 10 Feb\n"
                "Job: Master film 30 sec\n"
                "Client: The Good Take\n"
                "Fees: 25k"
            )
            self.memory.cancel_form(user_id)
            self.memory.update_user_memory(user_id, {"awaiting_job_input": True})
            self._store_conversation(user_id, message, response)
            return {"operation": "smart_capture_edit", "response": response, "trigger_invoice": False, "invoice_data": {}}

        else:
            # Track retries — auto-cancel after 2 unrecognised replies
            retry_count = form.get("retry_count", 0) + 1
            form["retry_count"] = retry_count
            self.memory.start_form(user_id, [], form_override=form)
            if retry_count >= 2:
                self.memory.cancel_form(user_id)
                response = "I'll cancel that for now. Let me know if you'd like to try again."
                self._store_conversation(user_id, message, response)
                return {"operation": "form_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}
            response = "Please reply 'Yes' to save or 'Edit' to make changes."
            self._store_conversation(user_id, message, response)
            return {"operation": "smart_capture_confirm_retry", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_smart_capture_missing(self, user_id: str, message: str, form: Dict) -> Dict:
        """Handle response with missing required fields."""
        extracted = form.get("values", {})
        missing = form.get("missing_fields", [])

        # Try to extract fields from the user's response
        new_data = self.gemini.extract_job_fields(message)
        invalid_email_attempt = None
        if new_data:
            # Validate poc_email before accepting
            if new_data.get("poc_email") and not self._is_valid_email(new_data.get("poc_email")):
                invalid_email_attempt = new_data.get("poc_email")
                new_data["poc_email"] = None
            for k, v in new_data.items():
                if v is not None:
                    extracted[k] = v

        # Also catch a bare email-looking token the user typed even if Gemini missed it
        if not invalid_email_attempt:
            _bare = message.strip()
            if "@" in _bare and " " not in _bare and not self._is_valid_email(_bare):
                invalid_email_attempt = _bare

        if invalid_email_attempt:
            response = (
                f"'{invalid_email_attempt}' doesn't look like a valid email. "
                f"Please send a valid email (e.g. name@company.com). "
                f"It's required so we can email the invoice later."
            )
            form["values"] = extracted
            self.memory.start_form(user_id, [], form_override=form)
            self._store_conversation(user_id, message, response)
            return {"operation": "smart_capture_invalid_email", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Check if still missing required fields
        still_missing = [f for f in missing if not extracted.get(f)]
        if still_missing:
            field_labels = {
                "brand_name": "Brand",
                "job_date": "Date",
                "job_description_details": "Job details",
                "poc_name": "POC name",
                "poc_email": "POC email (required for invoicing)",
            }
            missing_str = ", ".join(field_labels.get(f, f) for f in still_missing)
            response = f"I still need: {missing_str}. Please provide them."
            # Update form with new values
            form["values"] = extracted
            form["missing_fields"] = still_missing
            self.memory.start_form(user_id, [], form_override=form)
            self._store_conversation(user_id, message, response)
            return {"operation": "smart_capture_missing_retry", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # All required fields present - show confirmation
        return self._show_smart_capture_confirmation(user_id, extracted)

    def _save_smart_capture_job(self, user_id: str, extracted: Dict) -> Dict:
        """Save the extracted job to database."""
        self.memory.cancel_form(user_id)

        # Map extracted fields to job_entries columns
        record = {"user_id": user_id}
        field_map = {
            "job_date": "job_date",
            "brand_name": "brand_name",
            "client_name": "client_name",
            "job_description_details": "job_description_details",
            "fees": "fees",
            "paid": "paid",
            "poc_name": "poc_name",
            "poc_email": "poc_email",
            "notes": "notes",
        }
        for src, dst in field_map.items():
            val = extracted.get(src)
            if val is not None:
                record[dst] = val

        insert_result = self.supabase.insert_job_entry(record)
        if insert_result.get("ok"):
            brand = extracted.get("brand_name", "")
            client = extracted.get("client_name", "")
            display_name = brand or client or "Job"
            response = f"Job saved! ✅ {display_name} has been added to your records."

            # Check if user had a compound intent (e.g. "add job and send invoice")
            user_mem = self.memory.get_user_memory(user_id)
            suggested_next = user_mem.get("suggested_next_action")
            if suggested_next:
                # Keep suggested_next_action in memory so the handler can use it
                self.memory.update_user_memory(user_id, {"awaiting_compound_response": True})
                response += f"\n\nYou also mentioned: \"{suggested_next}\"\nWant me to do that now? (Yes / No)"
                logger.info(f"[COMPOUND] Suggesting next action after job save: '{suggested_next}'")

            # Store last job context so user can reference "this job" in follow-up
            self.memory.update_user_memory(user_id, {
                "last_saved_job": {
                    "brand_name": brand,
                    "client_name": client,
                    "job_date": extracted.get("job_date"),
                    "job_description_details": extracted.get("job_description_details"),
                    "fees": extracted.get("fees"),
                    "db_client_name": record.get("client_name"),  # what's actually in client_name col
                }
            })
        else:
            logger.error(f"[SMART_CAPTURE] Insert failed: {insert_result.get('error')}")
            response = "I couldn't save the job. Please try again."
        # Build a summary of what was saved for conversation context
        summary = ", ".join(f"{k}: {v}" for k, v in extracted.items() if v is not None)
        self._store_conversation(user_id, f"Save job: {summary}", response)
        return {"operation": "form_complete", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _show_smart_capture_confirmation(self, user_id: str, extracted: Dict) -> Dict:
        """Show confirmation message and wait for Yes/Edit."""
        lines = ["Got it 👍\n"]
        field_labels = [
            ("brand_name", "Brand"),
            ("client_name", "Client"),
            ("job_date", "Date"),
            ("job_description_details", "Details"),
            ("fees", "Fees"),
            ("paid", "Payment"),
            ("poc_name", "POC name"),
            ("poc_email", "POC email"),
            ("notes", "Notes"),
        ]
        for key, label in field_labels:
            val = extracted.get(key)
            if val is not None:
                if key == "fees":
                    val = f"₹{val:,}" if isinstance(val, (int, float)) else val
                elif key == "paid":
                    val = "Paid ✅" if str(val).lower() in ("true", "yes", "1") else "Unpaid"
                lines.append(f"{label}: {val}")

        lines.append("\nSave this job? (Yes / Edit)")
        response = "\n".join(lines)

        # Store in form state for confirmation
        form_data = {
            "form_type": "smart_capture_confirm",
            "values": extracted,
            "fields": [],
            "step": 0,
        }
        self.memory.start_form(user_id, [], form_override=form_data)
        # Store the extracted details as user message for context
        summary = ", ".join(f"{k}: {v}" for k, v in extracted.items() if v is not None)
        self._store_conversation(user_id, f"Job details: {summary}", response)
        return {"operation": "smart_capture_confirm", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _start_smart_capture(self, user_id: str, message: str) -> Dict:
        """
        AI Smart Capture: extract job fields from natural language.
        If message only contains trigger words, prompt for details.
        If message contains job data, extract and confirm.
        """
        import re
        # Strip all job-intent phrases to isolate actual job content
        content = message.strip()
        # Remove leading "+" 
        if content.startswith("+"):
            content = content[1:].strip()
        # Remove common intent phrases (anywhere in the message)
        intent_phrases = [
            r"i\s+want\s+to\s+", r"i\'?d\s+like\s+to\s+", r"can\s+you\s+",
            r"please\s+", r"let\s*'?s\s+",
            r"add\s+(?:a\s+)?(?:new\s+)?job\s*", r"new\s+job\s*",
            r"log\s+(?:a\s+)?job\s*", r"record\s+(?:a\s+)?job\s*",
            r"create\s+(?:a\s+)?(?:new\s+)?(?:job|entry)\s*",
        ]
        content_clean = content
        for pat in intent_phrases:
            content_clean = re.sub(pat, "", content_clean, flags=re.IGNORECASE).strip()

        logger.info(f"[SMART_CAPTURE] Original='{message}' -> Cleaned='{content_clean}'")
        # If no meaningful content remains, prompt for details
        if not content_clean or len(content_clean) < 3:
            self.memory.update_user_memory(user_id, {"awaiting_job_input": True})
            response = (
                "Describe the job in one message.\n\n"
                "Example:\n"
                "Brand: Bridgestone\n"
                "Date: 10 Feb\n"
                "Job: Master film 30 sec + 4 cutdowns\n"
                "Client: The Good Take\n"
                "Fees: 25k\n"
                "POC: Rohan Mehta\n"
                "POC Email: rohan@thegoodtake.com"
            )
            self._store_conversation(user_id, message, response)
            return {"operation": "smart_capture_prompt", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Content available - extract fields
        return self._extract_and_confirm(user_id, content_clean)

    def _extract_and_confirm(self, user_id: str, content: str) -> Dict:
        """Extract fields from content and show confirmation or ask for missing."""
        self.memory.update_user_memory(user_id, {"awaiting_job_input": False})
        logger.info(f"[SMART_CAPTURE] Extracting fields from: '{content[:200]}'")
        extracted = self.gemini.extract_job_fields(content)
        logger.info(f"[SMART_CAPTURE] Result: {extracted}")

        # Validate poc_email format — drop if malformed so we re-prompt.
        if extracted and extracted.get("poc_email") and not self._is_valid_email(extracted.get("poc_email")):
            logger.info(f"[SMART_CAPTURE] Dropping invalid poc_email: {extracted.get('poc_email')!r}")
            extracted["poc_email"] = None

        # Treat all-null extraction as failure
        if extracted and all(v is None for v in extracted.values()):
            extracted = None

        if not extracted:
            # No fields extracted — user likely just expressed intent ("add a job")
            # without providing actual data. Show a friendly prompt, not an error.
            self.memory.update_user_memory(user_id, {"awaiting_job_input": True})
            response = (
                "Describe the job in one message.\n\n"
                "Example:\n"
                "Brand: Bridgestone\n"
                "Date: 10 Feb\n"
                "Job: Master film 30 sec + 4 cutdowns\n"
                "Client: The Good Take\n"
                "Fees: 25k\n"
                "POC: Rohan Mehta\n"
                "POC Email: rohan@thegoodtake.com"
            )
            self._store_conversation(user_id, content, response)
            return {"operation": "smart_capture_prompt", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Default job_date to today if not extracted
        if not extracted.get("job_date"):
            from datetime import date as _date
            extracted["job_date"] = _date.today().isoformat()

        # Check required fields — brand, date, and description are mandatory.
        # POC fields are optional: user can add them later or skip if not needed.
        required = ["brand_name", "job_description_details"]
        missing = [f for f in required if not extracted.get(f)]

        if missing:
            field_labels = {
                "brand_name": "Brand",
                "job_date": "Date",
                "job_description_details": "Job details",
                "poc_name": "POC name",
                "poc_email": "POC email",
            }
            missing_str = ", ".join(field_labels.get(f, f) for f in missing)

            # Show what we got so far + ask for missing
            lines = ["I got some of the details:\n"]
            field_display = [
                ("brand_name", "Brand"), ("client_name", "Client"), ("job_date", "Date"),
                ("job_description_details", "Details"), ("fees", "Fees"),
                ("poc_name", "POC name"), ("poc_email", "POC email"), ("notes", "Notes"),
            ]
            for key, label in field_display:
                val = extracted.get(key)
                if val is not None:
                    if key == "fees":
                        val = f"₹{val:,}" if isinstance(val, (int, float)) else val
                    lines.append(f"{label}: {val}")

            lines.append(f"\nI still need: {missing_str}")
            lines.append("Please send the missing info.")
            response = "\n".join(lines)

            form_data = {
                "form_type": "smart_capture_missing",
                "values": extracted,
                "missing_fields": missing,
                "fields": [],
                "step": 0,
            }
            self.memory.start_form(user_id, [], form_override=form_data)
            self._store_conversation(user_id, content, response)
            return {"operation": "smart_capture_missing", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # All required fields present - show confirmation
        return self._show_smart_capture_confirmation(user_id, extracted)

    def _handle_invoice_month_reply(self, user_id: str, message: str, user_mem: dict, data_user_id: str, conversation_history: list) -> Dict:
        """Handle user providing a month after bot asked 'Which month?' for invoice."""
        import re as _re
        from datetime import datetime
        # Clear the awaiting state
        client_name = user_mem.get("pending_invoice_client", "")
        send_email = user_mem.get("pending_invoice_send_email", False)
        self.memory.update_user_memory(user_id, {
            "awaiting_invoice_month": False,
            "pending_invoice_client": None,
            "pending_invoice_send_email": None,
        })

        # Extract month from user reply
        month_name = None
        msg_lower = message.strip().lower()
        _MONTHS = {
            "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
            "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7,
            "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
        }
        month_num = None
        for name, num in _MONTHS.items():
            if name in msg_lower:
                month_name = name.capitalize()
                month_num = num
                break

        if not month_num:
            response = f"I couldn't detect a month from your reply. Please say something like: 'March' or 'March 2025'."
            self._store_conversation(user_id, message, response)
            # Re-set awaiting state
            self.memory.update_user_memory(user_id, {
                "awaiting_invoice_month": True,
                "pending_invoice_client": client_name,
                "pending_invoice_send_email": send_email,
            })
            return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Extract year from user reply (e.g. "April 2025", "March 2024")
        # Look for 4-digit year in the message
        year_match = _re.search(r'\b(20\d{2})\b', message)
        if year_match:
            year_val = int(year_match.group(1))
        else:
            year_val = datetime.now().year

        # Build synthetic message preserving ALL extracted entities
        year_part = f" {year_val}" if year_match else ""
        synthetic_msg = f"Generate invoice for {client_name} for {month_name}{year_part}"
        if send_email:
            synthetic_msg = f"Send invoice for {client_name} for {month_name}{year_part} over email"
        logger.info(f"[INVOICE_FOLLOWUP] Resuming invoice flow: client={client_name}, month={month_name}, year={year_val}, synthetic='{synthetic_msg}'")
        # Re-enter process_request with the full synthetic message
        return self.process_request(user_id=user_id, message=synthetic_msg)

    def _handle_send_confirmation(self, user_id: str, message: str) -> Dict:
        """Handle user confirming/declining sending invoice to client email."""
        pending = self.memory.get_user_memory(user_id).get("pending_send_invoice", {})
        msg_lower = message.strip().lower()

        _YES = {"yes", "y", "yeah", "yep", "sure", "ok", "okay", "go ahead", "send it", "do it", "confirm", "yes please"}
        _NO = {"no", "n", "nope", "nah", "skip", "cancel", "not now", "later", "don't send", "dont send"}

        # Detect invoice feedback (e.g. "Invoice is missing client billing info")
        # instead of treating it as a decline
        _FEEDBACK_WORDS = ["missing", "wrong", "incorrect", "update", "change",
                           "fix", "edit", "add", "not correct", "doesn't have",
                           "not showing", "no client", "no billing", "no address"]
        is_feedback = any(w in msg_lower for w in _FEEDBACK_WORDS)

        if is_feedback:
            # Clear confirmation state but keep the cached invoice for the feedback handler
            self.memory.update_user_memory(user_id, {
                "awaiting_send_confirmation": False,
                "pending_send_invoice": None,
            })
            client_name = pending.get("client_name", "Client")
            response = (
                f"Got it — you'd like to update the invoice for {client_name} before sending. "
                f"Here's what you can set:\n\n"
                f"1. Your name/title/address/email on the invoice header — say: "
                f"\"Update invoice profile\"\n"
                f"2. Client billing details (billing name, address, GST) — say: "
                f"\"Update client billing for {client_name}\"\n\n"
                f"After updating, say \"Regenerate invoice for {client_name}\" "
                f"and I'll create a fresh PDF with the new details."
            )
            self._store_conversation(user_id, message, response)
            return {"operation": "invoice_feedback", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Clear the awaiting flag
        self.memory.update_user_memory(user_id, {
            "awaiting_send_confirmation": False,
            "pending_send_invoice": None,
        })

        if msg_lower in _YES:
            poc_email = pending.get("poc_email", "")
            client_name = pending.get("client_name", "Client")
            month_display = pending.get("month", "Request")
            year_val = pending.get("year")
            row_ids = pending.get("row_ids", [])

            if not poc_email:
                response = "I don't have the client email anymore. Please try again with 'Send invoice for ...'."
                self._store_conversation(user_id, message, response)
                return {"operation": "send_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # Find the generated PDF
            safe_client = client_name.replace(" ", "_")
            safe_month = month_display.replace(" ", "_")
            pdf_path = os.path.join("output", f"Invoice_{safe_client}_{safe_month}.pdf")

            if not os.path.exists(pdf_path):
                response = "I can't find the generated PDF. Please regenerate the invoice first."
                self._store_conversation(user_id, message, response)
                return {"operation": "send_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # Extract poc_name from pending data rows
            _poc_name = pending.get("poc_name") or ""
            _invoicer_name = pending.get("invoicer_name") or ""
            ok = self.email.send_invoice_email(
                to_email=poc_email,
                client_name=client_name,
                month=month_display,
                year=year_val,
                pdf_path=pdf_path,
                poc_name=_poc_name or None,
                invoicer_name=_invoicer_name or None,
                cc=self._get_user_invoice_email(user_id),
            )
            if ok:
                # Mark rows as actually-emailed:
                #   invoice_date  = "PDF exists" (CURRENT_DATE = today)
                #   bill_sent     = 'Yes' (text flag)
                #   bill_sent_at  = NOW() (precise timestamp — answers
                #                  "when was the invoice sent?")
                if row_ids:
                    ids_str = ",".join(f"'{rid}'" for rid in row_ids)
                    self.supabase.execute_sql(
                        f"UPDATE public.job_entries SET invoice_date = CURRENT_DATE, "
                        f"bill_sent = 'Yes', bill_sent_at = NOW() WHERE id IN ({ids_str})"
                    )
                    logger.info(f"[INVOICE] Marked bill_sent + bill_sent_at + invoice_date for {len(row_ids)} row(s)")
                response = f"Invoice has been sent to {poc_email}. ✅"
            else:
                response = "I couldn't send the invoice email. Please check the email configuration and try again."
            self._store_conversation(user_id, message, response)
            return {"operation": "send_confirmed", "response": response, "trigger_invoice": False, "invoice_data": {}}
        else:
            response = "👍 Got it, invoice not sent. You can say 'Send invoice for ...' anytime to email it."
            self._store_conversation(user_id, message, response)
            return {"operation": "send_declined", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_poc_email_response(self, user_id: str, message: str) -> Dict:
        """Handle user providing a client POC email after invoice generation."""
        import re
        user_mem = self.memory.get_user_memory(user_id)
        pending = user_mem.get("pending_send_invoice", {})

        # Clear awaiting state
        self.memory.update_user_memory(user_id, {"awaiting_poc_email": False})

        # Allow cancel
        if message.strip().lower() in ("cancel", "skip", "no", "nevermind"):
            self.memory.update_user_memory(user_id, {"pending_send_invoice": None})
            response = "No problem, skipped. You can add the client email later."
            self._store_conversation(user_id, message, response)
            return {"operation": "poc_email_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Validate email format
        email = message.strip()
        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
            # Not a valid email - re-prompt
            self.memory.update_user_memory(user_id, {"awaiting_poc_email": True})
            response = "That doesn't look like a valid email. Please send the client's email address (e.g. client@agency.com) or type 'skip'."
            self._store_conversation(user_id, message, response)
            return {"operation": "poc_email_retry", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Read client info from pending_send_invoice (the actual stored data)
        client_name = pending.get("client_name", "")
        month = pending.get("month", "")
        year = pending.get("year")
        row_ids = pending.get("row_ids", [])

        if not client_name:
            logger.warning(f"[POC] No client_name in pending_send_invoice for user {user_id}")

        # Save POC email to job entries for this client
        result = self.supabase.update_poc_email_for_client(user_id, client_name, email)

        if result.get("ok"):
            updated = result.get("updated", 0)

            # Try to send the invoice email now using the already-generated PDF
            safe_client = client_name.replace(" ", "_")
            safe_month = month.replace(" ", "_")
            pdf_path = os.path.join("output", f"Invoice_{safe_client}_{safe_month}.pdf")

            email_sent = False
            if os.path.exists(pdf_path):
                try:
                    _poc_name = pending.get("poc_name") or ""
                    _invoicer_name = pending.get("invoicer_name") or ""
                    ok = self.email.send_invoice_email(
                        to_email=email,
                        client_name=client_name,
                        month=month,
                        year=year or 2026,
                        pdf_path=pdf_path,
                        poc_name=_poc_name or None,
                        invoicer_name=_invoicer_name or None,
                        cc=self._get_user_invoice_email(user_id),
                    )
                    email_sent = ok
                except Exception as e:
                    logger.error(f"[POC] Failed to send invoice email after saving POC: {e}")
            else:
                logger.warning(f"[POC] PDF not found at {pdf_path} — cannot auto-send invoice")

            if email_sent:
                response = f"Saved! Email {email} has been added for {client_name} and the invoice has been sent. ✅"
                # Mark rows as actually-emailed (see _handle_send_confirmation for
                # why we set bill_sent + bill_sent_at alongside invoice_date).
                if row_ids:
                    try:
                        ids_str = ",".join(f"'{rid}'" for rid in row_ids)
                        self.supabase.execute_sql(
                            f"UPDATE public.job_entries SET invoice_date = CURRENT_DATE, "
                            f"bill_sent = 'Yes', bill_sent_at = NOW() WHERE id IN ({ids_str})"
                        )
                        logger.info(f"[INVOICE] Marked bill_sent + bill_sent_at + invoice_date for {len(row_ids)} row(s) after POC email save")
                    except Exception as e:
                        logger.warning(f"[INVOICE] Failed to update invoice_date after POC save: {e}")
            else:
                response = f"Saved! Email {email} has been added for {client_name} ({updated} job{'s' if updated != 1 else ''} updated)."
                if not email_sent and os.path.exists(pdf_path):
                    response += "\n\nI couldn't send the invoice email automatically. You can try 'Send invoice for " + client_name + "' again."
        else:
            response = f"I couldn't save the email: {result.get('error', 'Unknown error')}. Please try again."

        # Clean up memory
        self.memory.update_user_memory(user_id, {"pending_send_invoice": None})
        self._store_conversation(user_id, message, response)
        return {"operation": "poc_email_saved", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_client_billing_response(self, user_id: str, message: str) -> Dict:
        """Handle user providing client billing details before invoice generation."""
        user_mem = self.memory.get_user_memory(user_id)
        pending_invoice = user_mem.get("pending_invoice", {})
        client_name = user_mem.get("pending_billing_client", "")
        data_user_id = user_mem.get("pending_billing_user_id", user_id)

        # Clear awaiting state
        self.memory.update_user_memory(user_id, {
            "awaiting_client_billing": False,
            "pending_billing_client": None,
            "pending_billing_user_id": None,
        })

        # Billing details are mandatory — 'cancel' aborts the invoice, otherwise
        # the value is required.
        if message.strip().lower() in ("cancel", "stop", "abort", "nevermind"):
            self.memory.update_user_memory(user_id, {"pending_invoice": None})
            response = "No problem — invoice cancelled. Nothing was generated."
            self._store_conversation(user_id, message, response)
            return {"operation": "invoice_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Parse billing details from the message. Strip a stray leading label the
        # user often types ("Billing info is Spotify India…") so it doesn't get
        # stored and later printed on the invoice (#1).
        from services.invoice_generation_service import _strip_billing_label
        billing_text = _strip_billing_label(message.strip())

        # Save to all matching job_entries for this client
        if client_name:
            safe_client = client_name.replace("'", "''")
            safe_uid = data_user_id.replace("'", "''")
            safe_billing = billing_text.replace("'", "''")
            update_sql = (
                f"UPDATE public.job_entries SET client_billing_details = '{safe_billing}' "
                f"WHERE user_id = '{safe_uid}' "
                f"AND (client_name ILIKE '%{safe_client}%' OR brand_name ILIKE '%{safe_client}%' OR production_house ILIKE '%{safe_client}%') "
                f"AND (\"isDeleted\" IS NOT TRUE)"
            )
            result = self.supabase.execute_sql(update_sql)
            if not result.get("ok") and "production_house" in str(result.get("error", "")):
                update_sql = (
                    f"UPDATE public.job_entries SET client_billing_details = '{safe_billing}' "
                    f"WHERE user_id = '{safe_uid}' AND (client_name ILIKE '%{safe_client}%' OR brand_name ILIKE '%{safe_client}%') "
                    f"AND (\"isDeleted\" IS NOT TRUE)"
                )
                result = self.supabase.execute_sql(update_sql)

            if result.get("ok"):
                logger.info(f"[BILLING] Saved client billing details for {client_name}")
            else:
                logger.warning(f"[BILLING] Failed to save billing details: {result.get('error')}")

        # Re-enter the invoice flow so the gate prompts for the next missing field
        # (or generates once everything is present).
        if pending_invoice:
            return self._resume_invoice_flow(user_id, pending_invoice)

        response = f"Billing details saved for {client_name}."
        self._store_conversation(user_id, message, response)
        return {"operation": "billing_saved", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _persist_invoice_address(self, addr_uid: str, address: str) -> bool:
        """Save the invoicer's business address to profile preferences (merging, so
        name/etc. are preserved). Shared by the gate flow and the 'update my
        address' command."""
        prefs = {}
        existing = self.supabase.get_user_profile(addr_uid)
        if existing.get("ok") and existing.get("data"):
            _p = existing["data"].get("preferences") or {}
            if isinstance(_p, str):
                try:
                    prefs = json.loads(_p)
                except (json.JSONDecodeError, TypeError):
                    prefs = {}
            elif isinstance(_p, dict):
                prefs = _p
        prefs["invoice_address"] = address.strip()
        prefs.pop("invoice_address_skipped", None)
        platform = "telegram" if str(addr_uid).isdigit() else "whatsapp"
        try:
            self.supabase.upsert_user_profile(addr_uid, platform, {"preferences": prefs})
            return True
        except Exception as _e:
            logger.warning(f"[INVOICE_ADDR] Could not save invoice address: {_e}")
            return False

    def _handle_address_update(self, user_id: str, message: str, data_user_id: str) -> Dict:
        """'update my address' / 'my address is X' — let the user set or correct
        their saved business address at any time (not just during an invoice)."""
        # Try to pull the new address inline ("...address to X" / "...address is X").
        m = re.search(r'address\s*(?:to|is|:|=)\s*(.+)', message.strip(), re.IGNORECASE | re.DOTALL)
        inline = (m.group(1).strip() if m else "")
        if inline and len(inline) >= 4 and inline.lower() not in ("wrong", "incorrect"):
            self._persist_invoice_address(data_user_id, inline)
            response = f"Updated ✅ Your business address is now:\n\n{inline}\n\nIt'll appear on your invoices from now on."
            self._store_conversation(user_id, message, response)
            return {"operation": "address_updated", "response": response, "trigger_invoice": False, "invoice_data": {}}
        # No inline address → prompt for it (standalone update, no pending invoice).
        self.memory.update_user_memory(user_id, {
            "awaiting_invoice_address": True,
            "pending_address_user_id": data_user_id,
            "pending_invoice": None,
        })
        response = ("Sure — what's your business address for the invoice header?\n\n"
                    "(multiple lines are fine — send the full new address)")
        self._store_conversation(user_id, message, response)
        return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_invoice_address_response(self, user_id: str, message: str) -> Dict:
        """Store the invoicer's business address (mandatory) in profile preferences,
        then re-enter the invoice flow. 'cancel' aborts the invoice."""
        user_mem = self.memory.get_user_memory(user_id)
        pending_invoice = user_mem.get("pending_invoice", {})
        addr_uid = user_mem.get("pending_address_user_id") or user_id
        self.memory.update_user_memory(user_id, {
            "awaiting_invoice_address": False,
            "pending_address_user_id": None,
        })

        # 'cancel' aborts the invoice (or, for a standalone update, just stops).
        if message.strip().lower() in ("cancel", "stop", "abort", "nevermind"):
            self.memory.update_user_memory(user_id, {"pending_invoice": None})
            response = ("No problem — invoice cancelled. Nothing was generated." if pending_invoice
                        else "Okay, address unchanged.")
            self._store_conversation(user_id, message, response)
            return {"operation": "invoice_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        self._persist_invoice_address(addr_uid, message.strip())

        if pending_invoice:
            self._store_conversation(user_id, message, "Saved — that address will appear on your invoices.")
            return self._resume_invoice_flow(user_id, pending_invoice)
        ack = "Saved — that address will appear on your invoices."
        self._store_conversation(user_id, message, ack)
        return {"operation": "address_saved", "response": ack, "trigger_invoice": False, "invoice_data": {}}

    # ── Mandatory-fields gate for invoices ──────────────────────────────────
    # Every invoice must carry: client billing details, a POC name, a description
    # for each job, the user's bank account, and the user's business address.
    # _invoice_readiness_check returns a prompt for the FIRST missing field (and
    # arms the matching awaiting state + pending_invoice); the field's handler
    # saves it and re-enters the flow, so prompts chain until the invoice is
    # complete — only then does generation/email proceed. GST and POC email stay
    # optional. 'cancel' at any prompt aborts the invoice.
    def _invoice_readiness_check(self, user_id: str, data_user_id: str,
                                 invoice_data: dict, rows: Optional[List[Dict]] = None) -> Optional[Dict]:
        display_client = invoice_data.get("client_name", "Client")
        month_display = invoice_data.get("month", "")
        year_val = invoice_data.get("year")
        bill_number = invoice_data.get("bill_number")

        if rows is None:
            _mn = month_name_to_number(month_display) if month_display and month_display != "Request" else None
            if bill_number:
                _res = self.supabase.fetch_job_entries_for_invoice(client_name="", bill_no=bill_number, user_id=data_user_id)
            else:
                _res = self.supabase.fetch_job_entries_for_invoice(client_name=display_client, month=_mn, year=year_val, user_id=data_user_id)
            rows = _res.get("rows") or []
        if not rows:
            return None  # nothing to invoice — the main flow handles the empty case

        def _prompt(state: dict, text: str) -> Dict:
            patch = {"pending_invoice": invoice_data}
            patch.update(state)
            self.memory.update_user_memory(user_id, patch)
            self._store_conversation(user_id, "", text)
            return {"operation": "ACTION_TRIGGER", "response": text, "trigger_invoice": False, "invoice_data": {}}

        def _present(v) -> bool:
            s = str(v or "").strip()
            return bool(s) and s.lower() != "none"

        # 1. Client billing details
        if not any(_present(r.get("client_billing_details")) for r in rows):
            return _prompt(
                {"awaiting_client_billing": True, "pending_billing_client": display_client,
                 "pending_billing_user_id": data_user_id},
                f"To bill {display_client}, I need their billing details.\n\n"
                "Send the billing name + address (and GST if they have one), e.g.:\n"
                "Spotify India\nLower Parel, Mumbai\nGST: 27ABCDE1234F1Z5\n\n"
                "(or 'cancel' to stop)"
            )

        # 2. POC name
        if not any(_present(r.get("poc_name")) for r in rows):
            return _prompt(
                {"awaiting_poc_name": True, "pending_poc_client": display_client,
                 "pending_poc_user_id": data_user_id,
                 "pending_poc_row_ids": [r["id"] for r in rows if r.get("id")]},
                f"Who should the invoice for {display_client} be addressed to? "
                "(the point-of-contact name — or 'cancel' to stop)"
            )

        # 3. Job description — every line item needs one
        _missing = [r for r in rows if not _present(r.get("job_description_details"))]
        if _missing:
            _r = _missing[0]
            _d = str(_r.get("job_date") or "")[:10]
            _when = f" dated {_d}" if _d else ""
            return _prompt(
                {"awaiting_job_description": True, "pending_jobdesc_row_id": _r.get("id"),
                 "pending_jobdesc_user_id": data_user_id},
                f"One job{_when} for {display_client} has no description.\n\n"
                "What was the work? (e.g. '2 master films, English VO' — or 'cancel' to stop)"
            )

        # 4. Bank account number — the client can't pay without it
        _bank = self.supabase.get_user_bank_details(data_user_id)
        _bd = _bank.get("data") if _bank.get("ok") else None
        if not _bd or not str(_bd.get("bank_account_number") or "").strip():
            return _prompt(
                {"awaiting_bank_details": True},
                "Before I generate it, I need your bank details so the client can pay:\n\n"
                "Account Name: Your Name\nBank Name: HDFC Bank\nAccount Number: 1234567890\n"
                "IFSC: HDFC0001234\nUPI: you@upi (optional)\n\n"
                "(or 'cancel' to stop)"
            )

        # 5. Invoicer business address (mandatory — no skip)
        _prof = self.supabase.get_user_profile(data_user_id)
        _prefs = {}
        if _prof.get("ok") and _prof.get("data"):
            _prefs = _prof["data"].get("preferences") or {}
            if isinstance(_prefs, str):
                try:
                    _prefs = json.loads(_prefs)
                except (json.JSONDecodeError, TypeError):
                    _prefs = {}
        if not (_prefs.get("invoice_address") or "").strip():
            return _prompt(
                {"awaiting_invoice_address": True, "pending_address_user_id": data_user_id},
                "Last thing — what's your business address for the invoice header? "
                "It'll sit under your name.\n\n(multiple lines are fine — or 'cancel' to stop)"
            )

        return None  # everything mandatory is present → proceed

    def _resume_invoice_flow(self, user_id: str, invoice_data: dict) -> Dict:
        """Re-enter the invoice flow after a mandatory field was supplied, so the
        readiness gate re-runs and either prompts for the next field or generates."""
        self.memory.update_user_memory(user_id, {"pending_invoice": None})
        client = invoice_data.get("client_name", "")
        month = invoice_data.get("month", "")
        year = invoice_data.get("year")
        bill = invoice_data.get("bill_number")
        send = invoice_data.get("send_to_client")
        verb = "Send" if send else "Generate"
        if bill:
            synthetic = f"{verb} invoice for bill {bill}"
        else:
            _yr = f" {year}" if year else ""
            _mo = f" for {month}" if month and month != "Request" else ""
            synthetic = f"{verb} invoice for {client}{_mo}{_yr}"
        if send:
            synthetic += " over email"
        logger.info(f"[INVOICE_GATE] Field supplied — re-entering flow: '{synthetic}'")
        return self.process_request(user_id=user_id, message=synthetic)

    def _handle_job_description_response(self, user_id: str, message: str) -> Dict:
        """Save the supplied description to the pending job row, then re-enter the
        invoice flow (which prompts for the next missing field or generates)."""
        user_mem = self.memory.get_user_memory(user_id)
        pending_invoice = user_mem.get("pending_invoice", {})
        row_id = user_mem.get("pending_jobdesc_row_id")
        uid = user_mem.get("pending_jobdesc_user_id", user_id)
        self.memory.update_user_memory(user_id, {
            "awaiting_job_description": False,
            "pending_jobdesc_row_id": None,
            "pending_jobdesc_user_id": None,
        })

        if message.strip().lower() in ("cancel", "stop", "nevermind", "abort"):
            self.memory.update_user_memory(user_id, {"pending_invoice": None})
            response = "No problem — invoice cancelled. Nothing was generated."
            self._store_conversation(user_id, message, response)
            return {"operation": "invoice_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        desc = message.strip()
        if row_id and desc:
            _safe_desc = desc.replace("'", "''")
            _safe_id = str(row_id).replace("'", "''")
            _safe_uid = str(uid).replace("'", "''")
            self.supabase.execute_sql(
                f"UPDATE public.job_entries SET job_description_details = '{_safe_desc}' "
                f"WHERE id = '{_safe_id}' AND user_id = '{_safe_uid}' AND (\"isDeleted\" IS NOT TRUE)"
            )
            logger.info(f"[INVOICE_GATE] Saved job description for row {row_id}")

        if pending_invoice:
            return self._resume_invoice_flow(user_id, pending_invoice)
        response = "Got it, saved that job description."
        self._store_conversation(user_id, message, response)
        return {"operation": "jobdesc_saved", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_poc_name_response(self, user_id: str, message: str) -> Dict:
        """Handle user providing POC name before invoice generation."""
        user_mem = self.memory.get_user_memory(user_id)
        pending_invoice = user_mem.get("pending_invoice", {})
        client_name = user_mem.get("pending_poc_client", "")
        data_user_id = user_mem.get("pending_poc_user_id", user_id)
        row_ids = user_mem.get("pending_poc_row_ids", []) or []

        # Clear awaiting state
        self.memory.update_user_memory(user_id, {
            "awaiting_poc_name": False,
            "pending_poc_client": None,
            "pending_poc_user_id": None,
            "pending_poc_row_ids": None,
        })

        # POC name is mandatory — 'cancel' aborts the invoice.
        if message.strip().lower() in ("cancel", "stop", "abort", "nevermind"):
            self.memory.update_user_memory(user_id, {"pending_invoice": None})
            response = "No problem — invoice cancelled. Nothing was generated."
            self._store_conversation(user_id, message, response)
            return {"operation": "invoice_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        poc_name = message.strip()
        # Save to the specific rows queried for this invoice (or all matching client rows
        # if row_ids is empty as a fallback).
        if poc_name:
            safe_poc = poc_name.replace("'", "''")
            safe_uid = data_user_id.replace("'", "''")
            if row_ids:
                id_list = ", ".join(f"'{str(r).replace(chr(39), chr(39)+chr(39))}'" for r in row_ids)
                update_sql = (
                    f"UPDATE public.job_entries SET poc_name = '{safe_poc}' "
                    f"WHERE user_id = '{safe_uid}' AND id IN ({id_list}) "
                    f"AND (\"isDeleted\" IS NOT TRUE)"
                )
            elif client_name:
                safe_client = client_name.replace("'", "''")
                update_sql = (
                    f"UPDATE public.job_entries SET poc_name = '{safe_poc}' "
                    f"WHERE user_id = '{safe_uid}' "
                    f"AND (client_name ILIKE '%{safe_client}%' OR brand_name ILIKE '%{safe_client}%' OR production_house ILIKE '%{safe_client}%') "
                    f"AND (\"isDeleted\" IS NOT TRUE)"
                )
            else:
                update_sql = None
            if update_sql:
                result = self.supabase.execute_sql(update_sql)
                if result.get("ok"):
                    logger.info(f"[POC] Saved poc_name='{poc_name}' for {client_name}")
                else:
                    logger.warning(f"[POC] Failed to save poc_name: {result.get('error')}")

        if pending_invoice:
            return self._resume_invoice_flow(user_id, pending_invoice)

        response = f"POC name saved as {poc_name}."
        self._store_conversation(user_id, message, response)
        return {"operation": "poc_saved", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _prompt_bank_details_format(self, user_id: str, message: str) -> Dict:
        """Ask the user to send all bank details in a single structured message."""
        self.memory.update_user_memory(user_id, {"awaiting_bank_details": True})
        response = (
            "Sure! Please send your bank details in this format:\n\n"
            "Account Name: Darshit Mody\n"
            "Bank Name: HDFC Bank\n"
            "Account Number: 1234567890\n"
            "IFSC: HDFC0001234\n"
            "UPI: darshit@upi\n\n"
            "UPI is optional — skip it if you don't have one.\n"
            "Type 'cancel' to skip."
        )
        self._store_conversation(user_id, message, response)
        return {"operation": "bank_details_prompt", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_bank_details_response(self, user_id: str, message: str) -> Dict:
        """Parse a single structured message containing bank details and upsert."""
        # Clear the awaiting flag first
        self.memory.update_user_memory(user_id, {"awaiting_bank_details": False})

        if message.strip().lower() in ("cancel", "stop", "nevermind", "skip"):
            response = "No problem, bank details update cancelled."
            self._store_conversation(user_id, message, response)
            return {"operation": "bank_details_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}

        parsed = self._parse_bank_details_message(message)
        if not parsed:
            response = (
                "I couldn't find the bank details in your message. "
                "Please send them in this format:\n\n"
                "Account Name: Your Name\n"
                "Bank Name: HDFC Bank\n"
                "Account Number: 1234567890\n"
                "IFSC: HDFC0001234\n"
                "UPI: you@upi\n\n"
                "Or type 'cancel' to skip."
            )
            # Re-enable the awaiting flag so user can try again
            self.memory.update_user_memory(user_id, {"awaiting_bank_details": True})
            self._store_conversation(user_id, message, response)
            return {"operation": "bank_details_retry", "response": response, "trigger_invoice": False, "invoice_data": {}}

        result = self.supabase.upsert_user_config(user_id, parsed)
        if result.get("ok"):
            # Check if there's a pending invoice to generate
            user_mem = self.memory.get_user_memory(user_id)
            pending_invoice = user_mem.get("pending_invoice")
            if pending_invoice:
                # Bank saved — re-enter the invoice flow so the gate runs the
                # remaining mandatory checks (address, etc.) before generating.
                self._store_conversation(user_id, message, "Your bank details have been saved! ✅")
                return self._resume_invoice_flow(user_id, pending_invoice)
            else:
                response = "Your bank details have been saved successfully! Say 'my bank details' to view them."
        else:
            response = f"I couldn't save your bank details: {result.get('error', 'Unknown error')}. Please try again."
        self._store_conversation(user_id, message, response)
        return {"operation": "bank_config_complete", "response": response, "trigger_invoice": False, "invoice_data": {}}

    @staticmethod
    def _parse_bank_details_message(message: str) -> Optional[Dict[str, str]]:
        """
        Parse a structured message like:
          Account Name: Darshit Mody
          Bank Name: HDFC Bank
          Account Number: 1234567890
          IFSC: HDFC0001234
          UPI: darshit@upi
        Returns dict of bank fields or None if nothing was parseable.
        """
        import re
        text = message.strip()
        result = {}

        # Map of possible labels → db field name.
        # NOTE: bank_account_number includes a bare "account" label (negative-
        # lookahead excludes "account name"/"account holder") so inputs like
        # "Account: 123456" — common shorthand — are captured, not dropped.
        label_map = {
            "bank_account_name": [r"account\s*(?:holder\s*)?name", r"account\s*holder", r"holder\s*name", r"name\s*on\s*account"],
            "bank_name": [r"bank\s*name", r"bank"],
            "bank_account_number": [
                r"account\s*(?:no|number|num|#)",
                r"a/?c\s*(?:no|number|num|#)?",
                r"account(?!\s*(?:holder|name|holder\s*name))",
            ],
            "bank_ifsc": [r"ifsc\s*(?:code)?"],
            "upi_id": [r"upi\s*(?:id)?"],
        }

        for field, patterns in label_map.items():
            for pat in patterns:
                match = re.search(rf"(?:^|\n)\s*{pat}\s*[:=\-]\s*(.+)", text, re.IGNORECASE)
                if match:
                    val = match.group(1).strip().rstrip(",;")
                    if val.lower() not in ("", "none", "na", "n/a", "-", "skip"):
                        result[field] = val
                    break

        # Need at least account name + account number to be useful
        if not result.get("bank_account_name") and not result.get("bank_account_number"):
            return None
        return result if result else None

    def _show_bank_details(self, user_id: str, message: str) -> Dict:
        """Show stored bank details for the user with masked account number."""
        result = self.supabase.get_user_bank_details(user_id)
        if not result.get("ok"):
            response = f"I couldn't retrieve your bank details: {result.get('error', 'Unknown error')}."
        elif not result.get("data"):
            response = "You haven't set up bank details yet. Say 'update bank details' to add them."
        else:
            bd = result["data"]
            acct = bd.get("bank_account_number") or ""
            masked_acct = f"****{acct[-4:]}" if len(acct) >= 4 else acct or "Not set"
            lines = [
                "Your stored bank details:\n",
                f"Account Holder: {bd.get('bank_account_name') or 'Not set'}",
                f"Bank Name: {bd.get('bank_name') or 'Not set'}",
                f"Account Number: {masked_acct}",
                f"IFSC Code: {bd.get('bank_ifsc') or 'Not set'}",
                f"UPI ID: {bd.get('upi_id') or 'Not set'}",
                "\nSay 'update bank details' to change these.",
            ]
            response = "\n".join(lines)
        self._store_conversation(user_id, message, response)
        return {"operation": "bank_details_view", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _detect_small_talk(self, message: str, user_id: str = None) -> Optional[str]:
        """
        Returns a canned response if the message is pure small talk, else None.
        Short messages with no data keywords are matched against _SMALL_TALK_TRIGGERS.
        """
        import hashlib
        msg = message.strip().lower().rstrip("!?.,:;")

        data_keywords = {
            "invoice", "bill", "payment", "fees", "client", "job",
            "remind", "overdue", "due", "total", "billing", "record",
            "add", "show", "get", "send", "fetch", "how much", "how many",
            "query", "list", "find", "search", "last", "latest",
            "bank", "update",
        }

        is_exact = msg in self._SMALL_TALK_TRIGGERS
        is_short = len(msg.split()) <= 6
        has_data = any(kw in msg for kw in data_keywords)

        if has_data:
            return None
        if not is_exact:
            if not is_short:
                return None
            multi_match = any(trigger in msg for trigger in self._SMALL_TALK_TRIGGERS if " " in trigger)
            if not multi_match:
                return None

        def _pick(options):
            idx = int(hashlib.md5(message.encode()).hexdigest(), 16) % len(options)
            return options[idx]

        msg = message.strip().lower()
        user_name = self._get_user_name(user_id)
        
        bye_words = {"bye", "goodbye", "good bye", "see you", "see ya", "cya", "ttyl",
                     "alvida", "phir milenge", "kal milte hain", "baad mein baat karte hain"}
        thanks_words = {"thanks", "thank you", "thx", "ty", "cheers",
                        "shukriya", "dhanyavaad", "shukran"}
        how_words = {"how are you", "how r u", "how are u", "how are you doing",
                     "how\'s it going", "hows it going", "what\'s up", "whats up", "wassup",
                     "kya haal hai", "kya hal hai", "kya haal h", "kya hal h",
                     "kaise ho", "kaisa chal raha hai", "kya chal raha hai", "kya scene hai"}
        time_words = {"good morning", "good afternoon", "good evening", "good night",
                      "morning", "afternoon", "evening"}
        affirmation_words = {"ok", "okay", "cool", "got it", "great", "nice", "awesome",
                             "haan", "haan ji"}
        dismissal_words = {
            "nothing", "nothing thanks", "nothing thank you", "nothing, thanks",
            "no thanks", "no thank you", "nope thanks", "nah thanks",
            "all good", "all good thanks", "i'm good", "im good", "i'm fine", "im fine",
            "that's all", "thats all", "that's it", "thats it",
            "no need", "not needed", "never mind", "nevermind", "nvm",
            "i'm ok", "im ok", "i'm okay", "im okay",
            "nahi", "nahi ji", "bas karo", "rehne do", "chhoddo",
            "sab theek", "sab thik", "theek hoon", "thik hoon", "mast hoon",
        }

        # Get base response
        if msg in bye_words:
            response = _pick(self._SMALL_TALK_RESPONSES["bye"])
        elif msg in thanks_words or any(tw in msg for tw in thanks_words):
            response = _pick(self._SMALL_TALK_RESPONSES["thanks"])
        elif msg in dismissal_words or any(d in msg for d in dismissal_words):
            response = _pick(self._SMALL_TALK_RESPONSES["affirmation"])
        elif any(hw in msg for hw in how_words):
            response = _pick(self._SMALL_TALK_RESPONSES["how_are_you"])
        elif msg in time_words:
            if any(w in msg for w in ("good night", "night")):
                response = _pick(self._SMALL_TALK_RESPONSES["good_night"])
            elif any(w in msg for w in ("good evening", "evening")):
                response = _pick(self._SMALL_TALK_RESPONSES["good_evening"])
            elif any(w in msg for w in ("good afternoon", "afternoon")):
                response = _pick(self._SMALL_TALK_RESPONSES["good_afternoon"])
            else:
                response = _pick(self._SMALL_TALK_RESPONSES["good_morning"])
        elif msg in affirmation_words:
            response = _pick(self._SMALL_TALK_RESPONSES["affirmation"])
        else:
            response = _pick(self._SMALL_TALK_RESPONSES["greeting"])
        
        # Personalize if we know the user's name
        if user_name and "Hi there" not in response:  # Avoid double personalization
            response = response.replace("Hey!", f"Hey {user_name}!")
            response = response.replace("Hi there!", f"Hi {user_name}!")
            response = response.replace("Hello!", f"Hello {user_name}!")
        
        return response

    def _handle_pending_reminder(self, user_id: str, message: str) -> Optional[Dict]:
        """
        Check if a WhatsApp user has pending reminders and is replying with
        a number (e.g. '1', '2') to send, or 'skip' to dismiss.
        Also handles overdue-audit replies ('paid 1', 'all paid', 'later').
        Returns a response dict if handled, None otherwise.
        """
        pending = get_pending(user_id)
        if not pending:
            return None

        # A pending reminder is a persistent, never-expiring background flag
        # (written by the cron worker, shared via the DB). It must NOT steal a
        # reply meant for an active sub-flow — adding a job, providing a POC
        # email, a yes/no confirmation, a disambiguation pick, etc. Those flows
        # are handled later in the pipeline; if one is active, yield so the user
        # isn't trapped. (A genuine reminder reply is a standalone number / skip /
        # all, which the user can send once the sub-flow is done.)
        _mem = self.memory.get_user_memory(user_id)
        _active_subflow = bool(_mem.get("pending_disambiguation")) or any(
            _mem.get(k) for k in (
                "awaiting_job_input", "awaiting_poc_email", "awaiting_invoice_month",
                "awaiting_send_confirmation", "awaiting_bank_details",
                "awaiting_client_billing", "awaiting_poc_name", "awaiting_name_change",
                "awaiting_link_id", "awaiting_modify_field", "awaiting_compound_response",
                "awaiting_invoice_address", "awaiting_job_description",
            )
        )
        if _active_subflow:
            logger.info("[REMINDER] Active sub-flow in progress — yielding so the reminder doesn't hijack the reply")
            return None

        msg = message.strip().lower()

        # ── Overdue-audit branch — pending entries tagged _audit_row=True ─
        is_audit = any(p.get("_audit_row") for p in pending)
        if is_audit:
            resp = self._handle_pending_audit_reply(user_id, message, msg, pending)
            if resp is not None:
                return resp
            # Fall through if the user typed something not recognized — let
            # the universal pipeline handle it (don't trap forever).
            return None

        # "skip" / "skip all" → clear pending
        if msg in ("skip", "skip all", "no", "cancel"):
            clear_pending(user_id)
            response = "⏭ Reminders skipped. You can always send them manually later."
            self._store_conversation(user_id, message, response)
            return {"operation": "reminder", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # "all" / "send all" → send every pending reminder
        if msg in ("all", "send all", "send all reminders"):
            return self._send_all_pending_reminders(user_id, message, pending)

        # A reminder reply is a STANDALONE number selection — "1", "2", "send 1",
        # "#1". A number buried in free text ("...date 5 may 2025, fees 20k") is
        # NOT a reminder reply; the old re.search grabbed the first digit anywhere
        # and hijacked add-job / query messages. Require the whole message to be a
        # number selection.
        num_match = re.fullmatch(r"(?:send\s+|reminder\s+|#)?\s*(\d+)\s*\.?", msg)
        if not num_match:
            return None  # Not a reminder reply, let normal flow handle it

        idx = int(num_match.group(1))
        if idx < 1 or idx > len(pending):
            response = f"Please reply with a number between 1 and {len(pending)}, or 'skip' to skip all."
            self._store_conversation(user_id, message, response)
            return {"operation": "reminder", "response": response, "trigger_invoice": False, "invoice_data": {}}

        reminder = pending[idx - 1]
        job_id = reminder.get("id")
        level = reminder.get("_reminder_level", "first")
        poc_email = reminder.get("poc_email")
        bill_no = reminder.get("bill_no") or "N/A"
        client_name = reminder.get("client_name") or "Client"
        poc_name = reminder.get("poc_name") or client_name
        fees = reminder.get("fees")

        if not poc_email:
            response = f"❌ No email on file for {client_name}. Please add a POC email first."
            self._store_conversation(user_id, message, response)
            return {"operation": "reminder", "response": response, "trigger_invoice": False, "invoice_data": {}}

        try:
            amount_str = f"₹{int(float(fees)):,}"
        except (ValueError, TypeError):
            amount_str = str(fees) if fees else "N/A"

        subject_map = {
            "first": f"First Payment Reminder – Invoice #{bill_no}",
            "second": f"Second Payment Reminder – Invoice #{bill_no}",
            "third": f"Final Payment Reminder – Invoice #{bill_no}",
        }
        subject = subject_map.get(level, f"Payment Reminder – Invoice #{bill_no}")

        # Get sender name
        profile = self.supabase.get_user_profile(user_id)
        sender_name = "Team"
        if profile.get("ok") and profile.get("data"):
            sender_name = profile["data"].get("name") or sender_name

        body = (
            f"Hi {poc_name},\n\n"
            f"This is a friendly reminder regarding invoice #{bill_no}.\n\n"
            f"Amount Due: {amount_str}\n\n"
            f"Please let us know if payment has already been processed.\n\n"
            f"Best regards,\n{sender_name}\n"
        )

        ok = self.email.send_email(to_email=poc_email, subject=subject, body=body)

        if not ok:
            response = f"❌ Failed to send reminder email to {poc_email}. Please try again later."
            self._store_conversation(user_id, message, response)
            return {"operation": "reminder", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Update DB flag
        flag_map = {
            "first": "first_reminder_sent",
            "second": "second_reminder_sent",
            "third": "third_reminder_sent",
        }
        flag_col = flag_map.get(level)
        if flag_col and job_id:
            update_sql = f"UPDATE public.job_entries SET {flag_col} = NOW() WHERE id = '{job_id}'"
            self.supabase.execute_sql(update_sql)

        # Remove this reminder from pending list
        remove_single(user_id, job_id)

        label_map = {"first": "First", "second": "Second", "third": "Final"}
        label = label_map.get(level, level.title())
        response = f"✅ {label} reminder sent to {poc_email} for invoice #{bill_no}."

        # If more pending, remind user
        remaining = get_pending(user_id)
        if remaining:
            response += f"\n\n{len(remaining)} reminder(s) still pending. Reply with a number or 'skip'."

        self._store_conversation(user_id, message, response)
        return {"operation": "reminder", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_pending_audit_reply(self, user_id: str, message: str, msg_lower: str, pending: list) -> Optional[Dict]:
        """Handle WhatsApp replies to the overdue-audit message.
        Accepts: 'paid <n>' / 'paid' / 'all paid' / 'paid all' / 'later' / 'skip'.
        Returns a response dict, or None if reply wasn't recognized."""
        # 'later' / 'remind later' — push the next nag out by stamping NOW().
        if msg_lower in ("later", "remind later", "remind me later", "next week", "not now"):
            ids = [p.get("id") for p in pending if p.get("id")]
            if ids:
                ids_sql = ",".join(f"'{i}'" for i in ids)
                self.supabase.execute_sql(
                    f"UPDATE public.job_entries SET overdue_audit_sent = NOW() WHERE id IN ({ids_sql})"
                )
            clear_pending(user_id)
            response = "⏸ Got it — I'll check back next week."
            self._store_conversation(user_id, message, response)
            return {"operation": "audit_later", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # 'all paid' / 'paid all' / 'mark all paid' — bulk
        if msg_lower in ("all paid", "paid all", "mark all paid", "all are paid", "yes all paid"):
            ids = [p.get("id") for p in pending if p.get("id")]
            if not ids:
                clear_pending(user_id)
                return {"operation": "audit_paid", "response": "Nothing to update.", "trigger_invoice": False, "invoice_data": {}}
            ids_sql = ",".join(f"'{i}'" for i in ids)
            res = self.supabase.execute_sql(
                f"UPDATE public.job_entries SET paid = 'Yes', payment_date = CURRENT_DATE "
                f"WHERE id IN ({ids_sql}) RETURNING id"
            )
            count = len(res.get("rows", []) or ids)
            clear_pending(user_id)
            response = f"✅ Marked {count} invoice(s) as paid. Nice — payments cleared."
            self._store_conversation(user_id, message, response)
            return {"operation": "audit_paid", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # 'paid <n>' / 'paid' / '<n> paid' / 'mark 2 paid'
        if "paid" in msg_lower:
            num_m = re.search(r"\b(\d+)\b", msg_lower)
            if num_m:
                idx = int(num_m.group(1))
                if 1 <= idx <= len(pending):
                    target = pending[idx - 1]
                    job_id = target.get("id")
                    if job_id:
                        self.supabase.execute_sql(
                            f"UPDATE public.job_entries SET paid = 'Yes', payment_date = CURRENT_DATE "
                            f"WHERE id = '{job_id}'"
                        )
                    remove_single(user_id, job_id)
                    client = target.get("client_name") or "the invoice"
                    try:
                        amt = f" — ₹{int(float(target.get('fees') or 0)):,}"
                    except Exception:
                        amt = ""
                    remaining = get_pending(user_id)
                    response = f"✅ Marked paid: {client}{amt}."
                    if remaining:
                        response += f"\n\n{len(remaining)} invoice(s) still pending. Reply 'paid <n>', 'all paid', or 'later'."
                    self._store_conversation(user_id, message, response)
                    return {"operation": "audit_paid", "response": response, "trigger_invoice": False, "invoice_data": {}}
                response = f"Please choose a number between 1 and {len(pending)}."
                self._store_conversation(user_id, message, response)
                return {"operation": "audit_paid", "response": response, "trigger_invoice": False, "invoice_data": {}}
            # 'paid' alone with only one pending row → mark it
            if len(pending) == 1:
                target = pending[0]
                job_id = target.get("id")
                if job_id:
                    self.supabase.execute_sql(
                        f"UPDATE public.job_entries SET paid = 'Yes', payment_date = CURRENT_DATE "
                        f"WHERE id = '{job_id}'"
                    )
                clear_pending(user_id)
                client = target.get("client_name") or "the invoice"
                response = f"✅ Marked paid: {client}."
                self._store_conversation(user_id, message, response)
                return {"operation": "audit_paid", "response": response, "trigger_invoice": False, "invoice_data": {}}

        return None  # not recognized — let other handlers try

    def _send_all_pending_reminders(self, user_id: str, message: str, pending: list) -> Dict:
        """Send reminder emails for every item in the pending list (WhatsApp 'send all')."""
        profile = self.supabase.get_user_profile(user_id)
        sender_name = "Team"
        if profile.get("ok") and profile.get("data"):
            sender_name = profile["data"].get("name") or sender_name

        flag_map = {
            "first": "first_reminder_sent",
            "second": "second_reminder_sent",
            "third": "third_reminder_sent",
        }
        label_map = {"first": "First", "second": "Second", "third": "Final"}

        sent = []
        failed = []

        for reminder in pending:
            job_id = reminder.get("id")
            level = reminder.get("_reminder_level", "first")
            poc_email = reminder.get("poc_email")
            bill_no = reminder.get("bill_no") or "N/A"
            client_name = reminder.get("client_name") or "Client"
            poc_name = reminder.get("poc_name") or client_name
            fees = reminder.get("fees")

            if not poc_email:
                failed.append(f"{client_name} (no email)")
                continue

            try:
                amount_str = f"₹{int(float(fees)):,}"
            except (ValueError, TypeError):
                amount_str = str(fees) if fees else "N/A"

            subject_map = {
                "first": f"First Payment Reminder – Invoice #{bill_no}",
                "second": f"Second Payment Reminder – Invoice #{bill_no}",
                "third": f"Final Payment Reminder – Invoice #{bill_no}",
            }
            subject = subject_map.get(level, f"Payment Reminder – Invoice #{bill_no}")

            body = (
                f"Hi {poc_name},\n\n"
                f"This is a friendly reminder regarding invoice #{bill_no}.\n\n"
                f"Amount Due: {amount_str}\n\n"
                f"Please let us know if payment has already been processed.\n\n"
                f"Best regards,\n{sender_name}\n"
            )

            ok = self.email.send_email(to_email=poc_email, subject=subject, body=body)
            if ok:
                sent.append(f"{client_name} → {poc_email}")
                flag_col = flag_map.get(level)
                if flag_col and job_id:
                    self.supabase.execute_sql(
                        f"UPDATE public.job_entries SET {flag_col} = NOW() WHERE id = '{job_id}'"
                    )
            else:
                failed.append(f"{client_name} ({poc_email})")

        # Clear all pending
        clear_pending(user_id)

        lines = [f"✅ Sent {len(sent)} reminder(s)."]
        for s in sent:
            lines.append(f"  • {s}")
        if failed:
            lines.append(f"\n❌ Failed for {len(failed)}:")
            for f_item in failed:
                lines.append(f"  • {f_item}")
        response = "\n".join(lines)
        self._store_conversation(user_id, message, response)
        return {"operation": "reminder", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def process_request(self, user_id: str, message: str) -> Dict:
        """
        Main handler: keyword-based branches for reminder/invoice/overdue;
        then LLM query plan → validate → resolve time → execute → format.
        """
        # ── Beta gate (BETA_GATE_ENABLED=true) ────────────────────────────
        # Block first-touch from non-allowlisted users so we can pilot with
        # specific test numbers. Existing users (those who already have a
        # profile with onboarded_at set) are always exempt — they were
        # auto-seeded into allowed_users when the gate shipped. Brand-new
        # users get a friendly "private beta" reply and NO profile gets
        # created until you add them via SQL/admin path.
        _beta_gate_on = (os.getenv("BETA_GATE_ENABLED", "").strip().lower()
                        in ("1", "true", "yes", "on"))
        if _beta_gate_on:
            try:
                _exists = self.supabase.get_user_profile(user_id) or {}
                _has_onboarded = bool((_exists.get("data") or {}).get("onboarded_at"))
            except Exception:
                _has_onboarded = False
            # Brand-new (no profile) OR mid-onboarding profile → check allowlist.
            if not _has_onboarded and not self.supabase.is_user_allowed(user_id):
                logger.info(f"[BETA_GATE] blocked first-touch from {user_id}")
                response = (
                    "Hey 👋 — Remyndly is in private beta right now.\n\n"
                    "To get added to the access list, reach out:\n"
                    "• Email: admin@remyndly.io\n"
                    "• WhatsApp: +91 70386 75067 or +91 99303 30887\n\n"
                    "Share your mobile number with us and you'll be in within minutes."
                )
                return {
                    "operation": "beta_gate_blocked",
                    "response": response,
                    "trigger_invoice": False,
                    "invoice_data": {},
                }

        # Check if user is new and needs onboarding
        profile = self.supabase.get_user_profile(user_id)
        if not profile.get("ok"):
            logger.error(f"Failed to check user profile for {user_id}: {profile.get('error')}")
            # Treat DB errors as new user → start onboarding (creates profile on success)
            return self._start_onboarding(user_id, message)
        elif not profile.get("data"):
            # New user - start onboarding
            return self._start_onboarding(user_id, message)
        elif not profile.get("data", {}).get("onboarded_at"):
            # User exists but not onboarded - continue onboarding
            return self._continue_onboarding(user_id, message, profile["data"])

        from services.business_logic_service import BusinessLogicService
        logic = BusinessLogicService()
        conversation_history = self.memory.get_conversation_history(user_id)
        # Bind user_mem early so the v2 block (which runs before the legacy
        # cascade re-fetches it) can read it safely. Without this top-of-fn
        # assignment, the genexpr inside `if _v2_enabled` triggers a
        # NameError because Python marks user_mem as a function-local but
        # the first textual assignment is later in the function.
        user_mem = self.memory.get_user_memory(user_id) or {}
        trigger_invoice = False
        invoice_data = {}

        # Resolve effective user_id for data queries (account linking)
        data_user_id = self._resolve_data_user_id(user_id, profile.get("data", {}))
        if data_user_id != user_id:
            logger.info(f"[LINK] Using linked data_user_id={data_user_id} for user {user_id}")

        try:
            # 0. Check for active form (smart capture confirmation / missing fields)
            form_state = self.memory.get_form_state(user_id)
            if form_state:
                return self._handle_form_step(user_id, message)

            # 0+. Check for pending payment reminders (WhatsApp reply flow)
            reminder_result = self._handle_pending_reminder(user_id, message)
            if reminder_result:
                return reminder_result

            # ── FlowMachine v2 — session 1 (classifier + IDLE leaf routing) ──
            # Behind a feature flag so production stays on the legacy path until
            # we explicitly flip FLOW_MACHINE_V2=true on Railway. v2 only takes
            # over for messages that arrive while the user is FULLY IDLE.
            # Read/write intents shadow-only (telemetry); SMALL_TALK / FEATURE_QUESTION
            # / UNKNOWN are owned by v2 for instant on-brand replies.
            try:
                _v2_enabled = os.getenv("FLOW_MACHINE_V2", "").strip().lower() in ("1", "true", "yes", "on")
            except Exception:
                _v2_enabled = False
            # Lift the verdict out of the v2 block so the legacy INVOICE_CHECK
            # below can defer to it for high-confidence READ_QUERY / READ_AGGREGATE
            # classifications. Without this, the legacy invoice keyword check
            # silently overrode v2's correct call (e.g. for Hinglish "kiska
            # invoice baki hai bhejna" — v2 said READ_QUERY but legacy said
            # 'looks like invoice, ask which client').
            _v2_verdict = None
            if _v2_enabled:
                try:
                    # First: apply TTL — long-idle flow auto-resets so the user
                    # isn't trapped in a stale state from hours ago. When the
                    # FlowMachine resets, also clear ALL legacy awaiting flags
                    # so they don't re-arm the flow on the next message.
                    if self.flow_machine.expire_if_stale(user_id):
                        _stale_clear = {
                            "awaiting_send_confirmation": False,
                            "pending_send_invoice":       None,
                            "awaiting_client_billing":    False,
                            "pending_billing_client":     None,
                            "pending_billing_user_id":    None,
                            "pending_invoice":            None,
                            "awaiting_poc_name":          False,
                            "pending_poc_client":         None,
                            "pending_poc_user_id":        None,
                            "pending_poc_row_ids":        None,
                            "awaiting_poc_email":         False,
                            "poc_email_client":           None,
                            "awaiting_job_input":         False,
                        }
                        self.memory.update_user_memory(user_id, _stale_clear)
                        # Also drop any in-progress smart-capture form.
                        try:
                            if self.memory.get_form_state(user_id):
                                self.memory.cancel_form(user_id)
                        except Exception:
                            pass
                        user_mem = self.memory.get_user_memory(user_id)
                except Exception as _ttl_err:
                    logger.warning(f"[V2] TTL check failed: {_ttl_err}")

                # Reconciliation: legacy code paths still arm awaiting_* flags
                # in dozens of places (we haven't migrated each arm site to
                # write FlowMachine directly). Once per message, if v2 thinks
                # IDLE but a legacy flag is set, sync FlowMachine to match so
                # dispatch_in_flow can take over.
                try:
                    self._reconcile_legacy_to_flow_machine(user_id, user_mem)
                except Exception as _rec_err:
                    logger.warning(f"[V2] reconcile failed: {_rec_err}")

                from services.classifier import classify as _v2_classify
                from services.flow_dispatcher import dispatch_idle as _v2_dispatch_idle
                from services.flow_dispatcher import dispatch_in_flow as _v2_dispatch_in_flow
                from services.flow_machine import FLOW_IDLE
                _v2_current_flow = self.flow_machine.current_flow(user_id)
                _v2_current_state = self.flow_machine.get_state(user_id)
                _v2_in_owned_flow = _v2_current_flow != FLOW_IDLE
                _schema_summary = ", ".join(
                    c for c in JOB_ENTRIES_COLUMNS if not c.startswith("_")
                )[:1500]

                if _v2_in_owned_flow:
                    # In a v2-owned flow — classify with flow context, route through dispatch_in_flow.
                    try:
                        _verdict = _v2_classify(
                            message, self.gemini,
                            conversation_history=conversation_history,
                            schema_summary=_schema_summary,
                            current_flow=_v2_current_flow,
                            current_context=_v2_current_state.get("context") or {},
                        )
                        if _verdict:
                            _result = _v2_dispatch_in_flow(
                                _verdict,
                                intent_service=self,
                                user_id=user_id,
                                current_flow=_v2_current_flow,
                                current_context=_v2_current_state.get("context") or {},
                                conversation_history=conversation_history,
                            )
                            if _result is not None:
                                return _result
                            # Shadow → fall through to legacy.
                    except Exception as _v2_err:
                        logger.warning(f"[V2] in-flow dispatch failed, falling back: {_v2_err}")
                else:
                    # IDLE path — also gated on no legacy awaiting flag set so
                    # mid-migration flows still use legacy handlers.
                    _idle_blockers = (
                        "awaiting_job_input", "awaiting_invoice_month", "awaiting_poc_email",
                        "awaiting_send_confirmation", "awaiting_client_billing",
                        "awaiting_poc_name", "awaiting_bank_details", "awaiting_name_change",
                        "awaiting_modify_field", "pending_disambiguation",
                    )
                    _is_idle = (
                        not any(user_mem.get(k) for k in _idle_blockers)
                        and not self.memory.get_form_state(user_id)
                    )
                    if _is_idle:
                        try:
                            _verdict = _v2_classify(
                                message, self.gemini,
                                conversation_history=conversation_history,
                                schema_summary=_schema_summary,
                            )
                            if _verdict:
                                _v2_verdict = _verdict  # lift for legacy override below
                                _result = _v2_dispatch_idle(
                                    _verdict,
                                    intent_service=self,
                                    user_id=user_id,
                                    conversation_history=conversation_history,
                                )
                                if _result is not None:
                                    return _result
                        except Exception as _v2_err:
                            logger.warning(f"[V2] idle dispatch failed, falling back: {_v2_err}")

            # 0a-. Small talk detection (greetings, thanks, etc.) — avoid expensive SQL path
            small_talk_resp = self._detect_small_talk(message, user_id=user_id)
            if small_talk_resp:
                self._store_conversation(user_id, message, small_talk_resp)
                return {"operation": "small_talk", "response": small_talk_resp, "trigger_invoice": False, "invoice_data": {}}

            # 0a. Check if user is responding with job data (awaiting smart capture input)
            user_mem = self.memory.get_user_memory(user_id)

            # An active invoice-email flow (asking for a POC email or a yes/no send
            # confirmation) takes precedence over a STALE disambiguation. Otherwise a
            # reply meant for the email prompt (an address, "yes"/"no") gets swallowed
            # by a leftover delete/select disambiguation. Drop the stale state here.
            _invoice_await_active = (
                user_mem.get("awaiting_poc_email")
                or user_mem.get("awaiting_send_confirmation")
            )
            if _invoice_await_active and user_mem.get("pending_disambiguation"):
                logger.info("[DISAMBIG] Invoice-email flow active — clearing stale disambiguation so the email reply is handled correctly")
                self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
                user_mem = self.memory.get_user_memory(user_id)

            # Handle pending disambiguation reply (user selecting a specific row by number)
            if user_mem.get("pending_disambiguation"):
                _disambig_result = self._handle_disambiguation_reply(user_id, message, user_mem["pending_disambiguation"])
                if _disambig_result is not None:
                    return _disambig_result
                # None means the handler detected a new query and cleared state — fall through

            # Handle compound intent follow-up ("Yes" after "You also mentioned: ...")
            if user_mem.get("awaiting_compound_response"):
                pending_action = user_mem.get("suggested_next_action", "")
                # Always clear the state first (interruption-safe)
                self.memory.update_user_memory(user_id, {
                    "awaiting_compound_response": False,
                    "suggested_next_action": None,
                })
                msg_lower_check = message.strip().lower().rstrip("., !")
                _YES_EXACT = {"yes", "y", "yeah", "yep", "sure", "ok", "okay", "go ahead", "do it", "yes please"}
                _YES_PREFIXES = ("yes,", "yes ", "yeah,", "yeah ", "sure,", "sure ", "ok,", "ok ", "okay,", "okay ")
                _is_yes = msg_lower_check in _YES_EXACT or msg_lower_check.startswith(_YES_PREFIXES)
                if _is_yes and pending_action:
                    # Preserve any qualifier the user added after "yes" (e.g.
                    # "yes along with bill numbers") so the pending action runs
                    # WITH that extra context, not in isolation.
                    remainder = ""
                    if msg_lower_check not in _YES_EXACT:
                        for _p in _YES_PREFIXES:
                            if msg_lower_check.startswith(_p):
                                remainder = message.strip()[len(_p):].strip(" ,.!")
                                break
                    merged = f"{pending_action} {remainder}".strip() if remainder else pending_action
                    logger.info(f"[COMPOUND] User confirmed next action: '{merged}' (pending='{pending_action}', qualifier='{remainder}')")
                    return self.process_request(user_id=user_id, message=merged)
                elif msg_lower_check in {"no", "nah", "nope", "skip", "not now", "later"}:
                    response = "👍 No problem. Let me know if you need anything else."
                    self._store_conversation(user_id, message, response)
                    return {"operation": "compound_declined", "response": response, "trigger_invoice": False, "invoice_data": {}}
                # else: user said something unrelated — fall through to normal processing

            # 0a2. Modify / update / change a job — AI-extracted update intent
            #      Triggers: explicit modify verb in the message, OR the user is
            #      currently in a "what field do you want to change?" follow-up.
            _MODIFY_TRIGGERS = (
                "modify ", "modify\n", "update ", "update\n", "change ", "change\n",
                "edit ", "edit\n", "set ", "mark ",
            )
            _MODIFY_EQUALS = ("modify", "update", "change", "edit")
            _msg_l = message.strip().lower()
            _has_modify_verb = any(_msg_l.startswith(t) for t in _MODIFY_TRIGGERS) or _msg_l in _MODIFY_EQUALS
            _awaiting_modify = bool(user_mem.get("awaiting_modify_field"))
            # Allow user to escape the modify state with standard cancel words.
            if _awaiting_modify and _msg_l in ("cancel", "stop", "quit", "exit", "nevermind", "nvm", "no", "abort", "skip"):
                self.memory.update_user_memory(user_id, {"awaiting_modify_field": False, "modify_row_id": None})
                response = "No problem, cancelled. What else can I help with?"
                self._store_conversation(user_id, message, response)
                return {"operation": "modify_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}
            if _has_modify_verb or _awaiting_modify:
                _resp = self._handle_modify_intent(user_id, message, user_mem)
                if _resp is not None:
                    return _resp
                # else: fall through (extraction failed AND no row context — let
                # normal pipeline handle it; better than dead-ending here).

            # 0b. Check for "add job" / "+" trigger → AI Smart Capture
            msg_stripped = message.strip()
            add_job_triggers = ["add job", "add a job", "add new job", "add a new job",
                               "new job", "log a job", "log job", "record job", "record a job",
                               "create job", "create a job", "create a new job",
                               "add client", "add a client", "add new client", "add a new client",
                               "new client", "log a client", "log client", "record client", "record a client",
                               "create client", "create a client", "create a new client",
                               "add entry", "add an entry", "add a new entry", "new entry"]
            is_add_job = any(t in msg_stripped.lower() for t in add_job_triggers)
            is_plus = msg_stripped.startswith("+") and len(msg_stripped) > 1
            if is_add_job or is_plus:
                # Check for compound intent using AI (e.g. "add a job and send invoice")
                first_part_msg = message
                if len(message.split()) >= 6:  # only check if message is long enough
                    intents = self.gemini.decompose_compound_intent(message)
                    if intents and len(intents) > 1:
                        first_part_msg = intents[0]
                        suggested_next = intents[1]
                        logger.info(f"[COMPOUND] AI split: first='{first_part_msg}', next='{suggested_next}'")
                        self.memory.update_user_memory(user_id, {
                            "suggested_next_action": suggested_next,
                        })
                return self._start_smart_capture(user_id, first_part_msg)

            if user_mem.get("awaiting_job_input"):
                # Escape hatch: if the user's message clearly looks like a new query
                # (question word + verb, or a known query/command pattern), don't
                # treat it as job-form input — clear the sticky state and fall through.
                _msg_l_jobform = message.strip().lower().rstrip(".!?")
                _question_starts = (
                    "who ", "what ", "when ", "where ", "how ", "why ", "which ",
                    "show ", "list ", "find ", "tell ", "give ", "fetch ", "get me ",
                    "do you ", "can you ", "are you ", "is there ",
                    "delete ", "remove ", "update ", "modify ", "change ",
                    "generate invoice", "send invoice", "mark ",
                )
                _looks_like_query = (
                    any(_msg_l_jobform.startswith(s) for s in _question_starts)
                    or "?" in message
                    or self.gemini.is_new_query_not_response(
                        message,
                        "free-text job description (brand, date, fees, client, POC) for the smart-capture form"
                    )
                )
                if _looks_like_query:
                    logger.info(
                        f"[SMART_CAPTURE] Sticky awaiting_job_input cleared — message "
                        f"looks like a new query: {message[:80]!r}"
                    )
                    self.memory.update_user_memory(user_id, {"awaiting_job_input": False})
                    user_mem = self.memory.get_user_memory(user_id)
                    # Fall through to normal pipeline
                else:
                    return self._extract_and_confirm(user_id, message)

            # Universal intent-shift guard: if the bot is in any single-question awaiting state
            # and the user's message looks like a brand-new query, clear the pending state and
            # continue with the new request instead of silently treating it as a (wrong) answer.
            _PENDING_STATES = {
                "awaiting_invoice_month": "the month name for a pending invoice (e.g. 'March')",
                "awaiting_poc_email":     "a client POC email address",
                "awaiting_send_confirmation": "a yes/no confirmation to send the invoice over email",
                "awaiting_client_billing":   "client billing details (name, address, GST)",
                "awaiting_poc_name":         "a POC name to address the invoice to",
                "awaiting_bank_details":     "the user's own bank details",
                "awaiting_name_change":      "the user's new display name",
            }
            _active_pending = [k for k in _PENDING_STATES if user_mem.get(k)]
            if _active_pending:
                # Short-circuit: well-known response tokens are NEVER a new query.
                # This guards against AI overreach (e.g. classifying "skip" — which is
                # literally what we asked the user to type — as a fresh command).
                _RESPONSE_TOKENS = {
                    "skip", "cancel", "yes", "y", "yeah", "yep", "yup", "yes please",
                    "no", "n", "nope", "nah", "no thanks", "not now", "later",
                    "ok", "okay", "sure", "go ahead", "do it", "confirm", "send it",
                    "don't send", "dont send",
                }
                _msg_stripped = message.strip().lower().rstrip(".!?")
                _is_response_token = _msg_stripped in _RESPONSE_TOKENS
                if _is_response_token:
                    logger.info(f"[INTENT_SHIFT] '{_msg_stripped}' is a response token — keeping pending state {_active_pending}")
                else:
                    ctx_desc = "; ".join(_PENDING_STATES[k] for k in _active_pending)
                    if self.gemini.is_new_query_not_response(message, ctx_desc):
                        logger.info(f"[INTENT_SHIFT] Clearing pending states {_active_pending} — user typed a new query")
                        _clear_patch = {k: False for k in _active_pending}
                        _clear_patch.update({
                            "pending_send_invoice": None,
                            "pending_invoice_client": None,
                            "pending_poc_email_client": None,
                            "pending_poc_name_client": None,
                        })
                        self.memory.update_user_memory(user_id, _clear_patch)
                        user_mem = self.memory.get_user_memory(user_id)  # refresh after clear

            # 0b1.4. Check if user is providing the month for a pending invoice
            if user_mem.get("awaiting_invoice_month"):
                return self._handle_invoice_month_reply(user_id, message, user_mem, data_user_id, conversation_history)

            # 0b1.5. Check if user is providing a client POC email
            if user_mem.get("awaiting_poc_email"):
                return self._handle_poc_email_response(user_id, message)

            # 0b1.6. Check if user is confirming sending invoice to client email
            if user_mem.get("awaiting_send_confirmation"):
                return self._handle_send_confirmation(user_id, message)

            # 0b1.7. Check if user is providing client billing details
            if user_mem.get("awaiting_client_billing"):
                return self._handle_client_billing_response(user_id, message)

            # 0b1.8. Check if user is providing POC name for an invoice
            if user_mem.get("awaiting_poc_name"):
                return self._handle_poc_name_response(user_id, message)

            # 0b1.9. Check if user is providing their business address for the invoice (#2)
            if user_mem.get("awaiting_invoice_address"):
                return self._handle_invoice_address_response(user_id, message)

            # 0b1.10. Check if user is providing a missing job description (#3)
            if user_mem.get("awaiting_job_description"):
                return self._handle_job_description_response(user_id, message)

            # 0b2. Check if user is responding with bank details (awaiting state)
            if user_mem.get("awaiting_bank_details"):
                return self._handle_bank_details_response(user_id, message)

            # 0b3. "update bank details" — ask user for details in a specific format
            msg_lower = message.strip().lower()
            # Direct one-shot bank details: "My account is HDFC 1234 IFSC HDFC0001234, UPI x@y"
            _has_bank_inline = (
                ("my account" in msg_lower or "account number" in msg_lower)
                and ("ifsc" in msg_lower or "upi" in msg_lower or "account is" in msg_lower)
            )
            if _has_bank_inline:
                return self._handle_bank_details_response(user_id, message)
            if any(t in msg_lower for t in self._UPDATE_BANK_TRIGGERS):
                return self._prompt_bank_details_format(user_id, message)

            # 0b3.5. "change my name" / "update my name" — update user profile name
            _NAME_CHANGE_TRIGGERS = [
                "change my name", "update my name", "set my name",
                "rename me", "my name is wrong", "fix my name",
            ]
            if any(t in msg_lower for t in _NAME_CHANGE_TRIGGERS):
                return self._handle_name_change(user_id, message)

            # Check if user is awaiting name change (providing new name)
            if user_mem.get("awaiting_name_change"):
                return self._process_name_change(user_id, message)

            # 0b3.5b. "update my address" / "change my business address" — set or
            # correct the saved invoice business address at any time.
            _ADDRESS_UPDATE_TRIGGERS = [
                "update my address", "change my address", "update my business address",
                "change my business address", "update business address", "change business address",
                "set my address", "edit my address", "update invoice address", "change invoice address",
                "my address is", "my business address is", "wrong address", "address is wrong",
                "fix my address", "correct my address",
            ]
            if any(t in msg_lower for t in _ADDRESS_UPDATE_TRIGGERS):
                return self._handle_address_update(user_id, message, data_user_id)

            # 0b3.6. "what is my user id" / "my user id" — show user_id for account linking
            _USER_ID_TRIGGERS = ["my user id", "what is my id", "what's my id", "show my id", "my id"]
            if any(t in msg_lower for t in _USER_ID_TRIGGERS):
                platform = "Telegram" if user_id.isdigit() else "WhatsApp"
                response = f"Your {platform} user ID is:\n`{user_id}`\n\nShare this with your other platform to link accounts."
                self._store_conversation(user_id, message, response)
                return {"operation": "show_user_id", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 0b3.7. "link account" / "link telegram" — cross-platform account linking
            _LINK_TRIGGERS = [
                "link account", "link my account", "link telegram",
                "link my telegram", "link whatsapp", "link my whatsapp",
                "connect account", "connect telegram", "connect whatsapp",
            ]
            if any(t in msg_lower for t in _LINK_TRIGGERS):
                return self._handle_link_account(user_id, message)

            # Check if user is providing a link ID
            if user_mem.get("awaiting_link_id"):
                return self._process_link_id(user_id, message)

            # 0b4. "my bank details" / "show bank details" — show stored (masked)
            if any(t in msg_lower for t in self._VIEW_BANK_TRIGGERS):
                return self._show_bank_details(user_id, message)

            # 0b5. Negative intent — user declining a follow-up question
            _NEGATIVE_RESPONSES = {
                "no", "nope", "nah", "not required", "not needed", "no thanks",
                "no thank you", "skip", "don't need", "dont need", "i'm good",
                "im good", "pass", "no need", "that's fine", "thats fine",
                "all good", "not now", "maybe later", "no its fine",
                "no it's fine", "not right now", "i'm fine", "im fine",
            }
            _FOLLOWUP_MARKERS = [
                "would you like", "do you want", "shall i", "want me to",
                "should i", "need a breakdown", "like a breakdown",
                "want a breakdown", "like to see", "want to see",
                "interested in", "like more detail", "want more detail",
            ]
            if msg_lower in _NEGATIVE_RESPONSES:
                # Check if last assistant message was a follow-up question
                is_followup = False
                if conversation_history:
                    last_msgs = [m for m in conversation_history if m.get("role") == "assistant"]
                    if last_msgs:
                        last_assistant = last_msgs[-1].get("content", "").lower()
                        is_followup = any(marker in last_assistant for marker in _FOLLOWUP_MARKERS) or last_assistant.rstrip().endswith("?")
                # A bare decline reaching this point has no actionable target (every
                # awaiting-state handler already ran above). Acknowledge gracefully
                # rather than letting "maybe later" / "not now" fall through to the
                # SQL pipeline, where it parses to nothing and surfaces as an error.
                _PURE_SOCIAL_DECLINE = {
                    "maybe later", "not now", "not right now", "no thanks",
                    "no thank you", "i'm good", "im good", "all good", "pass",
                    "that's fine", "thats fine", "no need", "no its fine",
                    "no it's fine", "i'm fine", "im fine",
                }
                if is_followup or msg_lower in _PURE_SOCIAL_DECLINE:
                    response = "👍 Got it. Let me know if you need anything else."
                    self._store_conversation(user_id, message, response)
                    return {"operation": "decline_followup", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 0b6. Clearly out-of-scope requests (book a flight/cab/hotel, order food,
            # etc.) — give a deterministic on-brand refusal instead of letting them hit
            # the SQL pipeline, where they parse to nothing and surface as an error.
            _OOS = re.search(
                r'\b(book|order|reserve|buy|get\s+me|find\s+me)\b.{0,25}'
                r'\b(flight|flights|uber|ola|cab|taxi|ride|hotel|room|ticket|tickets|'
                r'food|pizza|lunch|dinner|coffee|groceries|grocery)\b',
                msg_lower,
            )
            if _OOS:
                response = self.gemini.answer_feature_question(message, conversation_history=conversation_history)
                if not response or not response.strip():
                    response = unsupported_feature_phrase(message[:80])
                self._store_conversation(user_id, message, response)
                return {"operation": "unsupported", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # ── Context Reconstruction ──────────────────────────────────
            # For short / ambiguous messages, merge with stored last_intent
            # to produce a fully self-contained query before main pipeline.
            original_message = message
            message = self._reconstruct_message(user_id, message, conversation_history)
            if message != original_message:
                msg_lower = message.strip().lower()  # refresh after reconstruction

            # 0c. Payment reminder queries
            reminder_keywords = [
                "remind clients",
                "approaching due",
                "upcoming due",
                "due soon",
            ]
            is_reminder_query = any(k in message.lower() for k in reminder_keywords)
            if is_reminder_query:
                logger.info("[REMINDER] Detected payment reminder query")
                approaching_days = 7
                payment_terms_days = 30
                targets = self.supabase.fetch_reminder_targets(
                    approaching_days=approaching_days,
                    payment_terms_days=payment_terms_days,
                    user_id=data_user_id,
                )
                logger.info(f"[REMINDER] Loaded {len(targets)} reminder targets from Supabase")

                sent = 0
                failed = 0
                sent_details = []
                from datetime import datetime as dt_now
                for t in targets:
                    to_email = (t.get("poc_email") or "").strip()
                    client = (t.get("client_name") or "Client").strip()
                    invoice_number = (t.get("bill_no") or "N/A")
                    if isinstance(invoice_number, (int, float)):
                        invoice_number = str(invoice_number)
                    fees_val = t.get("fees") or 0
                    try:
                        amount_due = f"₹{float(fees_val):,.2f}"
                    except (TypeError, ValueError):
                        amount_due = "₹0.00"
                    due_date_str = (t.get("due_date") or "").strip()[:10] or "N/A"

                    if not to_email:
                        continue
                    ok = self.email.send_payment_reminder(
                        to_email=to_email,
                        client_name=client,
                        invoice_number=invoice_number,
                        amount_due=amount_due,
                        due_date_str=due_date_str,
                    )
                    if ok:
                        row_id = t.get("id")
                        if row_id:
                            self.supabase.update_job_entry_field(row_id, "first_reminder_sent", dt_now.utcnow().isoformat())
                        sent += 1
                        sent_details.append(f"{client} ({invoice_number}) - {to_email}")
                    else:
                        failed += 1

                if not targets:
                    response = format_response(
                        REMINDER_MODE,
                        clarification_hint="Would you like me to check a different window or list overdue items?",
                        reminder_sent_count=0,
                    )
                    self._store_conversation(user_id, message, response)
                    return {
                        "operation": "ACTION_TRIGGER",
                        "response": response,
                        "trigger_invoice": False,
                    }

                response = format_response(
                    REMINDER_MODE,
                    reminder_sent_count=sent,
                    reminder_details=sent_details,
                )
                if failed > 0:
                    response = response.rstrip() + f"\n\nFailed to send: {failed}."
                self._store_conversation(user_id, message, response)
                return {
                    "operation": "ACTION_TRIGGER",
                    "response": response,
                    "trigger_invoice": False,
                }

            # 1b. "Send to client" follow-up — use AI to detect intent, no pattern list
            msg_lower = message.lower()
            cached_invoice = user_mem.get("last_generated_invoice")

            # TTL: expire cached invoice after 30 minutes
            if cached_invoice:
                from datetime import datetime, timedelta
                cached_at = cached_invoice.get("cached_at", "")
                if cached_at:
                    try:
                        cache_time = datetime.fromisoformat(cached_at)
                        if datetime.now() - cache_time > timedelta(minutes=30):
                            logger.info(f"[SEND_CHECK] Cached invoice expired (>30min), clearing")
                            self.memory.update_user_memory(user_id, {"last_generated_invoice": None})
                            cached_invoice = None
                    except (ValueError, TypeError):
                        pass

            if cached_invoice:
                cached_client = cached_invoice.get("client_name", "")
                # Get last bot message for context
                last_bot_msg = ""
                if conversation_history:
                    bot_msgs = [m for m in conversation_history if m.get("role") == "assistant"]
                    if bot_msgs:
                        last_bot_msg = bot_msgs[-1].get("content", "")
                is_send_to_client = self.gemini.is_send_to_client_intent(message, last_bot_msg, cached_client=cached_client)
                logger.info(f"[SEND_CHECK] AI determined send_to_client={is_send_to_client} for msg='{message[:60]}' (cached_client={cached_client})")
            else:
                is_send_to_client = False

            if is_send_to_client and cached_invoice:
                cached_client = cached_invoice.get("client_name", "Client")
                cached_month = cached_invoice.get("month", "Request")
                cached_year = cached_invoice.get("year")
                poc_email = cached_invoice.get("poc_email", "")
                cached_row_ids = cached_invoice.get("row_ids", [])

                if not poc_email:
                    response = (
                        f"I have the invoice for {cached_client} ({cached_month}) ready, "
                        f"but there's no contact email (poc_email) on file.\n\n"
                        f"Please provide the client's email so I can send it:\n"
                        f"Example: client@agency.com"
                    )
                    self.memory.update_user_memory(user_id, {
                        "awaiting_poc_email": True,
                        "pending_send_invoice": {
                            "client_name": cached_client,
                            "month": cached_month,
                            "year": cached_year,
                            "row_ids": cached_row_ids,
                        },
                    })
                    self._store_conversation(user_id, message, response)
                    return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

                # We have the PDF and the email — ask for confirmation
                self.memory.update_user_memory(user_id, {
                    "awaiting_send_confirmation": True,
                    "pending_send_invoice": {
                        "client_name": cached_client,
                        "month": cached_month,
                        "year": cached_year,
                        "poc_email": poc_email,
                        "row_ids": cached_row_ids,
                    },
                })
                response = (
                    f"I have the invoice for {cached_client} ({cached_month}) ready.\n\n"
                    f"Should I email it to **{poc_email}**?\n"
                    f"Reply 'Yes' to send or 'No' to skip."
                )
                logger.info(f"[INVOICE] Using cached invoice for send-to-client: {cached_client} {cached_month}")
                self._store_conversation(user_id, message, response)
                return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 1b-bis. "Send that/it to the client" but NO cached invoice to send.
            # Don't fall through to the query pipeline (which runs a confusing SELECT
            # disambiguation on "client"). Give a clear, actionable message instead.
            if not cached_invoice and re.search(
                r'\bsend\s+(?:that|it|this|the\s+invoice|the\s+bill)\b.*\b(client|them|over|email)\b',
                msg_lower,
            ):
                response = (
                    "I don't have a recently generated invoice to send. "
                    "Generate one first — e.g. 'Generate invoice for Nike for March' — "
                    "and then I can email it to the client."
                )
                self._store_conversation(user_id, message, response)
                return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 1c. Invoice feedback — user complains about the just-generated invoice
            _INVOICE_FEEDBACK_WORDS = ["missing", "wrong", "incorrect", "update", "change",
                                       "fix", "edit", "add", "remove", "doesn't have",
                                       "not showing", "no client", "no billing", "no address"]
            if cached_invoice and not is_send_to_client:
                has_invoice_ref = "invoice" in msg_lower or "bill" in msg_lower or "pdf" in msg_lower
                has_feedback = any(w in msg_lower for w in _INVOICE_FEEDBACK_WORDS)
                if has_invoice_ref and has_feedback:
                    cached_client = cached_invoice.get("client_name", "Client")
                    response = (
                        f"Got it — you'd like to update the invoice for {cached_client}. "
                        f"You can customize the invoice by updating your profile settings. "
                        f"Here's what you can set:\n\n"
                        f"1. Your name/title/address/email on the invoice header — say: "
                        f"\"Update invoice profile\"\n"
                        f"2. Client billing details (billing name, address, GST) — say: "
                        f"\"Update client billing for {cached_client}\"\n\n"
                        f"After updating, say \"Regenerate invoice for {cached_client}\" "
                        f"and I'll create a fresh PDF with the new details."
                    )
                    self._store_conversation(user_id, message, response)
                    return {"operation": "invoice_feedback", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 1c-bis. Invoice feedback with NO cached invoice (prior generation failed or
            # session was fresh). "Invoice is wrong" / "the bill is incorrect" must still be
            # recognised as feedback and answered helpfully — never fall through to the query
            # pipeline (which errors on this phrasing).
            if not cached_invoice and not is_send_to_client:
                _fb_words = ("wrong", "incorrect", "missing", "not right", "is off", "looks off", "error in")
                _has_inv_ref = ("invoice" in msg_lower or re.search(r'\bbill\b', msg_lower) or "pdf" in msg_lower)
                if _has_inv_ref and any(w in msg_lower for w in _fb_words):
                    response = (
                        "Sorry the invoice isn't right. Tell me what's off and I'll fix it:\n\n"
                        "• Wrong amount or missing job → say which client and month\n"
                        "• Wrong header details (your name/address/email) → say \"Update invoice profile\"\n"
                        "• Wrong client billing details → say \"Update client billing for [client]\"\n\n"
                        "Then say \"Regenerate invoice for [client]\" and I'll produce a fresh PDF."
                    )
                    self._store_conversation(user_id, message, response)
                    return {"operation": "invoice_feedback", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 2. Invoice retrieval (keyword-based; use LLM to extract params, fetch from Supabase)
            _INVOICE_VERBS = ["get", "download", "send", "give", "show", "retrieve", "fetch",
                              "generate", "create", "make", "prepare", "need", "want", "share",
                              "regenerate", "redo", "rebuild"]
            has_verb = any(w in msg_lower for w in _INVOICE_VERBS)
            # Also catch common verb typos
            _VERB_TYPOS = ["genrate", "generat", "crete", "creat", "mke", "prepre", "prepar"]
            has_verb = has_verb or any(t in msg_lower for t in _VERB_TYPOS)
            # Catch common invoice/bill typos: invoce, invoic, invoise, incoice, bll, bil
            _INVOICE_TYPOS = ["invoce", "invoic", "invoise", "incoice", "invioce", "invocice"]
            # "bill" must match as a whole word — "billing"/"billed"/"billable" are
            # financial/revenue terms and must NOT route to the invoice pipeline.
            _bill_as_word = bool(re.search(r'\bbill\b', msg_lower))
            has_invoice_word = (
                "invoice" in msg_lower or _bill_as_word or "pdf" in msg_lower
                or any(t in msg_lower for t in _INVOICE_TYPOS)
            )
            # If a keyword match suggests this might be an invoice request, confirm with AI
            # — many queries mention "invoice"/"bill" without actually asking to generate one
            # (e.g. "jobs with invoice_date older than 60 days", "show unpaid invoices",
            # "how many invoices last month"). AI is far more reliable than keyword lists.
            is_retrieval = has_invoice_word
            # Hard-keyword shortcut: if the message starts with a clear action verb
            # followed by "invoice/bill/invoce/etc.", it is unambiguously an invoice
            # action regardless of what any downstream AI classifier says. This prevents
            # v2 READ_QUERY over-triggering on "Generate invoice for X for March".
            _ACTION_VERBS = ("generate", "genrate", "generat", "create", "make", "build",
                             "prepare", "send", "share", "email", "mail", "give", "get",
                             "download", "fetch", "show me the invoice", "regenerate")
            _invoice_action_definite = (
                has_invoice_word
                and (
                    any(msg_lower.startswith(v) for v in _ACTION_VERBS)
                    or any(f"{v} invoice" in msg_lower or f"{v} bill" in msg_lower
                           or any(f"{v} {t}" in msg_lower for t in _INVOICE_TYPOS)
                           for v in _ACTION_VERBS)
                )
            )
            if is_retrieval:
                # First-line guard: defer to the v2 classifier when it confidently
                # called this a READ. The legacy invoice keyword check used to
                # silently override v2 — that was the "kiska invoice baki hai
                # bhejna" bug. v2 has full multilingual context + the column
                # registry's semantic mappings; it should win on reads.
                _v2_says_read = (
                    not _invoice_action_definite  # never downgrade a clear action verb
                    and _v2_verdict is not None
                    and _v2_verdict.get("intent") in ("READ_QUERY", "READ_AGGREGATE")
                    and float(_v2_verdict.get("confidence") or 0) >= 0.85
                )
                _v2_says_invoice = (
                    _v2_verdict is not None
                    and _v2_verdict.get("intent") == "WRITE_INVOICE"
                )
                if _v2_says_read:
                    logger.info(
                        f"[INVOICE_CHECK] v2 classifier confidently said "
                        f"{_v2_verdict.get('intent')} (conf={_v2_verdict.get('confidence')}) — "
                        f"routing to query pipeline."
                    )
                    is_retrieval = False
                    has_invoice_word = False
                elif _invoice_action_definite or _v2_says_invoice:
                    # Clear action verb or v2 confirmed WRITE_INVOICE — skip the
                    # redundant is_invoice_action_request LLM call.
                    logger.info(f"[INVOICE_CHECK] definite={_invoice_action_definite} v2_invoice={_v2_says_invoice} — skipping secondary AI check")
                elif not self.gemini.is_invoice_action_request(message):
                    logger.info(f"[INVOICE_CHECK] AI rejected invoice routing for msg='{message[:80]}' — routing to query pipeline")
                    is_retrieval = False
                    has_invoice_word = False
            logger.info(f"[INVOICE_CHECK] msg='{message[:80]}' has_verb={has_verb} has_invoice={has_invoice_word} is_retrieval={is_retrieval}")
            if is_retrieval:
                # For definite invoice actions (generate/create/send + invoice word), skip
                # parse_user_intent — it sometimes returns GEMINI_ERROR and silently falls
                # through to the query pipeline. Instead extract client and month with a
                # lightweight regex pass, then fall back to parse_user_intent only when that
                # doesn't yield a client name.
                if _invoice_action_definite:
                    _direct_month = None
                    _direct_client = None
                    _direct_bill = None
                    _month_names = {
                        "january": 1, "february": 2, "march": 3, "april": 4,
                        "may": 5, "june": 6, "july": 7, "august": 8,
                        "september": 9, "october": 10, "november": 11, "december": 12,
                        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7,
                        "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
                    }
                    for _mn, _mv in _month_names.items():
                        if f" {_mn}" in msg_lower or msg_lower.endswith(_mn):
                            _direct_month = _mv
                            break
                    # Detect a bill/invoice number BEFORE client extraction so
                    # "Generate invoice for bill INV-001" doesn't treat "bill INV-001"
                    # as a client name. Matches: "bill INV-001", "bill no INV-001",
                    # "invoice INV-001", "#INV-001", or a bare "INV-001"/"INV001".
                    _bill_match = re.search(
                        r'\b(?:bill|invoice|inv)\s*(?:no\.?|number|#)?\s*#?\s*'
                        r'((?:inv[-\s]?)?\d{1,6}|[a-z]{2,5}[-]\d{1,6})\b',
                        msg_lower, re.IGNORECASE,
                    )
                    # Only treat as a bill number if it actually looks like one
                    # (contains a digit and isn't just a plain month/year).
                    if _bill_match:
                        _cand = _bill_match.group(1).strip()
                        if re.search(r'\d', _cand) and not re.fullmatch(r'20\d{2}', _cand):
                            # Normalise to the stored format, e.g. "inv-001" → "INV-001".
                            _norm = _cand.upper().replace(" ", "")
                            if _norm.isdigit():
                                _direct_bill = _norm
                            else:
                                _direct_bill = re.sub(r'^INV-?', 'INV-', _norm)
                            logger.info(f"[INVOICE_SHORTCUT] Detected bill number: {_direct_bill!r}")
                    # Extract client name only when no bill number was found.
                    if not _direct_bill:
                        _for_match = re.search(
                            r'\bfor\s+(.+?)(?:\s+for\s+|\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*|$)',
                            msg_lower, re.IGNORECASE
                        )
                        if _for_match:
                            _cand_client = _for_match.group(1).strip().title()
                            # Pronouns ("them", "it", "this client") are NOT client
                            # names — leave _direct_client null so downstream context
                            # resolution maps them to the remembered client.
                            _PRONOUNS = {
                                "them", "it", "this", "that", "this client", "that client",
                                "the client", "him", "her", "they", "this one", "that one",
                            }
                            if _cand_client.lower() not in _PRONOUNS:
                                _direct_client = _cand_client
                    if _direct_bill:
                        intent_result = {
                            "operation": "ACTION_TRIGGER",
                            "parameters": {"client_name": None, "month": None, "year": None, "bill_number": _direct_bill},
                            "confidence": 0.97,
                            "clarification_question": None,
                        }
                    elif _direct_client:
                        intent_result = {
                            "operation": "ACTION_TRIGGER",
                            "parameters": {"client_name": _direct_client, "month": None, "year": None},
                            "confidence": 0.95,
                            "clarification_question": None,
                        }
                        # Override month from regex
                        if _direct_month:
                            intent_result["parameters"]["month"] = list(_month_names.keys())[
                                list(_month_names.values()).index(_direct_month)].capitalize()
                        logger.info(f"[INVOICE_SHORTCUT] Direct extract: client={_direct_client!r} month={_direct_month}")
                    else:
                        schema_info = logic.get_schema_for_intent() if hasattr(logic, "get_schema_for_intent") else None
                        intent_result = self.gemini.parse_user_intent(message, conversation_history=conversation_history, schema_info=schema_info)
                else:
                    schema_info = logic.get_schema_for_intent() if hasattr(logic, "get_schema_for_intent") else None
                    intent_result = self.gemini.parse_user_intent(message, conversation_history=conversation_history, schema_info=schema_info)
                # Resilience: a definite invoice action ("generate invoice") must NOT fall
                # through to the query pipeline just because the intent LLM flaked
                # (GEMINI_ERROR). Synthesise a minimal ACTION_TRIGGER so downstream context
                # resolution (last_saved_job / uscf_context / last_intent) still runs.
                if _invoice_action_definite and intent_result.get("operation") == "GEMINI_ERROR":
                    logger.info("[INVOICE_SHORTCUT] parse_user_intent flaked on a definite invoice action — synthesising minimal intent for context resolution")
                    intent_result = {
                        "operation": "ACTION_TRIGGER",
                        "parameters": {
                            "client_name": _direct_client,
                            "month": (number_to_month_name(_direct_month) if _direct_month else None),
                            "year": None,
                            "bill_number": _direct_bill,
                        },
                        "confidence": 0.8,
                        "clarification_question": None,
                    }
                params = intent_result.get("parameters", {})
                # If the intent parser asked for clarification, surface that question
                # — but only when the existing flow can't handle it. Missing month or
                # client name is handled downstream with richer context (lists available
                # months, fuzzy-matches typos like 'Samsun' → 'Samsung'), so we let
                # those fall through. Other clarifications (e.g. "which items to keep")
                # are surfaced directly to avoid silent invoice regeneration.
                if intent_result.get("operation") == "NEED_CLARIFICATION":
                    clar_q = (intent_result.get("clarification_question") or "").strip()
                    clar_q_lower = clar_q.lower()
                    _handled_downstream = any(
                        kw in clar_q_lower for kw in ("month", "year", "client name", "client?", "which client")
                    )
                    if clar_q and not _handled_downstream:
                        self._store_conversation(user_id, message, clar_q)
                        return {"operation": "ACTION_TRIGGER", "response": clar_q, "trigger_invoice": False, "invoice_data": {}}

                # AI confirmed the request is out of scope (not query/invoice/action) →
                # use the feature-aware AI responder so we can humorously confirm what
                # Remyndly does/doesn't do, grounded in REMYNDLY_FEATURES.md.
                if intent_result.get("operation") == "UNKNOWN":
                    response = self.gemini.answer_feature_question(message, conversation_history=conversation_history)
                    if not response or not response.strip():
                        response = unsupported_feature_phrase(message[:80])
                    self._store_conversation(user_id, message, response)
                    return {"operation": "unsupported", "response": response, "trigger_invoice": False, "invoice_data": {}}

                if intent_result.get("operation") != "GEMINI_ERROR":
                    # Email-specific override: if user explicitly mentions sending over email,
                    # treat this as SEND_EMAIL instead of a generic ACTION_TRIGGER.
                    email_keywords = [
                        "email invoice",
                        "send invoice over email",
                        "send over email",
                        "mail the invoice",
                        "mail invoice",
                        "share invoice via email",
                        "forward invoice",
                        "send invoice to client",
                        "send it to client",
                        "send to client",
                        "send invoice to the client",
                    ]
                    if "email" in msg_lower or "e-mail" in msg_lower or "to client" in msg_lower or any(k in msg_lower for k in email_keywords):
                        intent_result["operation"] = "SEND_EMAIL"

                    client_name = (params.get("client_name") or "").strip()
                    # A pronoun is never a real client — clear it so context
                    # resolution (last_saved_job / uscf_context / last_intent) maps
                    # "them"/"it"/"this client" to the remembered client.
                    if client_name.lower() in (
                        "them", "it", "this", "that", "this client", "that client",
                        "the client", "him", "her", "they", "this one", "that one",
                    ):
                        client_name = ""
                    month_name = (params.get("month") or "").strip()
                    year_val = params.get("year")
                    bill_number = (params.get("bill_number") or "").strip() or None
                    month_num = month_name_to_number(month_name) if month_name else None

                    # Validate: if user's message contains an explicit year but LLM
                    # missed it, extract it directly from the message text.
                    if not year_val:
                        _year_match = re.search(r'\b(20\d{2})\b', message)
                        if _year_match:
                            year_val = int(_year_match.group(1))
                            logger.info(f"[INVOICE] LLM missed year; extracted {year_val} from message text")
                    if not year_val:
                        from datetime import datetime
                        year_val = datetime.now().year

                    # Fuzzy-match client_name against actual DB clients
                    # (Gemini often normalizes names, e.g. "Bridgestone12" → "Bridgestone")
                    if client_name:
                        safe_uid = data_user_id.replace("'", "''")
                        clients_sql = (
                            f"SELECT DISTINCT client_name, brand_name, production_house FROM public.job_entries "
                            f"WHERE user_id = '{safe_uid}' AND (\"isDeleted\" IS NOT TRUE)"
                        )
                        clients_result = self.supabase.execute_sql(clients_sql)
                        if clients_result.get("ok"):
                            db_clients = []
                            for r in (clients_result.get("rows") or []):
                                for _f in ("client_name", "brand_name", "production_house"):
                                    _v = (r.get(_f) or "").strip() if r.get(_f) else ""
                                    if _v and _v.lower() != "none":
                                        db_clients.append(_v)
                            # De-duplicate (case-insensitive)
                            _seen = set()
                            db_clients = [x for x in db_clients if not (x.lower() in _seen or _seen.add(x.lower()))]

                            cn_lower = client_name.lower()
                            # Exact match first
                            exact = [c for c in db_clients if c.lower() == cn_lower]
                            if exact:
                                client_name = exact[0]
                            else:
                                # SAFETY: short queries (<= 3 chars like "MS", "AB") must
                                # NOT do raw substring matching — "ms" appears inside
                                # "samsung", which is dangerous. Require a word-boundary
                                # match against DB candidates for short queries.
                                if len(cn_lower) <= 3:
                                    _pat = re.compile(rf"\b{re.escape(cn_lower)}\b")
                                    partial = [c for c in db_clients if _pat.search(c.lower())]
                                else:
                                    partial = [
                                        c for c in db_clients
                                        if cn_lower in c.lower() or c.lower() in cn_lower
                                    ]
                                if len(partial) == 1:
                                    logger.info(f"[INVOICE] Fuzzy matched '{client_name}' → '{partial[0]}'")
                                    client_name = partial[0]
                                elif len(partial) > 1:
                                    # Multiple partial matches — check the original message for the best one
                                    msg_low = message.lower()
                                    for p in partial:
                                        if p.lower() in msg_low:
                                            logger.info(f"[INVOICE] Matched '{client_name}' → '{p}' from message text")
                                            client_name = p
                                            break

                    # Resolve "this job" / missing client from context
                    if not client_name and not bill_number:
                        # 1. Check last_saved_job (from smart capture)
                        last_job = user_mem.get("last_saved_job")
                        if last_job:
                            client_name = last_job.get("db_client_name") or last_job.get("brand_name", "")
                            if not month_name and last_job.get("job_date"):
                                try:
                                    job_month = int(last_job["job_date"][5:7])
                                    month_name = number_to_month_name(job_month)
                                    month_num = job_month
                                    year_val = int(last_job["job_date"][:4])
                                except (ValueError, IndexError):
                                    pass
                            logger.info(f"[INVOICE] Resolved from last_saved_job: client={client_name}, month={month_name}")

                    if not client_name and not bill_number:
                        # 2. Check uscf_context (from recent query/update results)
                        ctx = user_mem.get("uscf_context", {})
                        last_row = ctx.get("last_row_data", {})
                        if last_row.get("client_name"):
                            client_name = last_row["client_name"]
                            logger.info(f"[INVOICE] Resolved from uscf_context: client={client_name}")
                        # 3. Check last_intent (from recent interactions)
                        elif user_mem.get("last_intent", {}).get("client_name"):
                            client_name = user_mem["last_intent"]["client_name"]
                            logger.info(f"[INVOICE] Resolved from last_intent: client={client_name}")

                    if not client_name and not bill_number:
                        # Save intent so follow-up can provide client name
                        op_name = intent_result.get("operation", "invoice")
                        self._save_last_intent(user_id, operation=op_name, entity="invoice",
                                               month=month_name, year=year_val,
                                               pending_clarification="client_name")
                        response = "I need a client name or bill number to find an invoice. For example: 'Send invoice for Garnier for March'."
                        self._store_conversation(user_id, message, response)
                        return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

                    # Validate client exists in DB before proceeding
                    # Search client_name, brand_name, and production_house (legacy)
                    if client_name and not bill_number:
                        safe_uid = data_user_id.replace("'", "''")
                        safe_cn = client_name.replace("'", "''")
                        check_sql = (
                            f"SELECT DISTINCT client_name, brand_name, production_house FROM public.job_entries "
                            f"WHERE user_id = '{safe_uid}' "
                            f"AND (client_name ILIKE '%{safe_cn}%' OR brand_name ILIKE '%{safe_cn}%' OR production_house ILIKE '%{safe_cn}%') "
                            f"AND (\"isDeleted\" IS NOT TRUE)"
                        )
                        check_result = self.supabase.execute_sql(check_sql)
                        # Fallback if production_house column doesn't exist
                        if not check_result.get("ok") and "production_house" in str(check_result.get("error", "")):
                            check_sql = (
                                f"SELECT DISTINCT client_name, brand_name FROM public.job_entries "
                                f"WHERE user_id = '{safe_uid}' AND (client_name ILIKE '%{safe_cn}%' OR brand_name ILIKE '%{safe_cn}%') "
                                f"AND (\"isDeleted\" IS NOT TRUE)"
                            )
                            check_result = self.supabase.execute_sql(check_sql)
                        # A row matches if ANY of client_name / brand_name / production_house
                        # is non-null — jobs added with only a brand (e.g. '+Sunrich ...')
                        # have client_name=NULL but brand_name='Sunrich' and must still be found.
                        matching_clients = []
                        for r in (check_result.get("rows") or []):
                            for _f in ("client_name", "brand_name", "production_house"):
                                _v = (r.get(_f) or "").strip() if r.get(_f) else ""
                                if _v:
                                    matching_clients.append(_v)
                                    break
                        if not matching_clients:
                            # No matching client — show available clients and stop
                            all_clients_sql = (
                                f"SELECT DISTINCT client_name FROM public.job_entries "
                                f"WHERE user_id = '{safe_uid}' AND client_name IS NOT NULL "
                                f"AND (\"isDeleted\" IS NOT TRUE) ORDER BY client_name"
                            )
                            all_result = self.supabase.execute_sql(all_clients_sql)
                            available = [r["client_name"] for r in (all_result.get("rows") or []) if r.get("client_name")]
                            if available:
                                client_list = "\n".join(f"• {c}" for c in available)
                                response = (
                                    f"I couldn't find a client named \"{client_name}\". "
                                    f"Please check for typos.\n\n"
                                    f"Your clients on record:\n{client_list}\n\n"
                                    f"Try again with the correct name."
                                )
                            else:
                                response = f"I couldn't find a client named \"{client_name}\" and you don't have any job entries yet."
                            logger.info(f"[INVOICE] Client '{client_name}' not found for user {user_id}")
                            self._store_conversation(user_id, message, response)
                            return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

                    if client_name and not month_num and not bill_number:
                        # Check if user explicitly asked to send via email
                        send_email = "email" in msg_lower or "e-mail" in msg_lower
                        months_result = self.supabase.get_available_months_for_client(client_name, user_id=data_user_id)
                        _months = months_result.get("months") or []
                        # If only one month exists for this client, auto-proceed without prompting.
                        if months_result.get("ok") and len(_months) == 1:
                            try:
                                month_num = int(_months[0]["month"])
                                year_val = int(_months[0]["year"])
                                month_name = number_to_month_name(month_num)
                                logger.info(f"[INVOICE] Single available month — auto-using {month_name} {year_val} for {client_name}")
                            except (KeyError, ValueError, TypeError) as _e:
                                logger.warning(f"[INVOICE] Could not auto-pick single month: {_e}")
                        if months_result.get("ok") and _months and not month_num:
                            month_options = "\n".join(f"• {m['label']}" for m in _months)
                            response = f"I see you want an invoice for {client_name}. Which month?\n\n{month_options}\n\nReply with the month, e.g. 'Send invoice for {client_name} for March 2025'."
                        elif not month_num:
                            response = f"I see you want an invoice for {client_name}. Which month? For example: 'Send invoice for {client_name} for March'."
                        # Only prompt the user when month is still unresolved.
                        # If we auto-picked the single available month above, fall through
                        # to invoice generation.
                        if not month_num:
                            # Save intent so follow-up "March" reconstructs to full query
                            # Do NOT store inferred year — only store confirmed fields.
                            op_name = "SEND_EMAIL" if send_email else intent_result.get("operation", "invoice")
                            self._save_last_intent(user_id, operation=op_name, client_name=client_name,
                                                   entity="invoice",
                                                   pending_clarification="month")
                            # Set awaiting state so the next reply routes to invoice month handler
                            self.memory.update_user_memory(user_id, {
                                "awaiting_invoice_month": True,
                                "pending_invoice_client": client_name,
                                "pending_invoice_send_email": send_email,
                            })
                            self._store_conversation(user_id, message, response)
                            return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

                    if bill_number:
                        result = self.supabase.fetch_job_entries_for_invoice(client_name="", bill_no=bill_number, user_id=data_user_id)
                    else:
                        result = self.supabase.fetch_job_entries_for_invoice(client_name=client_name, month=month_num, year=year_val, user_id=data_user_id)
                    if not result.get("ok"):
                        response = result.get("error", "I couldn't fetch invoice data. Please try again.")
                        self._store_conversation(user_id, message, response)
                        return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}
                    rows = result.get("rows") or []
                    if not rows:
                        # Check what months actually have data for this client
                        hint = ""
                        if client_name:
                            safe_client = client_name.replace("'", "''")
                            safe_uid = data_user_id.replace("'", "''")
                            avail_sql = (
                                f"SELECT DISTINCT TO_CHAR(job_date, 'Month YYYY') AS period "
                                f"FROM public.job_entries "
                                f"WHERE user_id = '{safe_uid}' "
                                f"AND (client_name ILIKE '%{safe_client}%' OR brand_name ILIKE '%{safe_client}%' OR production_house ILIKE '%{safe_client}%') "
                                f"AND job_date IS NOT NULL AND (\"isDeleted\" IS NOT TRUE) ORDER BY period"
                            )
                            avail = self.supabase.execute_sql(avail_sql)
                            if not avail.get("ok") and "production_house" in str(avail.get("error", "")):
                                avail_sql = (
                                    f"SELECT DISTINCT TO_CHAR(job_date, 'Month YYYY') AS period "
                                    f"FROM public.job_entries "
                                    f"WHERE user_id = '{safe_uid}' AND (client_name ILIKE '%{safe_client}%' OR brand_name ILIKE '%{safe_client}%') "
                                    f"AND job_date IS NOT NULL AND (\"isDeleted\" IS NOT TRUE) ORDER BY period"
                                )
                                avail = self.supabase.execute_sql(avail_sql)
                            periods = [r["period"].strip() for r in (avail.get("rows") or [])]
                            if periods:
                                hint = f"\n\nI do have records for {client_name} in: {', '.join(periods)}."
                            else:
                                # No DATED records — but the client may still have rows
                                # with a NULL job_date. Don't claim "no records at all"
                                # without checking the actual row count first.
                                _cnt_sql = (
                                    f"SELECT COUNT(*) AS cnt FROM public.job_entries "
                                    f"WHERE user_id = '{safe_uid}' "
                                    f"AND (client_name ILIKE '%{safe_client}%' OR brand_name ILIKE '%{safe_client}%' OR production_house ILIKE '%{safe_client}%') "
                                    f"AND (\"isDeleted\" IS NOT TRUE)"
                                )
                                _cnt_res = self.supabase.execute_sql(_cnt_sql)
                                _cnt = 0
                                if _cnt_res.get("ok") and _cnt_res.get("rows"):
                                    try:
                                        _cnt = int(_cnt_res["rows"][0].get("cnt", 0))
                                    except (ValueError, TypeError):
                                        _cnt = 0
                                if _cnt > 0:
                                    hint = (
                                        f"\n\nI have {_cnt} record{'s' if _cnt != 1 else ''} for {client_name}, "
                                        f"but none have a job date set, so I can't filter by month. "
                                        f"Try 'Generate invoice for {client_name}' without a month."
                                    )
                                else:
                                    hint = f"\n\nI don't have any records for {client_name} at all."
                        if client_name and month_num:
                            response = f"I found no jobs for {client_name} in {month_name or month_num} {year_val}.{hint}"
                        else:
                            response = f"I don't see any records for {client_name or 'that bill'} in my records.{hint}"
                        # If exactly one alternate month is available, persist it so a short
                        # confirmation ("okay generate", "yes", "sure") can act on it.
                        if client_name and periods and len(periods) == 1:
                            try:
                                _alt = periods[0]  # e.g. "February  2026"
                                _parts = _alt.split()
                                _alt_month = _parts[0] if _parts else None
                                _alt_year = int(_parts[-1]) if _parts and _parts[-1].isdigit() else year_val
                                if _alt_month:
                                    self._save_last_intent(
                                        user_id,
                                        operation="generate_invoice",
                                        client_name=client_name,
                                        month=_alt_month,
                                        year=_alt_year,
                                        entity="invoice",
                                        pending_clarification="confirm_alt_month",
                                    )
                            except Exception as _e:
                                logger.warning(f"[INVOICE] Could not persist alt-month context: {_e}")
                        self._store_conversation(user_id, message, response)
                        return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}
                    display_client = (rows[0].get("client_name") or client_name or "Client").strip()
                    month_display = month_name
                    if not month_display and rows and rows[0].get("job_date"):
                        jd = str(rows[0]["job_date"])[:10]
                        if len(jd) >= 7:
                            try:
                                month_display = number_to_month_name(int(jd[5:7]))
                            except (ValueError, TypeError):
                                pass
                    if not month_display:
                        month_display = "Request"
                    # Detect explicit "regenerate" intent — only then bypass the cached PDF.
                    _regen_keywords = (
                        "regenerate", "regen ", "regen.", "re-generate", "re generate",
                        "fresh copy", "fresh invoice", "new copy", "new pdf",
                        "redo invoice", "remake invoice", "rebuild invoice", "recreate invoice",
                        "force regenerate", "generate again", "make it again",
                    )
                    _force_regen = any(kw in msg_lower for kw in _regen_keywords)
                    invoice_data = {
                        "client_name": display_client,
                        "month": month_display,
                        "bill_number": bill_number,
                        "year": year_val,
                        "force_regenerate": _force_regen,
                    }
                    if _force_regen:
                        logger.info(f"[INVOICE] User requested regeneration — bypassing cache for {display_client} {month_display} {year_val}")

                    # Extract poc_name and invoicer_name for email personalization
                    _inv_poc_name = ""
                    for _r in rows:
                        _v = (_r.get("poc_name") or "").strip()
                        if _v and _v.lower() != "none":
                            _inv_poc_name = _v
                            break
                    _inv_invoicer_name = ""
                    try:
                        _prof = self.supabase.get_user_profile(data_user_id)
                        if _prof.get("ok") and _prof.get("data"):
                            _prefs = _prof["data"].get("preferences") or {}
                            if isinstance(_prefs, str):
                                import json as _json
                                try:
                                    _prefs = _json.loads(_prefs)
                                except Exception:
                                    _prefs = {}
                            _inv_invoicer_name = _prefs.get("invoice_name") or _prof["data"].get("name") or ""
                    except Exception:
                        pass

                    # ── Mandatory-fields gate ──────────────────────────────────
                    # Prompt for any missing required field (client billing, POC
                    # name, job description, bank account, business address) and
                    # only proceed once the invoice is complete. Runs for BOTH the
                    # generate and the email paths; each field's handler re-enters
                    # the flow so the prompts chain until everything is present.
                    _gate = self._invoice_readiness_check(user_id, data_user_id, invoice_data, rows)
                    if _gate is not None:
                        return _gate

                    # Decide between generating/sending invoice via WhatsApp/Telegram vs email
                    if intent_result.get("operation") == "SEND_EMAIL":
                        poc_email = (rows[0].get("poc_email") or "").strip()
                        if not poc_email:
                            row_ids = [r["id"] for r in rows if r.get("id")]
                            response = (
                                f"I have the invoice for {display_client} ({month_display}) ready, "
                                f"but there's no contact email on file.\n\n"
                                f"Please provide the client's email so I can send it:\n"
                                f"Example: client@agency.com"
                            )
                            self.memory.update_user_memory(user_id, {
                                "awaiting_poc_email": True,
                                "pending_send_invoice": {
                                    "client_name": display_client,
                                    "month": month_display,
                                    "year": year_val,
                                    "row_ids": row_ids,
                                    "poc_name": _inv_poc_name,
                                    "invoicer_name": _inv_invoicer_name,
                                },
                            })
                            self._store_conversation(user_id, message, response)
                            return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

                        # PDF delivery + email confirmation prompt are owned by
                        # main.py.process_and_send_invoice — don't duplicate them here.
                        invoice_data["send_to_client"] = True
                        trigger_invoice = True
                        response = f"On it — generating your invoice for {display_client} ({month_display}) now…"
                        self._store_conversation(user_id, message, response)
                        return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": trigger_invoice, "invoice_data": invoice_data}

                    # (All mandatory-field checks now run in _invoice_readiness_check above.)

                    # Default path: generate PDF, deliver via WhatsApp/Telegram, then
                    # automatically offer to email it to the POC. Two cases:
                    #   1) POC email on file  → ask "Should I email it to X?"
                    #   2) No POC email       → ask the user to provide one
                    trigger_invoice = True
                    self._save_last_intent(user_id, operation="generate_invoice",
                                           client_name=display_client, month=month_display,
                                           year=year_val, entity="invoice")

                    _row_ids = [r["id"] for r in rows if r.get("id")]
                    _auto_poc_email = ""
                    for _r in rows:
                        _e = (str(_r.get("poc_email") or "")).strip()
                        if _e and self._is_valid_email(_e):
                            _auto_poc_email = _e
                            break

                    if _auto_poc_email:
                        # PDF delivery + email confirmation prompt are owned by
                        # main.py.process_and_send_invoice — don't duplicate them here.
                        # Just acknowledge so the user has feedback during PDF generation.
                        response = f"On it — generating your invoice for {display_client} ({month_display}) now…"
                    else:
                        self.memory.update_user_memory(user_id, {
                            "awaiting_poc_email": True,
                            "pending_send_invoice": {
                                "client_name": display_client,
                                "month": month_display,
                                "year": year_val,
                                "row_ids": _row_ids,
                                "poc_name": _inv_poc_name,
                                "invoicer_name": _inv_invoicer_name,
                            },
                        })
                        response = (
                            f"On it — generating your invoice for {display_client} ({month_display}) now.\n\n"
                            f"Heads up: no client email on file. Reply with one (e.g. client@agency.com) "
                            f"and I'll send it over, or 'skip' to keep it offline."
                        )

                    self._store_conversation(user_id, message, response)
                    return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": trigger_invoice, "invoice_data": invoice_data}

            # 3. Overdue / payment followup (keyword-based; data from Supabase)
            overdue_keywords = ["overdue", "due date", "passed due", "past due", "late payment", "follow up", "followup", "payment followup", "payment status"]
            _ml = message.lower()
            is_overdue = any(k in _ml for k in overdue_keywords) and ("invoice" in _ml or "client" in _ml or "payment" in _ml)
            # "Remind clients about payments" / "send payment reminders" / "remind everyone
            # to pay" — a manual request to chase payments. Route to the same overdue
            # handler, which lists due invoices and offers to send reminders.
            _wants_remind = bool(re.search(r'\bremind(?:er|ers)?\b', _ml)) and (
                "payment" in _ml or "pay" in _ml or "client" in _ml or "invoice" in _ml or "due" in _ml or "everyone" in _ml
            )
            if is_overdue or _wants_remind:
                overdue_jobs = self.supabase.fetch_overdue_jobs(payment_terms_days=30, user_id=user_id)
                if not overdue_jobs:
                    response = "Great news! I don't see any invoices that have passed their due date."
                else:
                    lines = [f"I found {len(overdue_jobs)} invoice(s) past due:\n"]
                    for j in overdue_jobs[:20]:
                        client = (j.get("client_name") or "Unknown").strip()
                        due = (j.get("due_date") or "")[:10]
                        bill = j.get("bill_no") or ""
                        lines.append(f"• {client}" + (f" (Due: {due})" if due else "") + (f" — Bill #{bill}" if bill else ""))
                    lines.append("\nWant me to send payment reminders to these clients? Reply 'yes' and I'll draft them.")
                    response = "\n".join(lines)
                    # Remember which clients are pending so a "yes" reply can act on them.
                    self.memory.update_user_memory(user_id, {
                        "pending_reminder_offer": [
                            (j.get("client_name") or "").strip()
                            for j in overdue_jobs[:20] if (j.get("client_name") or "").strip()
                        ],
                    })
                self._store_conversation(user_id, message, response)
                return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 3b. Delete intent → soft-delete (SET "isDeleted" = true)
            _DELETE_TRIGGERS = ["delete", "remove", "erase", "trash", "discard"]
            is_delete = any(w in message.lower() for w in _DELETE_TRIGGERS) and any(
                w in message.lower() for w in ["job", "entry", "record", "row", "last", "this", "it", "that"]
            )
            if is_delete:
                return self._handle_soft_delete(user_id, message, data_user_id, conversation_history)

            # 4. SQL path: intent → generate SQL → validate → execute on Supabase → format → response
            columns = [c for c in JOB_ENTRIES_COLUMNS if not c.startswith("_")]

            if not self.supabase.db_url:
                response = format_response(
                    ERROR_MODE,
                    error_detail="Query service isn't configured right now. I can still help with payment reminders and invoice retrieval.",
                )
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # 4a. Check for invoice confirmation (Yes/No after being asked)
            msg_lower = message.strip().lower()
            if msg_lower in ("yes", "y", "sure", "ok", "okay", "please do", "generate", "create"):
                # Check if we recently asked about invoice generation
                conv = self.memory.get_conversation_history(user_id)
                if conv and len(conv) >= 2:
                    last_assistant = conv[-1].get("content", "").lower() if conv[-1].get("role") == "assistant" else ""

                    # "Would you like to see more/full details?" → re-run last query as SELECT *
                    _detail_markers = ("more details", "see details", "full details", "see more", "more information", "see other jobs", "would you like")
                    if any(m in last_assistant for m in _detail_markers) and "generate an invoice" not in last_assistant:
                        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
                        last_sql = ctx.get("last_sql")
                        if last_sql:
                            import re as _re
                            full_sql = _re.sub(r"^SELECT\s+.+?\s+FROM\s", "SELECT * FROM ", last_sql, count=1, flags=_re.IGNORECASE | _re.DOTALL)
                            logger.info(f"[FOLLOWUP] User confirmed details; re-running as SELECT *: {full_sql[:200]}")
                            exec_result = self.supabase.execute_sql(full_sql)
                            if exec_result.get("ok"):
                                rows = exec_result.get("rows", [])
                                if rows:
                                    self._update_sql_context(user_id, rows)
                                    ctx["last_sql"] = full_sql
                                    self.memory.update_user_memory(user_id, {"uscf_context": ctx})
                                    payload = build_clean_payload(rows, "select")
                                    response = self.gemini.synthesize_response(payload, message, conversation_history=conversation_history)
                                    if not response or not response.strip():
                                        response = "Here are the full details for your records."
                                    self._store_conversation(user_id, message, response)
                                    return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

                    if "generate an invoice" in last_assistant and "would you like" in last_assistant:
                        # Generate invoice for the last job we found
                        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
                        last_row = ctx.get("last_row_data") if ctx else None
                        if last_row:
                            client_name = last_row.get("client_name", "Client")
                            job_date = last_row.get("job_date")
                            month = None
                            year = None
                            if job_date:
                                try:
                                    from datetime import datetime
                                    if isinstance(job_date, str):
                                        job_dt = datetime.fromisoformat(job_date[:10])
                                    else:
                                        job_dt = job_date
                                    month = job_dt.strftime("%B")
                                    year = job_dt.year
                                except:
                                    pass
                            
                            invoice_data = {
                                "client_name": client_name,
                                "month": month or "Period",
                                "bill_number": None,
                                "year": year
                            }
                            response = f"Generating invoice for {client_name}..."
                            self._store_conversation(user_id, message, response)
                            return {
                                "operation": "ACTION_TRIGGER",
                                "response": response,
                                "trigger_invoice": True,
                                "invoice_data": invoice_data
                            }

            # 4b. Follow-up: answer from last result row via AI synthesis (no raw field:value)
            followup_answer = self._try_answer_from_context(user_id, message, columns)
            if followup_answer:
                logger.info(f"[FOLLOWUP] Answered from context (synthesized)")
                response = followup_answer
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # ── Deterministic-first routing ──────────────────────────────────
            # The ~20 common query shapes (counts, sums, lists, top-N, by-client,
            # paid/unpaid, date lookups) are mapped straight to SQL by the query
            # router — no LLM guesswork. This is far more reliable than the planner
            # for the high-frequency cases; the planner stays as the fallback below
            # for the long tail. See services/query_router.py for the route table.
            _routed = route_common_query(message, data_user_id)
            if _routed is not None:
                _routed_result = self._execute_routed_query(
                    _routed, user_id, data_user_id, message, conversation_history,
                )
                if _routed_result is not None:
                    return _routed_result
                # None → router matched but DB returned nothing usable; fall through
                # to the planner so the user still gets a best-effort answer.

            # Generate SQL via query planner pipeline (Classify → Plan → Resolve → Validate → SQL)
            conv_ctx = user_mem.get("uscf_context") or {}
            conv_ctx["last_saved_job"] = user_mem.get("last_saved_job")
            logger.info(f"[PIPELINE] Starting query plan for user {user_id}: {message[:100]}")
            plan_result = execute_query_plan(
                message, self.gemini, self.supabase,
                conversation_history, user_id=data_user_id,
                conversation_context=conv_ctx,
            )
            logger.info(f"[PIPELINE] Plan result: sql={'yes' if plan_result.get('sql') else 'no'}, error={plan_result.get('_error')}, clarification={plan_result.get('clarification')}")

            # Handle clarification from planner
            if plan_result.get("clarification"):
                plan_data = plan_result.get("plan", {}) if isinstance(plan_result.get("plan"), dict) else {}
                _clar_client = plan_data.get("filters", {}).get("client_name", "") if isinstance(plan_data.get("filters"), dict) else ""

                # If the planner is asking about invoice generation (it leaked into query pipeline),
                # redirect to the proper invoice month-selection flow instead of showing a confusing
                # AI-generated clarification (which may suggest unsupported options like "all outstanding").
                _clar_text = plan_result["clarification"].lower()
                _msg_has_invoice = any(t in msg_lower for t in ["invoice", "invoce", "invoic", "bill", "pdf"])
                if _clar_client and _msg_has_invoice:
                    months_result = self.supabase.get_available_months_for_client(_clar_client, user_id=data_user_id)
                    if months_result.get("ok") and months_result.get("months"):
                        month_options = "\n".join(f"• {m['label']}" for m in months_result["months"])
                        response = f"I see you want an invoice for {_clar_client}. Which month?\n\n{month_options}\n\nReply with the month."
                    else:
                        response = f"I see you want an invoice for {_clar_client}. Which month? For example: 'Invoice for {_clar_client} for March'."
                    self._save_last_intent(user_id, operation="invoice", client_name=_clar_client, entity="invoice", pending_clarification="month")
                    self.memory.update_user_memory(user_id, {
                        "awaiting_invoice_month": True,
                        "pending_invoice_client": _clar_client,
                        "pending_invoice_send_email": False,
                    })
                    self._store_conversation(user_id, message, response)
                    return {"operation": "ACTION_TRIGGER", "response": response, "trigger_invoice": False, "invoice_data": {}}

                # For UPDATE clarifications the planner often hallucinates "multiple records found"
                # without ever hitting the DB. Do a real pre-check and handle it properly.
                _plan_op = plan_data.get("operation", "")
                _plan_updates = plan_data.get("updates") if isinstance(plan_data.get("updates"), dict) else {}
                if _plan_op == "update" and _clar_client and _plan_updates:
                    _safe_clar_client = _clar_client.replace("'", "''")
                    _uid_safe = data_user_id.replace("'", "''")
                    _not_del = '("isDeleted" IS NOT TRUE)'
                    _pre_check_sql = (
                        f"SELECT id, client_name, brand_name, job_date, job_description_details, fees "
                        f"FROM public.job_entries "
                        f"WHERE user_id = '{_uid_safe}' AND {_not_del} "
                        f"AND (client_name ILIKE '%{_safe_clar_client}%' "
                        f"  OR brand_name ILIKE '%{_safe_clar_client}%' "
                        f"  OR production_house ILIKE '%{_safe_clar_client}%')"
                    )
                    _pre_check = self.supabase.execute_sql(_pre_check_sql)
                    _pre_rows = _pre_check.get("rows", []) if _pre_check.get("ok") else []
                    logger.info(f"[UPDATE_CLARIFY] Real DB pre-check for '{_clar_client}': {len(_pre_rows)} rows")

                    if len(_pre_rows) == 1:
                        # Exactly one — skip the AI clarification entirely, build and run UPDATE directly
                        _target_id = _pre_rows[0]["id"]
                        _set_clauses = ", ".join(
                            f"{col} = '{val}'" for col, val in _plan_updates.items()
                        )
                        # Normalise paid value
                        _set_clauses = re.sub(r"\bpaid\s*=\s*'(?:true|1|yes)'",  "paid = 'Yes'", _set_clauses, flags=re.IGNORECASE)
                        _set_clauses = re.sub(r"\bpaid\s*=\s*'(?:false|0|no)'",  "paid = 'No'",  _set_clauses, flags=re.IGNORECASE)
                        _direct_sql = (
                            f"UPDATE public.job_entries SET {_set_clauses} "
                            f"WHERE id = '{_target_id}' RETURNING *"
                        )
                        _direct_exec = self.supabase.execute_sql(_direct_sql)
                        if _direct_exec.get("ok") and _direct_exec.get("rowcount", 0) > 0:
                            _direct_rows = _direct_exec.get("rows", [])
                            if _direct_rows:
                                self._update_sql_context(user_id, _direct_rows)
                            _payload = build_clean_payload(_direct_rows or _pre_rows, "select")
                            response = self.gemini.synthesize_response(_payload, message, conversation_history=conversation_history)
                            if not response or not response.strip():
                                response = f"Done! Updated {_direct_exec.get('rowcount', 1)} record."
                            self._store_conversation(user_id, message, response)
                            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

                    elif len(_pre_rows) > 1:
                        # Genuinely multiple — show real numbered list.
                        # Inline values into SET (no %s placeholders) — execute_sql
                        # doesn't bind params here, so unbound %s breaks Postgres.
                        _set_clauses_disambig = ", ".join(
                            f"{col} = '{str(val).replace(chr(39), chr(39)*2)}'"
                            for col, val in _plan_updates.items()
                        )
                        _set_clauses_disambig = re.sub(
                            r"\bpaid\s*=\s*'(?:true|1|yes)'", "paid = 'Yes'",
                            _set_clauses_disambig, flags=re.IGNORECASE,
                        )
                        _set_clauses_disambig = re.sub(
                            r"\bpaid\s*=\s*'(?:false|0|no)'", "paid = 'No'",
                            _set_clauses_disambig, flags=re.IGNORECASE,
                        )
                        self.memory.update_user_memory(user_id, {
                            "pending_disambiguation": {
                                "sql": (
                                    f"UPDATE public.job_entries "
                                    f"SET {_set_clauses_disambig} "
                                    f"WHERE id = '{{id}}' RETURNING *"
                                ),
                                "rows": _pre_rows,
                                "data_user_id": data_user_id,
                                "updates": _plan_updates,
                            }
                        })
                        _opts = [f"I found {len(_pre_rows)} matching records. Which one did you mean?\n"]
                        for _i, _r in enumerate(_pre_rows[:10], 1):
                            _parts = [f"{_i}."]
                            _c = (_r.get("client_name") or _r.get("brand_name") or "").strip()
                            _d = str(_r.get("job_date") or "")[:10]
                            _desc = str(_r.get("job_description_details") or "")[:40].strip()
                            _f = _r.get("fees")
                            if _c: _parts.append(_c)
                            if _d: _parts.append(_d)
                            if _desc: _parts.append(_desc)
                            if isinstance(_f, (int, float)): _parts.append(f"₹{int(_f):,}")
                            _opts.append(" | ".join(_parts))
                        _opts.append("\nReply with a number to pick, or 'cancel' to abort.")
                        response = "\n".join(_opts)
                        self._store_conversation(user_id, message, response)
                        return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

                    else:
                        # 0 rows — genuine not found
                        response = f"I couldn't find any job matching '{_clar_client}'. Could you double-check the name?"
                        self._store_conversation(user_id, message, response)
                        return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

                # Normal clarification for non-invoice / non-update queries
                self._save_last_intent(
                    user_id, operation=plan_data.get("operation", "query"),
                    client_name=_clar_client,
                    entity="query",
                    pending_clarification="details",
                )
                # Generic / off-brand clarifications (e.g. "I'm a spreadsheet assistant…")
                # are useless to the user — route through the feature-aware AI responder
                # which has the catalog as a truth source.
                raw_clar = plan_result["clarification"]
                _off_brand_markers = (
                    "spreadsheet assistant", "spreadsheet or database",
                    "i'm a spreadsheet", "data assistant",
                    "how can i help you with your data",
                    "internal identifier", "malformed request",
                    "cannot fulfill this request",
                )
                _is_off_brand = any(m in raw_clar.lower() for m in _off_brand_markers)
                if _is_off_brand:
                    on_brand = self.gemini.answer_feature_question(
                        message, conversation_history=conversation_history
                    )
                    response = on_brand or unsupported_feature_phrase(message[:80])
                else:
                    response = raw_clar
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

            sql = plan_result.get("sql")
            planner_failed = plan_result.get("_error") or not sql

            # Detect history questions once here; used later for SELECT * rewrite and synthesis hint.
            _is_history_q = self.gemini.is_history_question(message)

            # Fallback to direct SQL generation if planner fails
            if planner_failed:
                logger.info(f"[PIPELINE] Planner failed ({plan_result.get('_error')}), falling back to direct SQL generation")
                sql_result = generate_sql(message, self.gemini, self.supabase, conversation_history, user_id=data_user_id)
                if sql_result.get("_error"):
                    logger.warning(f"[PIPELINE] Fallback also failed for user {user_id}: {sql_result.get('_error')}")
                    # Third fallback: deterministic keyword-based SQL (no LLM needed)
                    sql = self._keyword_sql_fallback(message, data_user_id)
                    if not sql:
                        response = clarify_phrase(["How many jobs?", "Total fees for Garnier", "Last payment date"])
                        self._store_conversation(user_id, message, response)
                        return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
                    logger.info(f"[PIPELINE] Keyword fallback generated SQL: {sql[:200]}")
                else:
                    sql = sql_result.get("sql")

            # Normalise `paid` values to 'Yes'/'No' regardless of what the AI wrote
            # (the column is text; existing rows use 'Yes'/'No', not 'true'/'false'/booleans)
            sql = re.sub(r"SET\s+paid\s*=\s*'(?:true|1|yes)'",  "SET paid = 'Yes'",  sql, flags=re.IGNORECASE)
            sql = re.sub(r"SET\s+paid\s*=\s*'(?:false|0|no)'",  "SET paid = 'No'",   sql, flags=re.IGNORECASE)

            # History questions need SELECT * so that the `notes` change-log reaches
            # Gemini. But ONLY for plain row SELECTs — NEVER for aggregate / GROUP BY
            # queries: rewriting "SELECT job_description, AVG(fees) AS result ... ORDER
            # BY result" to "SELECT *" drops the aliased aggregate while ORDER BY/HAVING
            # still reference it → Postgres "column result does not exist". (is_history_question
            # also false-positives on phrases like "all time", so this guard matters.)
            _is_agg_sql = _is_aggregate_sql(sql)
            if _is_history_q and sql and sql.upper().lstrip().startswith("SELECT") and not _is_agg_sql:
                sql = re.sub(r"(?i)^\s*SELECT\s+(?!\*).+?\s+FROM\s+", "SELECT * FROM ", sql, count=1)
                logger.info("[PIPELINE] History question — rewrote SELECT to SELECT *")
            elif _is_history_q and _is_agg_sql:
                logger.info("[PIPELINE] History question but aggregate/GROUP BY SQL — skipping SELECT * rewrite (would break the aggregate alias)")

            # Expand AI-generated `client_name ILIKE 'X'` → `(client_name ILIKE '%X%' OR brand_name ILIKE '%X%')`
            # Users say "Nike" meaning the brand; the actual client_name may be a production company.
            # Also adds wildcard wrapping so partial names still match.
            def _expand_client_ilike(m):
                val = m.group(1)
                # Strip any existing % wildcards the AI may have added
                val = val.strip('%')
                return f"(client_name ILIKE '%{val}%' OR brand_name ILIKE '%{val}%' OR production_house ILIKE '%{val}%')"
            sql = re.sub(r"\bclient_name\s+ILIKE\s+'([^']*)'", _expand_client_ilike, sql, flags=re.IGNORECASE)

            # Preserve filter context columns in the projection.
            # When SQL filters by a semantic column (paid, invoice_date, *_reminder_sent)
            # but the SELECT drops it, the synthesizer sees a bare list and contradicts
            # itself ("I don't know which paid"). Append the filter column to SELECT
            # so the payload carries the meaning. This catches both planner SQL and
            # the direct generate_sql fallback (planner didn't catch the latter).
            try:
                _CONTEXT_COLS = ["paid", "invoice_date",
                                 "first_reminder_sent", "second_reminder_sent",
                                 "third_reminder_sent"]
                _select_match = re.match(
                    r"^\s*SELECT\s+(.+?)\s+FROM\s+",
                    sql, flags=re.IGNORECASE | re.DOTALL,
                )
                _where_match = re.search(
                    r"\bWHERE\s+(.+?)(?=\bGROUP\s+BY\b|\bORDER\s+BY\b|\bLIMIT\b|\bHAVING\b|$)",
                    sql, flags=re.IGNORECASE | re.DOTALL,
                )
                if _select_match and _where_match:
                    _proj = _select_match.group(1)
                    _where_clause = _where_match.group(1)
                    _proj_lower = _proj.lower()
                    _is_star = _proj.strip() == "*"
                    _is_agg = bool(re.search(r"\b(COUNT|SUM|AVG|MIN|MAX)\s*\(", _proj, re.IGNORECASE))
                    _has_group_by = bool(re.search(r"\bGROUP\s+BY\b", sql, re.IGNORECASE))
                    if not _is_star and not _is_agg:
                        _to_add = []
                        for _col in _CONTEXT_COLS:
                            if re.search(rf"\b{_col}\b", _where_clause, re.IGNORECASE) \
                               and not re.search(rf"\b{_col}\b", _proj, re.IGNORECASE):
                                # In GROUP BY queries, wrap with MAX() so the column
                                # is aggregate-safe. Since the WHERE already filters
                                # to a single semantic state (e.g. paid='Yes'),
                                # MAX(paid) returns that state per group.
                                if _has_group_by:
                                    _to_add.append(f"MAX({_col}) AS {_col}")
                                else:
                                    _to_add.append(_col)
                        if _to_add:
                            _new_proj = _proj.rstrip() + ", " + ", ".join(_to_add)
                            sql = sql.replace(_select_match.group(0),
                                              f"SELECT {_new_proj} FROM ", 1)
                            logger.info(f"[CONTEXT_COLS] Appended {_to_add} to SELECT for synthesizer context")
            except Exception as _e:
                logger.warning(f"[CONTEXT_COLS] Rewrite failed (non-fatal): {_e}")

            valid, sanitized_sql, err = validate_sql(sql)
            if not valid:
                logger.warning(f"[QUERY_FAIL] SQL validation failed for user {user_id}: {err} | SQL: {sql[:200]}")
                response = query_invalid_phrase()
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # INSERT (create): always show confirmation card before writing to DB.
            # The planner may have correctly parsed a Hinglish/typo job entry — but the user
            # must confirm before any row is inserted.  Route through _show_smart_capture_confirmation
            # which stores a form_state and waits for "Yes"/"Edit".
            if sanitized_sql.upper().lstrip().startswith("INSERT"):
                _plan_values = plan_result.get("plan", {})
                if isinstance(_plan_values, dict):
                    _plan_values = _plan_values.get("values") or {}
                # Parse field values out of the INSERT SQL if the plan didn't surface them
                if not _plan_values:
                    _col_m = re.search(r'INSERT INTO[^(]+\(([^)]+)\)\s*VALUES\s*\(([^)]+)\)', sanitized_sql, re.IGNORECASE)
                    if _col_m:
                        _cols = [c.strip() for c in _col_m.group(1).split(",")]
                        _vals_raw = [v.strip().strip("'") for v in _col_m.group(2).split(",")]
                        _plan_values = {c: v for c, v in zip(_cols, _vals_raw) if c != "user_id"}
                logger.info(f"[CREATE_CONFIRM] Routing INSERT to confirmation flow: {_plan_values}")
                return self._show_smart_capture_confirmation(user_id, _plan_values)

            # Disambiguation: if an UPDATE matches multiple rows, show real options before executing.
            # This covers soft-deletes (SET "isDeleted" = TRUE) and any other UPDATE.
            if sanitized_sql.upper().lstrip().startswith("UPDATE"):
                _where_m = re.search(r'WHERE\s+(.+?)(?=\s+RETURNING\b|$)', sanitized_sql, re.IGNORECASE | re.DOTALL)
                if _where_m:
                    _pre_sql = (
                        "SELECT id, client_name, job_date, job_description_details, fees "
                        f"FROM public.job_entries WHERE {_where_m.group(1).strip()}"
                    )
                    _pre = self.supabase.execute_sql(_pre_sql)
                    if _pre.get("ok") and len(_pre.get("rows", [])) > 1:
                        _cands = _pre["rows"]
                        self.memory.update_user_memory(user_id, {
                            "pending_disambiguation": {
                                "sql": sanitized_sql,
                                "rows": _cands,
                                "data_user_id": data_user_id,
                            }
                        })
                        _opts = [f"I found {len(_cands)} matching records. Which one did you mean?\n"]
                        for _i, _r in enumerate(_cands[:10], 1):
                            _parts = [f"{_i}."]
                            _c = (_r.get("client_name") or "").strip()
                            _d = str(_r.get("job_date") or "")[:10]
                            _desc = str(_r.get("job_description_details") or "")[:40].strip()
                            _f = _r.get("fees")
                            if _c: _parts.append(_c)
                            if _d: _parts.append(_d)
                            if _desc: _parts.append(_desc)
                            if isinstance(_f, (int, float)): _parts.append(f"₹{int(_f):,}")
                            _opts.append(" | ".join(_parts))
                        _opts.append("\nReply with a number to pick, or 'cancel' to abort.")
                        response = "\n".join(_opts)
                        self._store_conversation(user_id, message, response)
                        return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # For UPDATE queries: snapshot old values before overwriting so we can
            # append change history to the notes column.
            _pre_update_rows = {}  # {row_id: {field: old_value, ...}}
            if sanitized_sql.upper().lstrip().startswith("UPDATE"):
                _set_match = re.search(r'\bSET\b\s+(.+?)\s+\bWHERE\b', sanitized_sql, re.IGNORECASE | re.DOTALL)
                _where_match = re.search(r'\bWHERE\b\s+(.+?)(?:\bRETURNING\b|$)', sanitized_sql, re.IGNORECASE | re.DOTALL)
                if _set_match and _where_match:
                    _set_clause = _set_match.group(1)
                    _fields_being_set = re.findall(r'"?(\w+)"?\s*=', _set_clause)
                    _fields_being_set = [f for f in _fields_being_set if f not in ("notes",)]
                    if _fields_being_set:
                        _cols = ", ".join(f'"{f}"' for f in _fields_being_set) + ", id, notes"
                        _pre_sql = f"SELECT {_cols} FROM public.job_entries WHERE {_where_match.group(1).strip()}"
                        _pre_res = self.supabase.execute_sql(_pre_sql)
                        if _pre_res.get("ok"):
                            for _pr in (_pre_res.get("rows") or []):
                                _pre_update_rows[str(_pr.get("id"))] = _pr

            exec_result = self.supabase.execute_sql(sanitized_sql)
            if not exec_result.get("ok"):
                logger.error(f"[QUERY_FAIL] SQL execution failed for user {user_id}: {exec_result.get('error')} | SQL: {sanitized_sql[:200]}")
                response = format_response(
                    ERROR_MODE,
                    error_detail=error_calm_phrase(),
                )
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

            rows = exec_result.get("rows", [])
            op = exec_result.get("operation", "select")
            logger.info(f"[QUERY] op={op}, rows={len(rows)}, user={user_id}, msg='{message[:60]}'")

            if op == "update":
                # Append change history to notes for each updated row
                if _pre_update_rows and rows:
                    from datetime import date as _date
                    _today = _date.today().strftime('%d %b %Y')
                    for _updated_row in rows:
                        _rid = str(_updated_row.get("id"))
                        _old_snap = _pre_update_rows.get(_rid, {})
                        if not _old_snap:
                            continue
                        _history_parts = []
                        for _f, _old_v in _old_snap.items():
                            if _f in ("id", "notes"):
                                continue
                            _new_v = _updated_row.get(_f)
                            if str(_old_v) != str(_new_v):
                                _lbl = _f.replace("_", " ")
                                _old_d = f"₹{int(float(_old_v)):,}" if _f == "fees" and _old_v is not None else str(_old_v) if _old_v is not None else "—"
                                _new_d = f"₹{int(float(_new_v)):,}" if _f == "fees" and _new_v is not None else str(_new_v) if _new_v is not None else "—"
                                _history_parts.append(f"{_lbl}: {_old_d} → {_new_d}")
                        if _history_parts:
                            _entry = f"[{_today}] " + "; ".join(_history_parts)
                            _existing = (_old_snap.get("notes") or "").strip()
                            _new_notes = (_existing + "\n" + _entry).strip() if _existing else _entry
                            _safe_rid = _rid.replace("'", "''")
                            _safe_notes = _new_notes.replace("'", "''")
                            self.supabase.execute_sql(
                                f"UPDATE public.job_entries SET notes = '{_safe_notes}' WHERE id = '{_safe_rid}'"
                            )

                rowcount = exec_result.get("rowcount", 0)
                if rows:
                    self._update_sql_context(user_id, rows)
                    payload = build_clean_payload(rows, "select")
                    response = self.gemini.synthesize_response(payload, message, conversation_history=conversation_history)
                    if not response or not response.strip():
                        response = f"Done! Updated {rowcount} record{'s' if rowcount != 1 else ''}."
                else:
                    # 0 rows updated — the planner may have injected a stale date filter from
                    # memory context even though the user's message contained no explicit date.
                    # Retry by stripping the hallucinated job_date equality filter.
                    _msg_has_date = bool(re.search(
                        r'\b\d{4}-\d{2}-\d{2}\b'
                        r'|\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?'
                        r'|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b',
                        message.lower()
                    ))
                    _retry_sql = re.sub(
                        r"\s+AND\s+job_date\s*=\s*'[^']*'", "", sanitized_sql, flags=re.IGNORECASE
                    )
                    if not _msg_has_date and _retry_sql != sanitized_sql:
                        logger.info(f"[UPDATE_RETRY] Stripping stale date filter (0 rows). Retry SQL: {_retry_sql[:200]}")
                        _where_m2 = re.search(r'WHERE\s+(.+?)(?=\s+RETURNING\b|$)', _retry_sql, re.IGNORECASE | re.DOTALL)
                        if _where_m2:
                            _pre2_sql = (
                                "SELECT id, client_name, job_date, job_description_details, fees "
                                f"FROM public.job_entries WHERE {_where_m2.group(1).strip()}"
                            )
                            _pre2 = self.supabase.execute_sql(_pre2_sql)
                            if _pre2.get("ok"):
                                _pre2_rows = _pre2.get("rows", [])
                                if len(_pre2_rows) > 1:
                                    # Multiple candidates — ask the user to pick
                                    self.memory.update_user_memory(user_id, {
                                        "pending_disambiguation": {
                                            "sql": _retry_sql,
                                            "rows": _pre2_rows,
                                            "data_user_id": data_user_id,
                                        }
                                    })
                                    _opts = [f"I found {len(_pre2_rows)} matching records. Which one did you mean?\n"]
                                    for _i, _r in enumerate(_pre2_rows[:10], 1):
                                        _parts = [f"{_i}."]
                                        _c = (_r.get("client_name") or "").strip()
                                        _d = str(_r.get("job_date") or "")[:10]
                                        _desc = str(_r.get("job_description_details") or "")[:40].strip()
                                        _f = _r.get("fees")
                                        if _c: _parts.append(_c)
                                        if _d: _parts.append(_d)
                                        if _desc: _parts.append(_desc)
                                        if isinstance(_f, (int, float)): _parts.append(f"₹{int(_f):,}")
                                        _opts.append(" | ".join(_parts))
                                    _opts.append("\nReply with a number to pick, or 'cancel' to abort.")
                                    response = "\n".join(_opts)
                                    self._store_conversation(user_id, message, response)
                                    return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
                                elif len(_pre2_rows) == 1:
                                    # Exactly one match — execute the retry UPDATE
                                    _retry_exec = self.supabase.execute_sql(_retry_sql)
                                    if _retry_exec.get("ok") and _retry_exec.get("rowcount", 0) > 0:
                                        _retry_rows = _retry_exec.get("rows", [])
                                        if _retry_rows:
                                            self._update_sql_context(user_id, _retry_rows)
                                        payload = build_clean_payload(_retry_rows or _pre2_rows, "select")
                                        response = self.gemini.synthesize_response(payload, message, conversation_history=conversation_history)
                                        if not response or not response.strip():
                                            response = f"Done! Updated {_retry_exec.get('rowcount', 1)} record."
                                        self._store_conversation(user_id, message, response)
                                        return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
                    # Genuine no-match after retry
                    logger.warning(f"[UPDATE_FAIL] 0 rows updated after retry. SQL: {sanitized_sql[:200]}")
                    response = "I couldn't find a matching record to update. Could you be more specific (e.g. include the date or job description)?"
            elif op == "insert":
                if rows:
                    self._update_sql_context(user_id, rows)
                    response = format_response(ASSISTANT_MODE, insert_confirmation=True)
                else:
                    response = format_response(ASSISTANT_MODE, insert_confirmation=True)
                # Check for compound intent — suggest next action after job save
                insert_mem = self.memory.get_user_memory(user_id)
                suggested_next = insert_mem.get("suggested_next_action")
                if suggested_next:
                    self.memory.update_user_memory(user_id, {"awaiting_compound_response": True})
                    response += f"\n\nYou also mentioned: \"{suggested_next}\"\nWant me to do that now? (Yes / No)"
                    logger.info(f"[COMPOUND] Suggesting next action after insert: '{suggested_next}'")
            else:
                if not rows:
                    # Check if user has ANY data at all
                    count_result = self.supabase.execute_sql(
                        f"SELECT COUNT(*) AS cnt FROM public.job_entries WHERE user_id = '{data_user_id.replace(chr(39), chr(39)+chr(39))}' AND (\"isDeleted\" IS NOT TRUE)"
                    )
                    has_data = False
                    if count_result.get("ok") and count_result.get("rows"):
                        has_data = int(count_result["rows"][0].get("cnt", 0)) > 0
                    if not has_data:
                        logger.info(f"[QUERY_FAIL] User {user_id} has NO data at all (0 rows in job_entries)")
                        user_name = self._get_user_name(user_id)
                        greeting = f"{user_name}, you" if user_name else "You"
                        response = (
                            f"{greeting} don't have any jobs logged yet.\n\n"
                            "To get started, say something like:\n"
                            "• 'Add a job for [Client Name]'\n"
                            "• '+ClientName, job details, 5000'\n\n"
                            "Once you have jobs, I can answer queries, send reminders, and generate invoices!"
                        )
                    else:
                        logger.warning(f"[QUERY_FAIL] 0 rows but user HAS data. SQL returned nothing for: {sanitized_sql[:200]}")
                        # Try deterministic keyword fallback before giving up — the planner may have
                        # injected stale context filters (e.g. a remembered client/date) that
                        # don't match the user's actual intent (e.g. "what was my last job").
                        kw_sql = self._keyword_sql_fallback(message, data_user_id)
                        if kw_sql:
                            logger.info(f"[PIPELINE] Keyword retry after 0-row planner result: {kw_sql[:200]}")
                            kw_exec = self.supabase.execute_sql(kw_sql)
                            kw_rows = kw_exec.get("rows", []) if kw_exec.get("ok") else []
                            if kw_rows:
                                self._update_sql_context(user_id, kw_rows)
                                ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
                                ctx["last_sql"] = kw_sql
                                self.memory.update_user_memory(user_id, {"uscf_context": ctx})
                                if len(kw_rows) > 4:
                                    excel_path = _generate_jobs_excel(kw_rows, data_user_id)
                                    response = f"Found {len(kw_rows)} results — here's a spreadsheet with all of them."
                                    self._store_conversation(user_id, message, response)
                                    return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}, "excel_path": excel_path}
                                if _is_full_job_row(kw_rows[0]) and not _is_history_q:
                                    response = _format_job_cards(kw_rows)
                                else:
                                    payload = build_clean_payload(kw_rows, "select")
                                    response = self.gemini.synthesize_response(payload, message, history_question=_is_history_q, conversation_history=conversation_history)
                                    if not response or not response.strip():
                                        response = self._format_sql_result(kw_rows)
                                self._store_conversation(user_id, message, response)
                                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
                        # AI-driven helpful suggestion when SQL returned 0 rows but user has data.
                        try:
                            _cols = self._get_user_columns(data_user_id) if hasattr(self, "_get_user_columns") else []
                        except Exception:
                            _cols = []
                        _suggest = self.gemini.suggest_for_empty_result(
                            message, recent_columns=_cols, applied_sql=sanitized_sql or "",
                        )
                        response = _suggest or "Nothing matched that exact filter. Try a related query — for example, ask 'list my clients' or 'show unpaid jobs'."
                    self._store_conversation(user_id, message, response)
                    return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
                self._update_sql_context(user_id, rows)
                # Store last SQL so "Yes, show details" follow-ups can re-run it as SELECT *
                ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
                ctx["last_sql"] = sanitized_sql
                self.memory.update_user_memory(user_id, {"uscf_context": ctx})
                if len(rows) > 4:
                    excel_path = _generate_jobs_excel(rows, data_user_id)
                    response = f"Found {len(rows)} results — here's a spreadsheet with all of them."
                    logger.info(f"[QUERY] Excel generated: {excel_path} ({len(rows)} rows)")
                    self._store_conversation(user_id, message, response)
                    return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}, "excel_path": excel_path}
                if _is_full_job_row(rows[0]) and not _is_history_q:
                    response = _format_job_cards(rows)
                    logger.info(f"[QUERY] Success: {len(rows)} rows (structured card format)")
                else:
                    payload = build_clean_payload(rows, "select")
                    response = self.gemini.synthesize_response(payload, message, history_question=_is_history_q, conversation_history=conversation_history)
                    if not response or not response.strip():
                        logger.warning(f"[QUERY_FAIL] synthesize_response returned empty for {len(rows)} rows, msg='{message[:60]}'")
                        # Deterministic fallback for simple aggregate results — never show "couldn't format"
                        response = _format_aggregate_fallback(payload, message)
                    else:
                        logger.info(f"[QUERY] Success: {len(rows)} rows, response length={len(response)}")

        except Exception as e:
            logger.error(f"[QUERY_FAIL] Exception for user {user_id}, msg='{message[:60]}': {e}", exc_info=True)
            user_name = self._get_user_name(user_id)
            if user_name:
                response = format_response(ERROR_MODE, error_detail=f"Sorry {user_name}, {error_calm_phrase().lower()}")
            else:
                response = format_response(ERROR_MODE, error_detail=error_calm_phrase())

        self._store_conversation(user_id, message, response)
        return {
            "operation": "query",
            "response": response,
            "trigger_invoice": trigger_invoice,
            "invoice_data": invoice_data
        }

    def _execute_routed_query(self, routed, user_id: str, data_user_id: str,
                              message: str, conversation_history: list) -> Optional[Dict]:
        """Execute a deterministic RoutedQuery and render its rows.

        Returns a response dict on success, or None to fall through to the LLM
        planner (router matched but the DB returned nothing usable). All four
        render kinds share the same context-update + storage tail so follow-ups
        ("yes, show details") keep working.
        """
        logger.info(f"[ROUTER] matched route '{routed.name}' ({routed.render}): {routed.sql[:200]}")
        exec_result = self.supabase.execute_sql(routed.sql)
        if not exec_result.get("ok"):
            logger.warning(f"[ROUTER] SQL failed for route '{routed.name}': {exec_result.get('error')}")
            return None
        rows = exec_result.get("rows", []) or []

        # Remember the SQL + rows so context-dependent follow-ups still work.
        if rows:
            self._update_sql_context(user_id, rows)
            ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
            ctx["last_sql"] = routed.sql
            self.memory.update_user_memory(user_id, {"uscf_context": ctx})

        def _finish(resp: str) -> Dict:
            self._store_conversation(user_id, message, resp)
            return {"operation": "query", "response": resp, "trigger_invoice": False, "invoice_data": {}}

        # ── Deterministic renders (no LLM) ──
        if routed.render == _RENDER_CLIENT_LIST:
            return _finish(format_client_list(rows, routed.meta.get("status", "all")))

        if routed.render == _RENDER_PAYMENT_STATUS:
            if not rows:
                # Client not found — let the planner try (typo / different phrasing).
                return None
            return _finish(format_payment_status(rows, routed.meta))

        if routed.render == _RENDER_AGGREGATE:
            payload = build_clean_payload(rows, "select")
            resp = self.gemini.synthesize_response(payload, message, conversation_history=conversation_history)
            if not resp or not resp.strip():
                resp = _format_aggregate_fallback(payload, message)
            return _finish(resp)

        # ── ROWS: full job rows → cards / spreadsheet / synthesiser ──
        if not rows:
            return None  # nothing to show — hand to planner for a helpful empty-state reply
        if len(rows) > 4 and _is_full_job_row(rows[0]):
            excel_path = _generate_jobs_excel(rows, data_user_id)
            resp = f"Found {len(rows)} results — here's a spreadsheet with all of them."
            self._store_conversation(user_id, message, resp)
            return {"operation": "query", "response": resp, "trigger_invoice": False, "invoice_data": {}, "excel_path": excel_path}
        if _is_full_job_row(rows[0]):
            return _finish(_format_job_cards(rows))
        payload = build_clean_payload(rows, "select")
        resp = self.gemini.synthesize_response(payload, message, conversation_history=conversation_history)
        if not resp or not resp.strip():
            resp = _format_aggregate_fallback(payload, message)
        return _finish(resp)

    def _keyword_sql_fallback(self, message: str, user_id: str) -> Optional[str]:
        """
        Deterministic SQL fallback based on keyword matching — no LLM needed.
        Returns a SQL string for common query patterns, or None if no pattern matches.

        NOTE: the primary path is now services/query_router.route_common_query,
        run BEFORE the planner. This method remains as the planner-FAILURE safety
        net (planner errored or returned 0 rows) and delegates to the same router
        so there is a single source of truth for common-query SQL.
        """
        # Single source of truth: try the deterministic router first.
        _routed = route_common_query(message, user_id)
        if _routed is not None:
            return _routed.sql

        msg = message.strip().lower()
        uid = user_id.replace("'", "''")
        _not_deleted = "(\"isDeleted\" IS NOT TRUE)"
        base = f"SELECT * FROM public.job_entries WHERE user_id = '{uid}' AND {_not_deleted}"
        _client_expr = (
            "COALESCE(NULLIF(client_name,''),NULLIF(brand_name,''),NULLIF(production_house,''))"
        )

        # "biggest/top/largest client" — group by client, sum fees, top 1
        if re.search(r'\b(biggest|top|largest|best|highest[- ]paying)\b.{0,30}\b(client|brand|company)\b', msg):
            return (
                f"SELECT {_client_expr} AS client_name, SUM(fees) AS result "
                f"FROM public.job_entries WHERE user_id='{uid}' AND {_not_deleted} "
                f"GROUP BY 1 HAVING {_client_expr} IS NOT NULL ORDER BY result DESC LIMIT 1"
            )

        # "highest paying job" / "most expensive job" — single row, max fees
        if re.search(r'\b(highest[- ]paying|most expensive|biggest|top[- ]earning)\b.{0,20}\b(job|project|work|gig)\b', msg) \
                or re.search(r'\b(highest|most|max(imum)?)\b.{0,20}\b(pay(ing)?|fee|earning|income)\b', msg):
            return (
                f"SELECT * FROM public.job_entries WHERE user_id='{uid}' AND {_not_deleted} "
                f"AND fees IS NOT NULL ORDER BY fees DESC NULLS LAST LIMIT 1"
            )

        # "earnings by client" / "fees per client" / "breakdown by client" — all clients grouped
        if re.search(r'\b(earnings?|fees?|billing|revenue|income)\b.{0,20}\b(by|per|for each|breakdown)\b.{0,20}\b(client|brand|company)\b', msg) \
                or re.search(r'\b(by|per|for each)\b.{0,20}\b(client|brand)\b.{0,20}\b(earnings?|fees?|billing)\b', msg) \
                or re.search(r'\b(show|list).{0,20}\b(earnings?|income|revenue).{0,20}\b(client|brand)\b', msg):
            return (
                f"SELECT {_client_expr} AS client_name, SUM(fees) AS result "
                f"FROM public.job_entries WHERE user_id='{uid}' AND {_not_deleted} "
                f"GROUP BY 1 HAVING {_client_expr} IS NOT NULL ORDER BY result DESC"
            )

        # "average fees per job" / "average billing" / "औसत"
        if re.search(r'\b(average|avg|औसत)\b.{0,30}\b(fees?|billing|earnings?|amount|income)\b', msg):
            return (
                f"SELECT AVG(fees) AS result FROM public.job_entries "
                f"WHERE user_id='{uid}' AND {_not_deleted} AND fees IS NOT NULL"
            )

        # "how much does X owe me" / "X ka paisa" — client + unpaid SUM
        _owe_m = re.search(
            r'\b(how\s+much\s+does\s+(.+?)\s+owe\s+me'
            r'|(.+?)\s+(?:ka\s+paisa|se\s+paisa\s+aaya|ka\s+payment)\b)',
            msg,
        )
        if _owe_m:
            _client_raw = (_owe_m.group(2) or _owe_m.group(3) or "").strip().strip("?").strip()
            if _client_raw and len(_client_raw) > 1:
                _c = _client_raw.replace("'", "''")
                return (
                    f"SELECT SUM(fees) AS result FROM public.job_entries "
                    f"WHERE user_id='{uid}' AND {_not_deleted} "
                    f"AND ({_client_expr} ILIKE '%{_c}%') "
                    f"AND (paid IS NULL OR TRIM(COALESCE(paid,''))='' OR LOWER(paid) NOT IN ('true','t','yes','1','paid'))"
                )

        # "last job" / "latest job" / "most recent job" / "recent job"
        if re.search(r'\b(last|latest|most\s+recent|recent)\b.*\b(jobs?|entr(?:y|ies)?|work|project|gig)\b', msg):
            return f"{base} ORDER BY job_date DESC NULLS LAST LIMIT 1"

        # "how many jobs" / "count" / "total jobs"
        if re.search(r'\b(how\s+many|count|total\s+number\s+of|number\s+of)\b.*\b(jobs?|entr(?:y|ies)?|records?|work)\b', msg):
            return f"SELECT COUNT(*) AS result FROM public.job_entries WHERE user_id = '{uid}' AND (\"isDeleted\" IS NOT TRUE)"

        # "total fees" / "total earnings" / "sum of fees"
        if re.search(r'\b(total|sum|overall)\b.*\b(fees?|earnings?|income|revenue|billing)\b', msg):
            return f"SELECT SUM(fees) AS result FROM public.job_entries WHERE user_id = '{uid}' AND (\"isDeleted\" IS NOT TRUE)"

        # "show all my clients" / "list clients" / "which clients" (NOT how many)
        if re.search(r'\b(show|list|all|which|my)\b.{0,20}\b(clients?|brands?|companies|compan(?:y|ies))\b', msg) \
                and not re.search(r'\b(how\s+many|count|kitne|number\s+of)\b', msg):
            return (
                f"SELECT DISTINCT {_client_expr} AS client_name "
                f"FROM public.job_entries WHERE user_id='{uid}' AND {_not_deleted} "
                f"AND {_client_expr} IS NOT NULL ORDER BY 1"
            )

        # "show all jobs" / "list jobs" / "my jobs"
        if re.search(r'\b(show|list|all|my)\b.*\b(jobs?|entr(?:y|ies)?|records?|work)\b', msg):
            return f"{base} ORDER BY job_date DESC NULLS LAST LIMIT 25"

        # Hinglish/Roman Hindi: "pichhle quarter/mahine mein kitna paisa aaya" — earnings SUM (no client filter)
        # "paisa aaya" alone (without "se" / "ka" indicating a client) → total SUM
        if re.search(r'\b(paisa\s+aaya|paisa\s+mila|kamai|kamaya|kitna\s+mila)\b', msg) \
                and not re.search(r'\b(se\s+paisa\s+aaya|ka\s+paisa|ka\s+payment|se\s+paisa\s+mila)\b', msg):
            return f"SELECT SUM(fees) AS result FROM public.job_entries WHERE user_id='{uid}' AND {_not_deleted}"

        # "what did I do on [date]" / "show jobs on [date]" — date-specific job lookup
        _date_on_m = re.search(
            r'\b(?:what\s+did\s+i\s+do|what\s+was|show\s+(?:me\s+)?jobs?)\s+on\s+'
            r'(\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*(?:\s+\d{4})?'
            r'|(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{1,2}(?:\s*,?\s*\d{4})?)',
            msg,
        )
        if _date_on_m:
            _raw_date = _date_on_m.group(1).strip()
            from datetime import datetime as _dt_
            _cur_year = _dt_.now().year
            for _fmt in ("%d %B", "%d %b", "%B %d", "%b %d"):
                try:
                    _parsed = _dt_.strptime(_raw_date, _fmt).replace(year=_cur_year)
                    _d = _parsed.strftime("%Y-%m-%d")
                    return f"{base} AND job_date = '{_d}' ORDER BY job_date DESC"
                except ValueError:
                    pass
            for _fmt in ("%d %B %Y", "%d %b %Y", "%B %d %Y", "%b %d, %Y"):
                try:
                    _d = _dt_.strptime(_raw_date, _fmt).strftime("%Y-%m-%d")
                    return f"{base} AND job_date = '{_d}' ORDER BY job_date DESC"
                except ValueError:
                    pass

        # "unpaid" / "pending payments"
        if re.search(r'\b(unpaid|pending|not\s+paid|outstanding)\b', msg):
            return f"{base} AND (paid IS NULL OR paid = '' OR paid = 'false' OR LOWER(paid) != 'true') ORDER BY job_date DESC NULLS LAST LIMIT 25"

        # Generic fallback for any question with "job" or "jobs"
        if re.search(r'\bjobs?\b', msg):
            return f"{base} ORDER BY job_date DESC NULLS LAST LIMIT 5"

        return None

    def _handle_soft_delete(self, user_id: str, message: str, data_user_id: str, conversation_history: list) -> Dict:
        """
        Soft-delete a job entry by setting "isDeleted" = true.
        For 'delete my last job': finds the most recent job and deletes it.
        For 'delete this job': uses the last shown row from context.
        Shows what was deleted, or asks to disambiguate if multiple rows match.
        """
        msg_lower = message.strip().lower()
        uid = data_user_id.replace("'", "''")
        _not_deleted = "(\"isDeleted\" IS NOT TRUE)"

        # Determine which job to delete
        # Case 1: "last job" / "latest job" → most recent by job_date
        is_last = re.search(r'\b(last|latest|most\s+recent|recent)\b', msg_lower)
        # Case 2: "this job" / "it" / "that" → use last shown row from context
        ctx = self.memory.get_user_memory(user_id).get("uscf_context", {})
        last_row = ctx.get("last_row_data")

        # Detect "all" — bulk delete mode (e.g. "delete all Nike jobs", "remove all entries")
        is_bulk = bool(re.search(r'\ball\b', msg_lower))
        # Try to pull a client/brand name out of the message:
        #   "delete all Nike jobs"   → "Nike"
        #   "delete Nike jobs"       → "Nike"
        #   "remove all from Bisleri"→ "Bisleri"
        client_hint = ""
        _patterns = [
            r'\b(?:delete|remove|erase|trash|discard)\s+(?:all\s+)?(?:my\s+)?(.+?)\s+(?:jobs?|entries|records?|rows?)\b',
            r'\b(?:delete|remove|erase|trash|discard)\s+(?:all\s+)?(?:jobs?|entries|records?|rows?)\s+(?:for|from|of)\s+(.+?)$',
            r'\b(?:all|every)\s+(.+?)\s+(?:jobs?|entries|records?)\b',
        ]
        # "last"/"latest"/"recent" are POSITIONAL references, not client names —
        # "delete my last job" must not extract "last" as a client (which then
        # matches nothing → "no jobs matching 'last'"). Skip hint extraction when
        # the message is a last/recent reference; is_last drives row selection below.
        _hint_stopwords = ("my", "the", "all", "every", "this", "that",
                           "last", "latest", "recent", "most", "most recent")
        if not is_last:
            for _pat in _patterns:
                _m = re.search(_pat, msg_lower, re.IGNORECASE)
                if _m:
                    _hint = _m.group(1).strip().strip("'\"")
                    # Reject empty / generic / positional words
                    if _hint and _hint not in _hint_stopwords:
                        client_hint = _hint
                        break

        if not is_last and last_row and any(w in msg_lower for w in ["this", "it", "that"]):
            # Delete the job that was most recently shown
            row_id = last_row.get("id", "")
            if not row_id:
                response = "I'm not sure which job you want to delete. Try 'delete my last job'."
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
            candidate_rows = [last_row]
        else:
            # Build the candidate fetch — filter by client/brand if we extracted one,
            # and drop the LIMIT for bulk deletes so we don't silently truncate.
            _where_extra = ""
            _params_label = ""
            if client_hint:
                _safe_hint = client_hint.replace("'", "''")
                _where_extra = (
                    f" AND (client_name ILIKE '%{_safe_hint}%' "
                    f"OR brand_name ILIKE '%{_safe_hint}%' "
                    f"OR production_house ILIKE '%{_safe_hint}%')"
                )
                _params_label = f" matching '{client_hint}'"
            _limit_clause = "" if is_bulk else "LIMIT 5"
            fetch_sql = (
                f"SELECT id, client_name, brand_name, job_date, job_description_details, fees "
                f"FROM public.job_entries "
                f"WHERE user_id = '{uid}' AND {_not_deleted}{_where_extra} "
                f"ORDER BY job_date DESC NULLS LAST {_limit_clause}"
            ).strip()
            fetch_result = self.supabase.execute_sql(fetch_sql)
            if not fetch_result.get("ok") or not fetch_result.get("rows"):
                response = f"You don't have any jobs{_params_label} to delete."
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
            candidate_rows = fetch_result["rows"]

        # If "last job" and multiple rows returned, take only the very last one
        if is_last:
            candidate_rows = candidate_rows[:1]

        # Bulk path: user said "all" — confirm before nuking everything we found.
        if is_bulk and len(candidate_rows) > 1:
            self.memory.update_user_memory(user_id, {
                "pending_disambiguation": {
                    "rows": candidate_rows,
                    "data_user_id": data_user_id,
                    "bulk_mode": True,
                    "client_hint": client_hint,
                }
            })
            _label = f"{len(candidate_rows)} jobs"
            if client_hint:
                _label = f"{len(candidate_rows)} {client_hint} jobs"
            preview = []
            for r in candidate_rows[:5]:
                d = str(r.get("job_date") or "")[:10]
                desc = str(r.get("job_description_details") or "")[:40].strip()
                bits = []
                if d: bits.append(d)
                if desc: bits.append(f"— {desc}")
                preview.append("• " + (" ".join(bits) if bits else "(no date)"))
            more = "" if len(candidate_rows) <= 5 else f"\n…and {len(candidate_rows) - 5} more"
            response = (
                f"I found {_label}. Reply 'Yes' to delete all of them, "
                f"or 'cancel' to abort.\n\n" + "\n".join(preview) + more
            )
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        if len(candidate_rows) > 1:
            # Disambiguate — reuse the existing disambiguation flow
            update_sql = (
                f"UPDATE public.job_entries SET \"isDeleted\" = true "
                f"WHERE user_id = '{uid}' AND {_not_deleted} RETURNING *"
            )
            self.memory.update_user_memory(user_id, {
                "pending_disambiguation": {
                    "sql": update_sql,
                    "rows": candidate_rows,
                    "data_user_id": data_user_id,
                }
            })
            opts = [f"I found {len(candidate_rows)} jobs. Which one do you want to delete?\n"]
            for i, r in enumerate(candidate_rows[:10], 1):
                parts = [f"{i}."]
                c = (r.get("client_name") or r.get("brand_name") or "").strip()
                d = str(r.get("job_date") or "")[:10]
                desc = str(r.get("job_description_details") or "")[:40].strip()
                if c: parts.append(c)
                if d: parts.append(d)
                if desc: parts.append(f"— {desc}")
                opts.append(" ".join(parts))
            opts.append("\nReply with a number, 'all' to delete every match, or 'cancel' to abort.")
            response = "\n".join(opts)
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Single row — soft-delete it
        target = candidate_rows[0]
        row_id = target.get("id", "")
        if not row_id:
            response = "Couldn't find that job. Please try again."
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        delete_sql = (
            f"UPDATE public.job_entries SET \"isDeleted\" = true "
            f"WHERE id = '{row_id}' AND user_id = '{uid}' RETURNING id"
        )
        del_result = self.supabase.execute_sql(delete_sql)
        if not del_result.get("ok"):
            response = "Something went wrong deleting that job. Please try again."
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # Build a human-readable description of what was deleted
        client = (target.get("client_name") or target.get("brand_name") or "").strip()
        date_str = str(target.get("job_date") or "")[:10]
        desc = str(target.get("job_description_details") or "")[:60].strip()
        detail_parts = []
        if client: detail_parts.append(client)
        if date_str: detail_parts.append(date_str)
        if desc: detail_parts.append(f"— {desc}")
        detail = " | ".join(detail_parts) if detail_parts else "the job"

        response = f"Done — deleted: {detail}.\n\nLet me know if you need anything else."
        self._store_conversation(user_id, message, response)
        return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_disambiguation_reply(self, user_id: str, message: str, pending: Dict) -> Dict:
        """User is replying to pick one row (by number) or all rows ('all'/'yes') from a disambiguation list."""
        msg = message.strip().lower().rstrip(".!?")
        if msg in ("cancel", "stop", "nevermind", "abort", "no"):
            self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
            response = "Cancelled. Let me know if you need anything else."
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        rows = pending.get("rows", [])
        _is_bulk_confirm = bool(pending.get("bulk_mode"))
        # "yes"/"y"/"do it" only mean "delete all" when we EXPLICITLY asked a bulk
        # confirmation ("Reply 'Yes' to delete all of them"). In a numbered
        # disambiguation we only offered number/all/cancel — a bare "yes" there is
        # ambiguous (often a stray confirmation meant for a DIFFERENT pending flow,
        # e.g. an invoice email prompt). Treating it as delete-all caused jobs to be
        # silently deleted. So: affirmatives delete-all ONLY in bulk_mode.
        _AFFIRM_TOKENS = {"yes", "y", "do it", "every one", "everyone"}
        _EXPLICIT_ALL = {"all", "delete all", "all of them"}
        _is_explicit_all = msg in _EXPLICIT_ALL or msg.startswith("delete all")
        _is_affirm = msg in _AFFIRM_TOKENS or msg.startswith("yes")
        if _is_affirm and not _is_bulk_confirm:
            # Ambiguous bare "yes" in a numbered disambiguation — do NOT delete.
            # Fall through so a competing pending state (invoice email send
            # confirmation, compound follow-up, etc.) can handle it instead.
            self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
            logger.info("[DISAMBIG] Bare 'yes' in numbered disambiguation — clearing and falling through (NOT deleting)")
            return None
        if _is_explicit_all or (_is_affirm and _is_bulk_confirm):
            if not rows:
                self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
                response = "Nothing to delete. Let me know if you need anything else."
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
            data_uid = pending.get("data_user_id", user_id).replace("'", "''")
            ids = [r.get("id") for r in rows if r.get("id")]
            if not ids:
                self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
                response = "Couldn't resolve which rows to delete. Please try again."
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
            ids_sql = ",".join(f"'{rid}'" for rid in ids)
            bulk_sql = (
                f"UPDATE public.job_entries SET \"isDeleted\" = true "
                f"WHERE id IN ({ids_sql}) AND user_id = '{data_uid}' RETURNING id"
            )
            exec_result = self.supabase.execute_sql(bulk_sql)
            self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
            if not exec_result.get("ok"):
                response = "Something went wrong with the bulk delete. Please try again."
                self._store_conversation(user_id, message, response)
                return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}
            deleted_count = len(exec_result.get("rows", []) or ids)
            hint = pending.get("client_hint") or ""
            label = f"{deleted_count} {hint} job{'s' if deleted_count != 1 else ''}".strip()
            response = f"Done — deleted {label}. Let me know if you need anything else."
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        num_match = re.search(r'\b(\d+)\b', message)
        if not num_match:
            # If the message is clearly a new query (ends with ?, starts with a question
            # word) — clear disambiguation and fall through so the user isn't trapped.
            _starts_with_query_word = any(
                msg.startswith(w + " ") or msg == w
                for w in ("who", "what", "show", "list", "how", "when", "which", "get", "find",
                          "kiska", "kitne", "kitna", "kaunsa", "kya")
            )
            _new_query_signals = "?" in message or _starts_with_query_word
            if _new_query_signals:
                self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
                return None  # fall through to normal pipeline
            response = "Please reply with a number to select the record, 'all' to delete every match, or 'cancel' to abort."
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        idx = int(num_match.group(1)) - 1
        if idx < 0 or idx >= len(rows):
            response = f"Please choose a number between 1 and {len(rows)}, or 'cancel' to abort."
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        chosen_id = rows[idx].get("id", "")
        data_uid = pending.get("data_user_id", user_id).replace("'", "''")
        original_sql = pending.get("sql", "")

        # Replace the WHERE clause with a precise id + user_id lookup
        targeted_sql = re.sub(
            r'WHERE\s+.+?(?=\s+RETURNING\b|$)',
            f"WHERE id = '{chosen_id}' AND user_id = '{data_uid}'",
            original_sql,
            flags=re.IGNORECASE | re.DOTALL,
        ).strip()

        self.memory.update_user_memory(user_id, {"pending_disambiguation": None})
        exec_result = self.supabase.execute_sql(targeted_sql)
        if not exec_result.get("ok"):
            response = "Something went wrong. Please try again."
            self._store_conversation(user_id, message, response)
            return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

        result_rows = exec_result.get("rows", [])
        _is_update = targeted_sql.strip().upper().startswith("UPDATE")
        if result_rows and _is_update:
            # Deterministic confirmation for the picked UPDATE — don't hand the row
            # to the synthesiser, which sometimes just re-describes it (and can read
            # the new state as the old, e.g. "it's currently unpaid" right after we
            # marked it paid). State plainly what changed.
            self._update_sql_context(user_id, result_rows)
            _row = result_rows[0]
            _client = (_row.get("client_name") or _row.get("brand_name") or "the job").strip()
            _bill = (_row.get("bill_no") or "").strip()
            _label = f"{_client} ({_bill})" if _bill else _client
            _updates = pending.get("updates") or {}
            _paid_set = any(str(k).lower() == "paid" for k in _updates) or "set paid" in targeted_sql.lower()
            _now_paid = str(_row.get("paid", "")).strip().lower() in ("yes", "true", "t", "1", "paid")
            if _paid_set and _now_paid:
                response = f"✅ Done — marked {_label} as paid."
            elif _updates:
                _changed = ", ".join(f"{k.replace('_', ' ')} → {v}" for k, v in _updates.items())
                response = f"✅ Done — updated {_label}: {_changed}."
            else:
                response = f"✅ Done — updated {_label}."
        elif result_rows:
            self._update_sql_context(user_id, result_rows)
            payload = build_clean_payload(result_rows, "select")
            # _handle_disambiguation_reply has no conversation_history parameter
            # (the caller is the form-step dispatcher, not process_request).
            # Re-fetch from memory so the synth call has context for the reply.
            _hist = self.memory.get_conversation_history(user_id)
            response = self.gemini.synthesize_response(payload, message, conversation_history=_hist)
            if not response or not response.strip():
                response = "Done! The selected record has been updated."
        else:
            response = "Done! The selected record has been updated."

        self._store_conversation(user_id, message, response)
        return {"operation": "query", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _resolve_data_user_id(self, user_id: str, profile_data: Dict) -> str:
        """
        Resolve the effective user_id for data queries.
        If a linked account exists in preferences, use that for job_entries queries.
        """
        prefs = profile_data.get("preferences") or {}
        if isinstance(prefs, str):
            try:
                prefs = json.loads(prefs)
            except (json.JSONDecodeError, TypeError):
                prefs = {}
        linked_id = prefs.get("linked_user_id")
        if linked_id:
            return linked_id
        return user_id

    def _handle_link_account(self, user_id: str, message: str) -> Dict:
        """Handle 'link account' request — extract ID inline or prompt."""
        msg_lower = message.strip().lower()
        # Try to extract ID inline: "link telegram 751256859"
        parts = message.strip().split()
        # Look for a numeric ID or whatsapp:+ ID in the message
        candidate = None
        for part in parts:
            clean = part.strip()
            if clean.isdigit() and len(clean) >= 5:
                candidate = clean
                break
            if clean.startswith("whatsapp:+"):
                candidate = clean
                break

        if candidate:
            return self._apply_link(user_id, message, candidate)

        # No inline ID — prompt for it
        self.memory.update_user_memory(user_id, {"awaiting_link_id": True})
        platform = "telegram" if user_id.isdigit() else "whatsapp"
        other = "WhatsApp" if platform == "telegram" else "Telegram"
        response = (
            f"To link your {other} account, I need the user ID from that platform.\n\n"
            f"You can find it by messaging the bot on {other} and asking 'what is my user id'.\n\n"
            "Please paste the ID here:"
        )
        self._store_conversation(user_id, message, response)
        return {"operation": "link_prompt", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _process_link_id(self, user_id: str, message: str) -> Dict:
        """Process the linked account ID after the user was prompted."""
        self.memory.update_user_memory(user_id, {"awaiting_link_id": False})
        candidate = message.strip()
        if candidate.lower() in ("cancel", "nevermind", "never mind", "no"):
            response = "No worries, account not linked."
            self._store_conversation(user_id, message, response)
            return {"operation": "link_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}
        return self._apply_link(user_id, message, candidate)

    def _apply_link(self, user_id: str, message: str, linked_id: str) -> Dict:
        """Store the linked account ID in user preferences."""
        platform = "telegram" if user_id.isdigit() else "whatsapp"
        # Get current preferences
        profile = self.supabase.get_user_profile(user_id)
        prefs = {}
        if profile.get("ok") and profile.get("data"):
            prefs = profile["data"].get("preferences") or {}
            if isinstance(prefs, str):
                try:
                    prefs = json.loads(prefs)
                except (json.JSONDecodeError, TypeError):
                    prefs = {}

        prefs["linked_user_id"] = linked_id
        result = self.supabase.upsert_user_profile(user_id, platform, {"preferences": json.dumps(prefs)})
        if result.get("ok"):
            response = f"Account linked! ✅ Your data from user ID '{linked_id}' is now accessible here."
            logger.info(f"[LINK] Linked {user_id} → {linked_id}")
        else:
            response = "Sorry, I couldn't link the account right now. Please try again."
            logger.error(f"[LINK] Failed to link {user_id} → {linked_id}: {result.get('error')}")
        self._store_conversation(user_id, message, response)
        return {"operation": "account_linked", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_name_change(self, user_id: str, message: str) -> Dict:
        """Handle 'change my name' request — check if name is inline or prompt."""
        msg_lower = message.strip().lower()
        # Try to extract name inline: "change my name to Akshaj"
        import re as _re
        m = _re.search(r'(?:name\s+to|name\s+as|rename\s+me\s+to?)\s+(.+)', msg_lower)
        if m:
            new_name = m.group(1).strip().title()
            return self._apply_name_change(user_id, message, new_name)

        # No inline name — prompt for it
        self.memory.update_user_memory(user_id, {"awaiting_name_change": True})
        current_name = self._get_user_name(user_id) or "unknown"
        response = f"Your current name is '{current_name}'. What would you like to change it to?"
        self._store_conversation(user_id, message, response)
        return {"operation": "name_change_prompt", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _process_name_change(self, user_id: str, message: str) -> Dict:
        """Process the new name after the user was prompted."""
        self.memory.update_user_memory(user_id, {"awaiting_name_change": False})
        new_name = message.strip()
        if new_name.lower() in ("cancel", "nevermind", "never mind", "no"):
            response = "No worries, name unchanged."
            self._store_conversation(user_id, message, response)
            return {"operation": "name_change_cancelled", "response": response, "trigger_invoice": False, "invoice_data": {}}
        return self._apply_name_change(user_id, message, new_name.title())

    def _apply_name_change(self, user_id: str, message: str, new_name: str) -> Dict:
        """Apply the name change to user_profiles.

        Also syncs preferences.invoice_name so the new name flows through to
        the invoice header (the PDF reads invoice_name first, falling back to
        name — without this sync, a stale invoice_name silently overrides the
        update).
        """
        platform = "telegram" if user_id.isdigit() else "whatsapp"

        # Merge new_name into preferences so invoice header picks it up too
        prefs = {}
        existing = self.supabase.get_user_profile(user_id)
        if existing.get("ok") and existing.get("data"):
            _p = existing["data"].get("preferences") or {}
            if isinstance(_p, str):
                try:
                    prefs = json.loads(_p)
                except (json.JSONDecodeError, TypeError):
                    prefs = {}
            elif isinstance(_p, dict):
                prefs = _p
        prefs["invoice_name"] = new_name

        result = self.supabase.upsert_user_profile(
            user_id, platform,
            {"name": new_name, "preferences": json.dumps(prefs)},
        )
        if result.get("ok"):
            response = f"Done! Your name has been updated to '{new_name}'. ✅"
        else:
            response = "Sorry, I couldn't update your name right now. Please try again."
        self._store_conversation(user_id, message, response)
        return {"operation": "name_changed", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _get_user_name(self, user_id: str) -> str:
        """Get user's name from profile, return None if not found."""
        profile = self.supabase.get_user_profile(user_id)
        if profile.get("ok") and profile.get("data"):
            return profile["data"].get("name")
        return None

    def _get_user_invoice_email(self, user_id: str) -> str:
        """Fetch the user's own email from preferences.invoice_email — used for CCing
        the user on outbound invoice/reminder mail so they have a copy."""
        if not user_id:
            return ""
        try:
            profile = self.supabase.get_user_profile(user_id)
            if not (profile.get("ok") and profile.get("data")):
                return ""
            prefs = profile["data"].get("preferences") or {}
            if isinstance(prefs, str):
                try:
                    prefs = json.loads(prefs)
                except Exception:
                    prefs = {}
            return (prefs.get("invoice_email") or "").strip()
        except Exception as e:
            logger.warning(f"[CC] Failed to fetch invoice_email for {user_id}: {e}")
            return ""

    def _start_onboarding(self, user_id: str, message: str) -> Dict:
        """Start onboarding for a new user."""
        # Determine platform from user_id format
        platform = "telegram" if user_id.isdigit() else "whatsapp"
        
        # Create initial profile
        self.supabase.upsert_user_profile(user_id, platform, {"platform": platform})
        
        response = self._get_welcome_message(platform)
        self._store_conversation(user_id, message, response)
        return {"operation": "onboarding_started", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _continue_onboarding(self, user_id: str, message: str, profile: Dict) -> Dict:
        """Continue onboarding. All three fields (name, email, industry) are required."""
        platform = profile.get("platform", "telegram")

        # Parse existing preferences (stored as JSONB string)
        prefs = profile.get("preferences") or {}
        if isinstance(prefs, str):
            try:
                prefs = json.loads(prefs)
            except Exception:
                prefs = {}

        # ── Step 1: Name (optional — 'skip' accepted) ────────────────────
        if not profile.get("name"):
            raw_name = message.strip()

            # User explicitly skips → assign generic name and advance
            if raw_name.lower() in ("skip", "n/a"):
                name = "User"
                self.supabase.upsert_user_profile(user_id, platform, {"name": name})
                response = (
                    "No problem! I'll refer to you as 'User' for now — you can change it anytime.\n\n"
                    "What's your company or industry? (e.g. Video Production, Photography, Design)\n"
                    "Type 'skip' to skip this too."
                )
                self._store_conversation(user_id, message, response)
                return {"operation": "onboarding_name", "response": response, "trigger_invoice": False, "invoice_data": {}}

            _GREETING_WORDS = {
                "hi", "hello", "helo", "hey", "heyy", "heyyy", "hii", "hiii",
                "yo", "sup", "hola", "howdy", "morning", "evening", "afternoon",
                "good", "whats", "what's", "up", "wassup", "namaste", "namaskar",
                "salaam", "salam", "hai", "haii", "ji", "bhai", "dost", "friend",
                "there", "everyone", "all",
            }
            _stripped = raw_name.lower().strip("!?.,:;'\"() ")
            _tokens = re.findall(r"[a-z']+", _stripped)
            _is_greeting_only = bool(_tokens) and all(t in _GREETING_WORDS for t in _tokens)
            if _stripped in {"good morning", "good evening", "good afternoon",
                             "whats up", "what's up"}:
                _is_greeting_only = True
            if _is_greeting_only or raw_name.lower() in ("no", ""):
                response = "Before we begin, please share your full name — or type 'skip' to continue without one."
                self._store_conversation(user_id, message, response)
                return {"operation": "onboarding_name_retry", "response": response, "trigger_invoice": False, "invoice_data": {}}

            ai_name = self.gemini.extract_name(raw_name)
            if ai_name:
                name = ai_name.title()
            else:
                # Fallback: pattern matching if AI call fails
                name_patterns = [
                    "मेरा नाम ", "मैं ",
                    "mera naam ", "mera name ", "main hoon ",
                    "my name is ", "i'm ", "i am ",
                    "call me ", "this is ", "it's ", "its ",
                ]
                _trailing_filler = [
                    " है", " हैं", " हूँ", " हूं", " हु",
                    " hai", " he", " hoon", " hun", " hu",
                ]
                name = raw_name
                for pattern in name_patterns:
                    if pattern.lower() in raw_name.lower():
                        idx = raw_name.lower().find(pattern.lower())
                        name = raw_name[idx + len(pattern):].strip()
                        break
                for filler in _trailing_filler:
                    if name.lower().endswith(filler.lower()):
                        name = name[: len(name) - len(filler)].strip()
                        break
                if len(name.split()) > 3:
                    name = " ".join(name.split()[:2])
                if name:
                    name = name.title()

            if not name or len(name.strip()) < 2:
                response = "I didn't catch your name. Please type your full name (e.g. 'Akshaj Kasliwal')."
                self._store_conversation(user_id, message, response)
                return {"operation": "onboarding_name_retry", "response": response, "trigger_invoice": False, "invoice_data": {}}

            result = self.supabase.upsert_user_profile(user_id, platform, {"name": name})
            if not result.get("ok"):
                logger.error(f"[ONBOARDING] Failed to save name for {user_id}: {result.get('error')}")

            response = (
                f"Nice to meet you, {name}! 🎉\n\n"
                "What's your email address? (used on invoices and for communication)"
            )
            self._store_conversation(user_id, message, response)
            return {"operation": "onboarding_name", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # ── Step 2: Email (optional — 'skip' accepted) ────────────────────
        elif not prefs.get("invoice_email"):
            raw = message.strip()

            # User skips email → use name as company/industry and complete onboarding
            if raw.lower() in ("skip", "n/a"):
                user_name = profile.get("name", "User")
                prefs["industry"] = user_name
                from datetime import datetime
                self.supabase.upsert_user_profile(user_id, platform, {
                    "preferences": prefs,
                    "onboarded_at": datetime.now().isoformat(),
                })
                response = (
                    f"Got it, {user_name}! You're all set! ✅\n\n"
                    "Here's how to use me:\n\n"
                    "📊 View data:\n"
                    "• 'How many jobs this month?'\n"
                    "• 'Total fees for Client X'\n\n"
                    "📄 Generate invoices:\n"
                    "• 'Send invoice to Client for March'\n\n"
                    "✏️ Add jobs:\n"
                    "• 'Add a job for Client X'\n\n"
                    "💳 Bank details:\n"
                    "• 'Update bank details'\n\n"
                    "Try it now! Say 'Add a job' to get started."
                )
                self._store_conversation(user_id, message, response)
                return {"operation": "onboarding_complete", "response": response, "trigger_invoice": False, "invoice_data": {}}

            # Try to extract an email if it's embedded in a sentence
            _m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", raw)
            candidate = _m.group(0) if _m else raw
            if not self._is_valid_email(candidate):
                response = (
                    "I need a valid email address to continue (e.g. name@company.com). "
                    "This appears on your invoices — or type 'skip' to set it up later."
                )
                self._store_conversation(user_id, message, response)
                return {"operation": "onboarding_email_retry", "response": response, "trigger_invoice": False, "invoice_data": {}}

            prefs["invoice_email"] = candidate
            self.supabase.upsert_user_profile(user_id, platform, {"preferences": prefs})

            response = (
                f"Got it. Saved {candidate} ✅\n\n"
                "Last step — what industry are you in?\n"
                "(e.g. Video Production, Photography, Design, Marketing, Consulting, etc.)"
            )
            self._store_conversation(user_id, message, response)
            return {"operation": "onboarding_email", "response": response, "trigger_invoice": False, "invoice_data": {}}

        # ── Step 3: Industry (optional — 'skip' accepted) ────────────────
        elif not prefs.get("industry"):
            industry = message.strip()
            if industry.lower() in ("skip", "no", "n/a", "") or len(industry) < 2:
                # Use user's name as default industry when skipped
                industry = profile.get("name", "Freelancer")
            if len(industry) > 80:
                industry = industry[:80]
            prefs["industry"] = industry

            from datetime import datetime
            self.supabase.upsert_user_profile(user_id, platform, {
                "preferences": prefs,
                "onboarded_at": datetime.now().isoformat(),
            })
            
            user_name = profile.get("name", "there")
            response = (
                f"Great, {user_name}! You're all set! ✅\n\n"
                "Here's how to use me:\n\n"
                "📊 View data:\n"
                "• 'How many jobs this month?'\n"
                "• 'Total fees for Client X'\n\n"
                "📄 Generate invoices:\n"
                "• 'Send invoice to Client for March'\n\n"
                "✏️ Add jobs:\n"
                "• 'Add a job for Client X'\n\n"
                "💳 Bank details:\n"
                "• 'Update bank details'\n\n"
                "Try it now! Say 'Add a job' to get started."
            )
            self._store_conversation(user_id, message, response)
            return {"operation": "onboarding_complete", "response": response, "trigger_invoice": False, "invoice_data": {}}
        
        else:
            # Shouldn't reach here, but complete onboarding if somehow stuck
            return self._complete_onboarding(user_id, message)

    def _get_welcome_message(self, platform: str) -> str:
        """Get platform-specific welcome message."""
        if platform == "telegram":
            return (
                "👋 Hi, I'm Remyndly!\n\n"
                "I help freelancers and creators stay on top of their business — "
                "log jobs, generate invoices, track payments, and chase clients who owe you money. All from chat.\n\n"
                "Let's set you up. What's your name?"
            )
        else:  # WhatsApp
            return (
                "👋 Hi, I'm Remyndly!\n\n"
                "I help freelancers and creators stay on top of their business — "
                "log jobs, generate invoices, track payments, and chase clients who owe you money. All from chat.\n\n"
                "Let's set you up. What's your name?"
            )

    def _handle_excel_import(self, user_id: str, message: str) -> Dict:
        """Handle Excel file import choice."""
        response = (
            "📎 To import from Excel:\n\n"
            "1. Download the template from: [Your template URL]\n"
            "2. Fill it with your job data\n"
            "3. Send the file here\n\n"
            "Or reply 'back' to choose another option."
        )
        self._store_conversation(user_id, message, response)
        return {"operation": "onboarding_excel", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_csv_import(self, user_id: str, message: str) -> Dict:
        """Handle CSV import choice."""
        response = (
            "📋 Paste your CSV data in this format:\n\n"
            "Client Name,Job Description,Date,Fees,Email\n"
            "Garnier,Short animation,2026-02-20,2000,email@example.com\n\n"
            "Send your data or reply 'back' to choose another option."
        )
        self._store_conversation(user_id, message, response)
        return {"operation": "onboarding_csv", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _handle_manual_entry(self, user_id: str, message: str) -> Dict:
        """Handle manual entry choice."""
        response = (
            "✏️ I'll help you add jobs manually!\n\n"
            "Let's add your first job. What's the client name?\n\n"
            "(Type 'cancel' anytime to stop)"
        )
        self._store_conversation(user_id, message, response)
        return {"operation": "onboarding_manual", "response": response, "trigger_invoice": False, "invoice_data": {}}

    def _complete_onboarding(self, user_id: str, message: str) -> Dict:
        """Complete the onboarding process."""
        # Mark as onboarded
        from datetime import datetime
        self.supabase.upsert_user_profile(user_id, "", {"onboarded_at": datetime.now().isoformat()})
        
        response = (
            "✅ You're all set! Here's how to use me:\n\n"
            "📊 View data:\n"
            "• 'How many jobs for Client X?'\n"
            "• 'Total fees this month'\n"
            "• 'Last payment date'\n\n"
            "📄 Generate invoices:\n"
            "• 'Send invoice to Client for March'\n"
            "• 'Generate invoice for last job'\n\n"
            "💳 Manage bank details:\n"
            "• 'Update bank details'\n"
            "• 'My bank details'\n\n"
            "Try: 'Show my jobs from last week'"
        )
        self._store_conversation(user_id, message, response)
        return {"operation": "onboarding_complete", "response": response, "trigger_invoice": False, "invoice_data": {}}

    @staticmethod
    def get_help_text() -> str:
        return (
            "I'm your conversational assistant! Here's what I can do:\n\n"
            "✏️ Add a job (one message!):\n"
            "Add job\n"
            "Bridgestone\n"
            "10 Feb\n"
            "Master film 30 sec + 4 cutdowns\n"
            "Client: The Good Take\n"
            "Fees: 25k\n\n"
            "Or ultra-fast: + Bridgestone 10 Feb 25k master film\n\n"
            "📄 Invoices: 'Send invoice to Garnier for April'\n"
            "📊 Queries: 'Total fees this month' / 'Jobs for Client X'\n"
            "💳 Bank: 'Update bank details' / 'My bank details'\n\n"
            "How can I help you today?"
        )
